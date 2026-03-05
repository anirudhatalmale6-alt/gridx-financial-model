"""
GRIDx Financial Dashboard v3 — Presentation-focused
Dashboard redesigned to match Revenue Model v3.1 Summary layout:
  LEFT: Color-coded financial summary table with section indicators
  RIGHT: Line chart (top) + Stacked bar chart (bottom)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

wb = Workbook()
SCENARIOS = ["3,000 Meters", "5,000 Meters", "10,000 Meters", "20,000 Meters", "50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]
NS = 5

# ── Shared styles ──
NAVY = "1B2A4A"; DARK_BLUE = "2C3E6B"; TEAL = "2E75B6"
hdr_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor="D6E4F0")
green_fill = PatternFill("solid", fgColor="E2EFDA")
orange_fill = PatternFill("solid", fgColor="FCE4D6")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
bold11 = Font(bold=True, size=11, name="Calibri")
label_font = Font(size=10, name="Calibri")
input_font = Font(color="0000FF", size=10, name="Calibri")
green_font = Font(color="008000", size=10, name="Calibri")
title_font = Font(bold=True, color=NAVY, size=14, name="Calibri")
navy_bold = Font(bold=True, color=NAVY, size=12, name="Calibri")

thin = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
total_bdr = Border(top=Side("medium", color=NAVY), bottom=Side("double", color=NAVY))

NUM = '#,##0'; PCT = '0.0%'; CURR = 'N$#,##0'; CURR_N = 'N$#,##0;(N$#,##0);"-"'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=r, column=cs+i, value=l)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin

def wl(ws, r, c, t, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=t); cell.font = f
    if fi: cell.fill = fi
    cell.border = thin; return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(row=r, column=c, value=v); cell.font = input_font; cell.fill = yellow_fill; cell.number_format = fmt; cell.border = thin; return cell

def wf(ws, r, c, formula, fmt=NUM, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = f
    if fi: cell.fill = fi
    cell.number_format = fmt; cell.border = thin; return cell

def fill_row(ws, r, cs, ce, fi):
    for c in range(cs, ce+1):
        ws.cell(r, c).fill = fi; ws.cell(r, c).border = thin

# ═══════════════════════════════════════════════════════════════
# SHEET 1: PERSONNEL COST (identical to v2)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Personnel Cost"; ws1.sheet_properties.tabColor = NAVY
ws1.column_dimensions['A'].width = 36
for i in range(NS): ws1.column_dimensions[get_column_letter(2+i)].width = 16
ws1.column_dimensions['H'].width = 30; ws1.column_dimensions['I'].width = 18; ws1.column_dimensions['J'].width = 18

wh(ws1, 1, ["Resource Type"] + SCENARIOS)
ws1.cell(1,1).alignment = Alignment(horizontal="left")
wl(ws1, 1, 8, "Position", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 9, "Monthly Salary (N$)", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 10, "Annual Salary (N$)", f=hdr_font, fi=hdr_fill)

departments = [
    ("LEADERSHIP & MANAGEMENT", [
        ("Chief Executive Officer (CEO)", [1,1,1,1,1], 35000),
        ("Chief Operations Officer (COO)", [0,0,1,1,1], 28000),
        ("Chief Financial Officer (CFO)", [0,1,1,1,1], 28000),
        ("Chief Technology Officer (CTO)", [0,1,1,1,1], 30000),
        ("Head of Business Development", [0,0,1,1,1], 25000),
        ("Head of Human Resources", [0,0,0,1,1], 22000),
        ("Legal & Compliance Officer", [0,0,0,1,1], 22000),
    ]),
    ("OPERATIONS & FIELD SERVICES", [
        ("Operations Manager", [0,1,1,1,1], 20000),
        ("Field Support Technicians", [2,3,4,5,8], 9000),
        ("Installation Team Lead", [0,1,1,2,3], 11000),
        ("Logistics Coordinator", [0,0,1,1,2], 8000),
    ]),
    ("ENGINEERING & QUALITY", [
        ("Engineering Manager", [0,0,1,1,1], 22000),
        ("Quality Assurance Engineers", [1,1,1,3,5], 12000),
        ("Firmware/Software Engineers", [0,1,2,3,5], 15000),
        ("Hardware Engineers", [0,0,1,2,3], 15000),
    ]),
    ("CUSTOMER SUCCESS & SUPPORT", [
        ("Customer Support Manager", [0,0,1,1,1], 18000),
        ("Customer Support Agents", [1,2,2,4,8], 7200),
        ("Technical Support Specialists", [0,0,1,2,3], 9000),
    ]),
    ("SALES & MARKETING", [
        ("Sales Manager", [0,0,1,1,1], 20000),
        ("Sales Representatives", [0,1,2,3,5], 10000),
        ("Marketing Coordinator", [0,0,1,1,2], 9000),
    ]),
    ("FINANCE & ADMINISTRATION", [
        ("Finance Manager", [0,0,1,1,1], 20000),
        ("Accountant", [0,1,1,2,3], 11000),
        ("Administrative Assistant", [0,0,1,2,3], 7000),
    ]),
]

r = 2; sal_row = 2
subtotal_rows = []
position_rows = {}

for dept_name, positions in departments:
    wl(ws1, r, 1, dept_name, f=bold11, fi=light_fill)
    fill_row(ws1, r, 2, 6, light_fill)
    r += 1
    start = r
    for name, counts, salary in positions:
        wl(ws1, r, 1, name)
        for i, cnt in enumerate(counts):
            wi(ws1, r, 2+i, cnt, NUM)
        position_rows[name] = r
        wl(ws1, sal_row, 8, name); wi(ws1, sal_row, 9, salary, CURR)
        wf(ws1, sal_row, 10, f"=I{sal_row}*12", CURR)
        sal_row += 1
        r += 1
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, f"Subtotal – {short}", f=bold_font)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws1, r, 2+i, f"=SUM({col}{start}:{col}{r-1})", NUM, bold_font)
    subtotal_rows.append(r)
    r += 2

wl(ws1, r, 1, "TOTAL PERSONNEL", f=navy_bold, fi=green_fill)
fill_row(ws1, r, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{col}{sr}" for sr in subtotal_rows])
    cell = wf(ws1, r, 2+i, f"={refs}", NUM, navy_bold, green_fill)
    cell.border = total_bdr
TOTAL_PERS_ROW = r

r += 3
wl(ws1, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=navy_bold)
r += 1
wh(ws1, r, ["Position", "Unit Cost (N$/year)"] + SCENARIOS)
r += 1

cost_subtotal_rows = []
for dept_name, positions in departments:
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, dept_name, f=bold_font, fi=light_fill)
    fill_row(ws1, r, 2, 2+NS, light_fill)
    r += 1
    dept_start = r
    for name, counts, salary in positions:
        hc_row = position_rows[name]
        annual = salary * 12
        wl(ws1, r, 1, name)
        wi(ws1, r, 2, annual, CURR)
        for i in range(NS):
            hc_col = get_column_letter(2+i)
            wf(ws1, r, 3+i, f"={hc_col}{hc_row}*$B{r}", CURR, green_font)
        r += 1
    wl(ws1, r, 1, f"{short} Subtotal", f=bold_font)
    for i in range(NS):
        col = get_column_letter(3+i)
        wf(ws1, r, 3+i, f"=SUM({col}{dept_start}:{col}{r-1})", CURR, bold_font)
    cost_subtotal_rows.append(r)
    r += 2

wl(ws1, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=navy_bold, fi=green_fill)
fill_row(ws1, r, 2, 2+NS, green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    refs = "+".join([f"{col}{sr}" for sr in cost_subtotal_rows])
    cell = wf(ws1, r, 3+i, f"={refs}", CURR, navy_bold, green_fill)
    cell.border = total_bdr
TOTAL_COST_ROW = r
print(f"Personnel: total headcount row={TOTAL_PERS_ROW}, cost row={TOTAL_COST_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: INFRASTRUCTURE & EQUIPMENT COST (identical to v2)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Infrastructure & Equipment"); ws2.sheet_properties.tabColor = TEAL
ws2.column_dimensions['A'].width = 36; ws2.column_dimensions['B'].width = 18
for i in range(NS): ws2.column_dimensions[get_column_letter(3+i)].width = 18

wl(ws2, 1, 1, "Infrastructure Costs (Annual)", f=title_font)
wh(ws2, 3, ["Cost Category", "Unit Cost"] + SCENARIOS)

infra = [
    ("Data Storage & Transmission", "N$48/meter", [144000,240000,480000,960000,2400000]),
    ("Office Space Rental", "N$200/sq m", [60000,120000,240000,360000,600000]),
    ("Utilities & Connectivity", "Lump sum", [36000,60000,120000,240000,480000]),
]
r = 4
for name, unit, vals in infra:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Annual Infrastructure Total", f=bold_font, fi=green_fill)
wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}4:{col}6)", CURR, bold_font, green_fill)
    cell.border = total_bdr
INFRA_TOT = r

r += 2
wl(ws2, r, 1, "Equipment Costs (One-Time Investment)", f=title_font)
r += 2
wh(ws2, r, ["Equipment Type", "Unit Cost"] + SCENARIOS)
r += 1

equip = [
    ("Service Vehicles", "N$250,000", [250000,500000,750000,750000,1250000]),
    ("Field Test Equipment Kits", "N$85,000", [85000,170000,340000,680000,1530000]),
    ("Spare Parts Inventory", "N$4,495 × 2%", [269700,449500,899000,1798000,4495000]),
    ("Workstations", "N$15,000", [75000,210000,420000,630000,930000]),
    ("Network Monitoring Tools", "N$120,000", [120000,120000,240000,360000,600000]),
    ("Office Equipment & Furniture", "Per employee", [50000,140000,280000,420000,620000]),
    ("Server/Cloud Infrastructure", "Lump sum", [100000,150000,250000,400000,800000]),
]
eq_start = r
for name, unit, vals in equip:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Equipment Investment", f=bold_font, fi=green_fill)
wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}{eq_start}:{col}{r-1})", CURR, bold_font, green_fill)
    cell.border = total_bdr
EQUIP_TOT = r
print(f"Infra: {INFRA_TOT}, Equip: {EQUIP_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: MANUFACTURING COST (identical to v2)
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Manufacturing Cost"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 28; ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 20; ws3.column_dimensions['D'].width = 36

wl(ws3, 1, 1, "Volume Discount Structure", f=title_font)
wh(ws3, 3, ["Volume Tier", "Discount", "Cost per Meter (N$)", "Calculation Basis"])
for i, (name, disc, cost, basis) in enumerate([
    ("Base (0-9,999 meters)", 0, 2000, "N$2,000 per meter"),
    ("10,000-19,999 meters", 0.15, 1700, "15% reduction from base"),
    ("20,000-49,999 meters", 0.30, 1400, "Additional 15% reduction"),
    ("50,000+ meters", 0.45, 1100, "Additional 15% reduction"),
]):
    wl(ws3, 4+i, 1, name); wi(ws3, 4+i, 2, disc, PCT); wi(ws3, 4+i, 3, cost, CURR); wl(ws3, 4+i, 4, basis)

wl(ws3, 10, 1, "Manufacturing Cost by Scenario", f=title_font)
wh(ws3, 12, ["Scenario", "Meters", "Cost per Meter (N$)", "Total Manufacturing Cost (N$)"])
mfg_data = [(3000, 2000), (5000, 2000), (10000, 1700), (20000, 1700), (50000, 1462)]
MFG_S = 13
for i, (m, c) in enumerate(mfg_data):
    r = MFG_S + i
    wl(ws3, r, 1, SCENARIOS[i]); wi(ws3, r, 2, m, NUM); wi(ws3, r, 3, c, CURR)
    wf(ws3, r, 4, f"=B{r}*C{r}", CURR, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 4: REVENUE (identical to v2)
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Revenue"); ws4.sheet_properties.tabColor = "ED7D31"
ws4.column_dimensions['A'].width = 36; ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 22; ws4.column_dimensions['D'].width = 36; ws4.column_dimensions['E'].width = 24

wl(ws4, 1, 1, "Revenue Model Assumptions", f=title_font)
wh(ws4, 3, ["Revenue Stream", "Rate", "Adoption Rate", "Calculation Basis", "Annual Amount per Meter (N$)"])

rev_a = [
    (4, "Meter Hardware Sale", "N$5,500 (one-time)", 1.0, "One-time revenue per installed meter", 5500),
    (5, "Meter Installation Fee", "N$700 (one-time)", 1.0, "One-time fee per new meter installation", 700),
    (6, "Meter Maintenance", "N$500 / year", 0.05, "Annual maintenance fee on installed meters", 25),
    (7, "Wi-Fi Subscription", "N$25 / month", "30% of installed base", "Recurring", 90),
    (8, "SMS Notifications", "N$2.50 / SMS", "50% adoption, 2 SMS/month", "Recurring", 30),
    (9, "Mobile App Subscription", "N$65 / month", "80% of installed base", "Recurring", 624),
    (10, "Real Estate Management Platform", "N$3,200 / block / month", "1 block per 500 meters", "Recurring", 76.80),
]
for r, name, rate, adoption, basis, annual in rev_a:
    wl(ws4, r, 1, name); wl(ws4, r, 2, rate)
    if isinstance(adoption, float): wi(ws4, r, 3, adoption, PCT)
    else: wl(ws4, r, 3, adoption)
    wl(ws4, r, 4, basis); wi(ws4, r, 5, annual, 'N$#,##0.00')

wl(ws4, 11, 1, "Total Year 1 Revenue per Meter", f=bold_font, fi=green_fill)
cell = wf(ws4, 11, 5, "=SUM(E4:E10)", 'N$#,##0.00', bold_font, green_fill)
cell.border = total_bdr

wl(ws4, 14, 1, "Revenue by Year", f=title_font)
wh(ws4, 15, ["Year"] + SCENARIOS)

wl(ws4, 16, 1, "Year 1")
for i in range(NS): wf(ws4, 16, 2+i, f"={MC[i]}*E11", CURR, green_font)

for yr in range(2, 6):
    wl(ws4, 15+yr, 1, f"Year {yr}")
    for i in range(NS): wf(ws4, 15+yr, 2+i, f"={MC[i]}*(E6+E7+E8+E9+E10)", CURR, green_font)

wl(ws4, 21, 1, "5-Year Total", f=bold_font, fi=green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws4, 21, 2+i, f"=SUM({col}16:{col}20)", CURR, bold_font, green_fill)
    cell.border = total_bdr

wl(ws4, 24, 1, "5-Year Revenue Breakdown by Stream", f=title_font)
wh(ws4, 25, ["Revenue Stream"] + SCENARIOS)

bd = [
    (26, "Hardware Sales (One-Time)", "E4"),
    (27, "Installation Fees (One-Time)", "E5"),
    (28, "Maintenance (Recurring)", "E6", 5),
    (29, "Wi-Fi Subscriptions", "E7", 5),
    (30, "SMS Notifications", "E8", 5),
    (31, "Mobile App Subscriptions", "E9", 5),
    (32, "Real Estate Management", "E10", 5),
]
for item in bd:
    r = item[0]; name = item[1]; ref = item[2]; mult = item[3] if len(item) > 3 else 1
    wl(ws4, r, 1, name)
    for i in range(NS):
        wf(ws4, r, 2+i, f"={MC[i]}*{ref}*{mult}", CURR, green_font)

wl(ws4, 33, 1, "5-Year Total", f=bold_font, fi=green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws4, 33, 2+i, f"=SUM({col}26:{col}32)", CURR, bold_font, green_fill)
    cell.border = total_bdr

# ═══════════════════════════════════════════════════════════════
# SHEET 5: PROFITABILITY ANALYSIS (identical to v2)
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Profitability Analysis"); ws5.sheet_properties.tabColor = "C00000"
ws5.column_dimensions['A'].width = 42
for i in range(NS): ws5.column_dimensions[get_column_letter(2+i)].width = 20

wl(ws5, 1, 1, "Updated Revenue & Profitability Analysis (Year 1)", f=title_font)
wh(ws5, 3, ["Metric"] + SCENARIOS)

wl(ws5, 4, 1, "REVENUE (N$)", f=bold11, fi=green_fill); fill_row(ws5, 4, 2, 6, green_fill)
wl(ws5, 5, 1, "Hardware Sales")
for i in range(NS): wf(ws5, 5, 2+i, f"={MC[i]}*Revenue!E4", CURR, green_font)
wl(ws5, 6, 1, "Installation Fees")
for i in range(NS): wf(ws5, 6, 2+i, f"={MC[i]}*Revenue!E5", CURR, green_font)
wl(ws5, 7, 1, "Service Revenues")
for i in range(NS): wf(ws5, 7, 2+i, f"={MC[i]}*(Revenue!E6+Revenue!E7+Revenue!E8+Revenue!E9+Revenue!E10)", CURR, green_font)
wl(ws5, 8, 1, "Total Revenue", f=bold11, fi=green_fill); fill_row(ws5, 8, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 8, 2+i, f"=SUM({col}5:{col}7)", CURR, bold11, green_fill)
    cell.border = total_bdr

wl(ws5, 9, 1, "MANUFACTURING COST (N$)", f=bold11, fi=orange_fill); fill_row(ws5, 9, 2, 6, orange_fill)
wl(ws5, 10, 1, "Cost per Meter")
for i in range(NS): wf(ws5, 10, 2+i, f"='Manufacturing Cost'!C{MFG_S+i}", CURR, green_font)
wl(ws5, 11, 1, "Total Manufacturing Cost")
for i in range(NS): wf(ws5, 11, 2+i, f"='Manufacturing Cost'!D{MFG_S+i}", CURR, green_font)

wl(ws5, 12, 1, "OPERATING COSTS (N$)", f=bold11, fi=orange_fill)
for i in range(NS):
    pc = get_column_letter(3+i)
    ic = get_column_letter(3+i)
    cell = wf(ws5, 12, 2+i, f"='Personnel Cost'!{pc}{TOTAL_COST_ROW}+'Infrastructure & Equipment'!{ic}{INFRA_TOT}", CURR, green_font, orange_fill)

wl(ws5, 13, 1, "EQUIPMENT INVESTMENT (N$)")
for i in range(NS):
    ec = get_column_letter(3+i)
    wf(ws5, 13, 2+i, f"='Infrastructure & Equipment'!{ec}{EQUIP_TOT}", CURR, green_font)

wl(ws5, 14, 1, "PROFITABILITY (N$)", f=bold11, fi=light_fill); fill_row(ws5, 14, 2, 6, light_fill)
wl(ws5, 15, 1, "Gross Profit", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 15, 2+i, f"={col}8-{col}11", CURR, bold_font)
wl(ws5, 16, 1, "Operating Profit (EBITDA)", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 16, 2+i, f"={col}15-{col}12-{col}13", CURR, bold_font)
wl(ws5, 17, 1, "Net Profit", f=navy_bold)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 17, 2+i, f"={col}16", CURR, navy_bold)
    cell.border = total_bdr

wl(ws5, 18, 1, "MARGINS", f=bold11, fi=light_fill); fill_row(ws5, 18, 2, 6, light_fill)
for rr, label, num_row in [(19, "Gross Margin %", 15), (20, "Operating Margin %", 16), (21, "Net Margin %", 17)]:
    wl(ws5, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws5, rr, 2+i, f"=IF({col}8<>0,{col}{num_row}/{col}8,0)", PCT, bold_font)

wl(ws5, 23, 1, "5-YEAR PROJECTIONS", f=title_font)
wh(ws5, 24, ["Metric"] + SCENARIOS)

wl(ws5, 25, 1, "5-Year Total Revenue", f=bold_font)
for i in range(NS):
    rc = get_column_letter(2+i)
    wf(ws5, 25, 2+i, f"=Revenue!{rc}21", CURR, green_font)

wl(ws5, 26, 1, "5-Year Total Costs (Mfg + OpCost×5 + Equip)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 26, 2+i, f"={col}11+{col}12*5+{col}13", CURR)

wl(ws5, 27, 1, "5-Year Net Profit", f=navy_bold)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 27, 2+i, f"={col}25-{col}26", CURR, navy_bold)
    cell.border = total_bdr

wl(ws5, 28, 1, "5-Year ROI %", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 28, 2+i, f"=IF({col}26<>0,{col}27/{col}26,0)", PCT, bold_font)

wl(ws5, 29, 1, "5-Year Net Margin %", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 29, 2+i, f"=IF({col}25<>0,{col}27/{col}25,0)", PCT, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 6: DASHBOARD — Presentation Style (Rev Model v3.1 layout)
# Table on LEFT (cols A-H) + Charts on RIGHT (cols J+)
# Color-coded sections with strip indicators
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Dashboard"); ws6.sheet_properties.tabColor = "002060"

# ── Column widths ──
ws6.column_dimensions['A'].width = 1.5    # Left margin
ws6.column_dimensions['B'].width = 2.5    # Section color strip
ws6.column_dimensions['C'].width = 40     # Metric labels
for i in range(NS):
    ws6.column_dimensions[get_column_letter(4+i)].width = 16  # D-H: scenario values
ws6.column_dimensions['I'].width = 2      # Gap between table and charts

# ── Dashboard-specific styles ──
# Section header fills
TEAL_H  = PatternFill("solid", fgColor="00897B")   # Revenue headers
TEAL_BG = PatternFill("solid", fgColor="E0F2F1")   # Revenue light bg
TEAL_TOT= PatternFill("solid", fgColor="B2DFDB")   # Revenue total bg
ORNG_H  = PatternFill("solid", fgColor="E65100")   # Cost headers
ORNG_BG = PatternFill("solid", fgColor="FFF3E0")   # Cost light bg
ORNG_TOT= PatternFill("solid", fgColor="FFE0B2")   # Cost total bg
GRN_H   = PatternFill("solid", fgColor="2E7D32")   # Profit headers
GRN_BG  = PatternFill("solid", fgColor="E8F5E9")   # Profit light bg
GRN_TOT = PatternFill("solid", fgColor="C8E6C9")   # Profit total bg
BLUE_H  = PatternFill("solid", fgColor="1565C0")   # Margin headers
BLUE_BG = PatternFill("solid", fgColor="E3F2FD")   # Margin light bg
DARK_H  = PatternFill("solid", fgColor="263238")   # 5-year headers
DARK_BG = PatternFill("solid", fgColor="ECEFF1")   # 5-year light bg
DARK_TOT= PatternFill("solid", fgColor="CFD8DC")   # 5-year total bg
GOLD_H  = PatternFill("solid", fgColor="F57F17")   # Assumptions headers
GOLD_BG = PatternFill("solid", fgColor="FFFDE7")   # Assumptions light bg

# Fonts
d_title  = Font(bold=True, size=16, name="Calibri", color="1B2A4A")
d_sub    = Font(italic=True, size=11, name="Calibri", color="546E7A")
d_sec    = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
d_label  = Font(size=10, name="Calibri", color="37474F")
d_val    = Font(size=10, name="Calibri", color="1B5E20")
d_bold   = Font(bold=True, size=10, name="Calibri", color="37474F")
d_total  = Font(bold=True, size=11, name="Calibri", color="1B2A4A")
d_hdr    = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
d_pct    = Font(bold=True, size=10, name="Calibri", color="0D47A1")
d_input  = Font(bold=True, size=10, name="Calibri", color="0000FF")
d_param  = Font(size=10, name="Calibri", color="37474F")

# Border styles
d_thin  = Border(
    left=Side("thin", color="CFD8DC"), right=Side("thin", color="CFD8DC"),
    top=Side("thin", color="CFD8DC"), bottom=Side("thin", color="CFD8DC"))
d_total_b = Border(top=Side("medium", color="263238"), bottom=Side("double", color="263238"))
d_sec_b = Border(bottom=Side("thin", color="90A4AE"))

# ── Helper functions for dashboard ──
def dsec(ws, r, label, hdr_fill, ce=8):
    """Section header with colored bg across cols B-H"""
    for c in range(2, ce+1):
        ws.cell(r, c).fill = hdr_fill
        ws.cell(r, c).border = d_thin
    ws.cell(r, 3, label).font = d_sec
    ws.cell(r, 3).fill = hdr_fill

def drow(ws, r, label, strip_fill, bg_fill, font=d_label, ce=8):
    """Data row: strip color in B, label in C, background fill"""
    ws.cell(r, 2).fill = strip_fill
    ws.cell(r, 3, label).font = font
    ws.cell(r, 3).fill = bg_fill
    ws.cell(r, 3).border = d_thin
    for c in range(4, ce+1):
        ws.cell(r, c).fill = bg_fill
        ws.cell(r, c).border = d_thin

def dtotal(ws, r, label, strip_fill, bg_fill, ce=8):
    """Total row with top+bottom border emphasis"""
    ws.cell(r, 2).fill = strip_fill
    ws.cell(r, 3, label).font = d_total
    ws.cell(r, 3).fill = bg_fill
    ws.cell(r, 3).border = d_total_b
    for c in range(4, ce+1):
        ws.cell(r, c).fill = bg_fill
        ws.cell(r, c).border = d_total_b

# ═══════════ BUILD THE DASHBOARD TABLE ═══════════

# ── Row 1-2: Title & subtitle ──
ws6.merge_cells('C1:H1')
ws6['C1'] = "GRIDx Financial Model — Summary Dashboard"
ws6['C1'].font = d_title
ws6['C1'].alignment = Alignment(vertical="center")
ws6.row_dimensions[1].height = 30

ws6.merge_cells('C2:H2')
ws6['C2'] = "Pulsar Electronic Solutions — Smart Meter Deployment (Namibia, N$)"
ws6['C2'].font = d_sub
ws6.row_dimensions[2].height = 20

# ── Row 4: Column headers ──
hdr_bg = PatternFill("solid", fgColor="37474F")
ws6.cell(4, 2).fill = hdr_bg
ws6.cell(4, 3, "Metric").font = d_hdr
ws6.cell(4, 3).fill = hdr_bg; ws6.cell(4, 3).border = d_thin
for i, s in enumerate(SCENARIOS):
    c = ws6.cell(4, 4+i, s)
    c.font = d_hdr; c.fill = hdr_bg
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = d_thin
ws6.row_dimensions[4].height = 28

# ═══════════ REVENUE SECTION (teal) — rows 5-13 ═══════════
dsec(ws6, 5, "REVENUE (N$)", TEAL_H)

# Individual revenue streams
rev_items = [
    (6,  "Hardware Sales (N$5,500/meter)",     "Revenue!E4",  1),
    (7,  "Installation Fees (N$700/meter)",    "Revenue!E5",  1),
    (8,  "Meter Maintenance",                  "Revenue!E6",  1),
    (9,  "Wi-Fi Subscriptions",                "Revenue!E7",  1),
    (10, "SMS Notifications",                  "Revenue!E8",  1),
    (11, "Mobile App Subscriptions",           "Revenue!E9",  1),
    (12, "Real Estate Management Platform",    "Revenue!E10", 1),
]
for dr, label, ref, mult in rev_items:
    drow(ws6, dr, label, TEAL_H, TEAL_BG)
    for i in range(NS):
        c = ws6.cell(dr, 4+i)
        c.value = f"={MC[i]}*{ref}*{mult}"
        c.font = d_val; c.number_format = CURR

# Total Revenue
dtotal(ws6, 13, "Total Revenue", TEAL_H, TEAL_TOT)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(13, 4+i)
    c.value = f"=SUM({col}6:{col}12)"
    c.font = d_total; c.number_format = CURR

# ═══════════ MANUFACTURING COST (orange) — rows 15-17 ═══════════
dsec(ws6, 15, "MANUFACTURING COST (N$)", ORNG_H)

drow(ws6, 16, "Cost per Meter", ORNG_H, ORNG_BG)
for i in range(NS):
    c = ws6.cell(16, 4+i)
    c.value = f"='Manufacturing Cost'!C{MFG_S+i}"
    c.font = d_val; c.number_format = CURR

dtotal(ws6, 17, "Total Manufacturing Cost", ORNG_H, ORNG_TOT)
for i in range(NS):
    c = ws6.cell(17, 4+i)
    c.value = f"='Manufacturing Cost'!D{MFG_S+i}"
    c.font = d_total; c.number_format = CURR

# ═══════════ OPERATING COSTS (orange) — rows 19-22 ═══════════
dsec(ws6, 19, "OPERATING COSTS (N$)", ORNG_H)

drow(ws6, 20, "Annual Personnel Cost", ORNG_H, ORNG_BG)
for i in range(NS):
    pc = get_column_letter(3+i)
    c = ws6.cell(20, 4+i)
    c.value = f"='Personnel Cost'!{pc}{TOTAL_COST_ROW}"
    c.font = d_val; c.number_format = CURR

drow(ws6, 21, "Annual Infrastructure Cost", ORNG_H, ORNG_BG)
for i in range(NS):
    ic = get_column_letter(3+i)
    c = ws6.cell(21, 4+i)
    c.value = f"='Infrastructure & Equipment'!{ic}{INFRA_TOT}"
    c.font = d_val; c.number_format = CURR

dtotal(ws6, 22, "Total Operating Costs", ORNG_H, ORNG_TOT)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(22, 4+i)
    c.value = f"={col}20+{col}21"
    c.font = d_total; c.number_format = CURR

# ═══════════ EQUIPMENT INVESTMENT (orange) — rows 24-25 ═══════════
dsec(ws6, 24, "EQUIPMENT INVESTMENT — One-Time (N$)", ORNG_H)

dtotal(ws6, 25, "Total Equipment Cost", ORNG_H, ORNG_TOT)
for i in range(NS):
    ec = get_column_letter(3+i)
    c = ws6.cell(25, 4+i)
    c.value = f"='Infrastructure & Equipment'!{ec}{EQUIP_TOT}"
    c.font = d_total; c.number_format = CURR

# ═══════════ PROFITABILITY — YEAR 1 (green) — rows 27-31 ═══════════
dsec(ws6, 27, "PROFITABILITY — YEAR 1 (N$)", GRN_H)

drow(ws6, 28, "Gross Profit (Revenue − Manufacturing)", GRN_H, GRN_BG, d_bold)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(28, 4+i)
    c.value = f"={col}13-{col}17"
    c.font = d_val; c.number_format = CURR

drow(ws6, 29, "Operating Profit (EBITDA)", GRN_H, GRN_BG, d_bold)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(29, 4+i)
    c.value = f"={col}28-{col}22-{col}25"
    c.font = d_val; c.number_format = CURR

dtotal(ws6, 30, "Net Profit (Year 1)", GRN_H, GRN_TOT)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(30, 4+i)
    c.value = f"={col}29"
    c.font = d_total; c.number_format = CURR

# Net Profit as % indicator
drow(ws6, 31, "Net Margin %", GRN_H, GRN_BG, d_pct)
for i in range(NS):
    col = get_column_letter(4+i)
    c = ws6.cell(31, 4+i)
    c.value = f"=IF({col}13<>0,{col}30/{col}13,0)"
    c.font = d_pct; c.number_format = PCT

# ═══════════ MARGINS (blue) — rows 33-36 ═══════════
dsec(ws6, 33, "MARGINS (%)", BLUE_H)

for dr, label, num_r in [(34, "Gross Margin", 28), (35, "Operating Margin", 29), (36, "Net Margin", 30)]:
    drow(ws6, dr, label, BLUE_H, BLUE_BG, d_bold if "Net" in label else d_label)
    for i in range(NS):
        col = get_column_letter(4+i)
        c = ws6.cell(dr, 4+i)
        c.value = f"=IF({col}13<>0,{col}{num_r}/{col}13,0)"
        c.font = d_pct; c.number_format = PCT

# ═══════════ 5-YEAR PROJECTIONS (dark) — rows 38-43 ═══════════
dsec(ws6, 38, "5-YEAR PROJECTIONS", DARK_H)

drow(ws6, 39, "5-Year Total Revenue", DARK_H, DARK_BG, d_bold)
for i in range(NS):
    rc = get_column_letter(2+i)
    c = ws6.cell(39, 4+i)
    c.value = f"='Profitability Analysis'!{rc}25"
    c.font = d_val; c.number_format = CURR

drow(ws6, 40, "5-Year Total Costs", DARK_H, DARK_BG)
for i in range(NS):
    rc = get_column_letter(2+i)
    c = ws6.cell(40, 4+i)
    c.value = f"='Profitability Analysis'!{rc}26"
    c.font = d_val; c.number_format = CURR

dtotal(ws6, 41, "5-Year Net Profit", DARK_H, DARK_TOT)
for i in range(NS):
    rc = get_column_letter(2+i)
    c = ws6.cell(41, 4+i)
    c.value = f"='Profitability Analysis'!{rc}27"
    c.font = d_total; c.number_format = CURR

drow(ws6, 42, "5-Year ROI %", DARK_H, DARK_BG, d_bold)
for i in range(NS):
    rc = get_column_letter(2+i)
    c = ws6.cell(42, 4+i)
    c.value = f"='Profitability Analysis'!{rc}28"
    c.font = d_pct; c.number_format = PCT

drow(ws6, 43, "5-Year Net Margin %", DARK_H, DARK_BG, d_bold)
for i in range(NS):
    rc = get_column_letter(2+i)
    c = ws6.cell(43, 4+i)
    c.value = f"='Profitability Analysis'!{rc}29"
    c.font = d_pct; c.number_format = PCT

# ═══════════ KEY ASSUMPTIONS (gold) — rows 45-50 ═══════════
dsec(ws6, 45, "KEY ASSUMPTIONS", GOLD_H)

assumptions = [
    (46, "Sale Price per Meter",           "N$5,500"),
    (47, "Total Revenue per Meter (Yr 1)", "N$7,045.80"),
    (48, "Recurring Revenue per Meter/Yr", "N$845.80"),
    (49, "Currency",                        "Namibian Dollar (N$)"),
    (50, "Projection Horizon",             "5 Years"),
    (51, "Total Personnel (3K–50K)",       "5 – 62 staff"),
]
for dr, label, value in assumptions:
    ws6.cell(dr, 2).fill = GOLD_H
    ws6.cell(dr, 3, label).font = d_param; ws6.cell(dr, 3).fill = GOLD_BG; ws6.cell(dr, 3).border = d_thin
    ws6.merge_cells(f'D{dr}:H{dr}')
    ws6.cell(dr, 4, value).font = d_input; ws6.cell(dr, 4).fill = GOLD_BG; ws6.cell(dr, 4).border = d_thin

# ═══════════ CHARTS — positioned to the RIGHT of the table ═══════════
# Category labels (scenario names) from row 4
cats = Reference(ws6, min_col=4, min_row=4, max_col=8, max_row=4)

# ── CHART 1: Revenue & Profit Growth (Line Chart) — TOP RIGHT ──
ch1 = LineChart()
ch1.title = "Revenue & Net Profit by Deployment Scale"
ch1.style = 10
ch1.y_axis.title = "Amount (N$)"
ch1.y_axis.numFmt = '#,##0'
ch1.x_axis.title = "Number of Meters"
ch1.width = 28; ch1.height = 16
ch1.y_axis.delete = False
ch1.x_axis.delete = False

# Total Revenue line
d1 = Reference(ws6, min_col=4, min_row=13, max_col=8, max_row=13)
ch1.add_data(d1, from_rows=True)
ch1.series[0].title = SeriesLabel(v="Total Revenue")
ch1.series[0].graphicalProperties.line.solidFill = "00897B"  # Teal
ch1.series[0].graphicalProperties.line.width = 32000

# Gross Profit line
d2 = Reference(ws6, min_col=4, min_row=28, max_col=8, max_row=28)
ch1.add_data(d2, from_rows=True)
ch1.series[1].title = SeriesLabel(v="Gross Profit")
ch1.series[1].graphicalProperties.line.solidFill = "4CAF50"  # Green
ch1.series[1].graphicalProperties.line.width = 28000

# Net Profit line
d3 = Reference(ws6, min_col=4, min_row=30, max_col=8, max_row=30)
ch1.add_data(d3, from_rows=True)
ch1.series[2].title = SeriesLabel(v="Net Profit")
ch1.series[2].graphicalProperties.line.solidFill = "1B5E20"  # Dark green
ch1.series[2].graphicalProperties.line.width = 28000
ch1.series[2].graphicalProperties.line.dashStyle = "dash"

ch1.set_categories(cats)
ch1.legend.position = 'b'
ws6.add_chart(ch1, "J3")

# ── CHART 2: Revenue & Cost Breakdown (Stacked Bar) — BOTTOM RIGHT ──
ch2 = BarChart()
ch2.type = "col"; ch2.grouping = "stacked"
ch2.title = "Revenue & Cost Breakdown by Scenario"
ch2.style = 10
ch2.y_axis.title = "Amount (N$)"
ch2.y_axis.numFmt = '#,##0'
ch2.width = 28; ch2.height = 18

# Revenue components (teal/green shades)
rev_series = [
    (6,  "Hardware Sales",   "004D40"),
    (7,  "Installation",     "00695C"),
    (8,  "Maintenance",      "00897B"),
    (9,  "Wi-Fi",            "26A69A"),
    (10, "SMS",              "4DB6AC"),
    (11, "Mobile App",       "80CBC4"),
    (12, "Real Estate Mgmt", "B2DFDB"),
]
for rr, nm, color in rev_series:
    d = Reference(ws6, min_col=4, min_row=rr, max_col=8, max_row=rr)
    ch2.add_data(d, from_rows=True)
    ch2.series[-1].title = SeriesLabel(v=nm)
    ch2.series[-1].graphicalProperties.solidFill = color

# Cost components (orange/red shades)
cost_series = [
    (17, "Manufacturing",  "BF360C"),
    (20, "Personnel",      "E65100"),
    (21, "Infrastructure", "FF8A65"),
    (25, "Equipment",      "FFAB91"),
]
for rr, nm, color in cost_series:
    d = Reference(ws6, min_col=4, min_row=rr, max_col=8, max_row=rr)
    ch2.add_data(d, from_rows=True)
    ch2.series[-1].title = SeriesLabel(v=nm)
    ch2.series[-1].graphicalProperties.solidFill = color

ch2.set_categories(cats)
ch2.legend.position = 'b'
ws6.add_chart(ch2, "J24")

# ── CHART 3: Margin Trends (Line) — below the stacked bar ──
ch3 = LineChart()
ch3.title = "Margin Trends by Scale (%)"
ch3.style = 10
ch3.y_axis.title = "Margin %"
ch3.y_axis.numFmt = '0%'
ch3.width = 28; ch3.height = 14

for rr, nm, color in [(34, "Gross Margin", "1565C0"), (35, "Operating Margin", "42A5F5"), (36, "Net Margin", "0D47A1")]:
    d = Reference(ws6, min_col=4, min_row=rr, max_col=8, max_row=rr)
    ch3.add_data(d, from_rows=True)
    ch3.series[-1].title = SeriesLabel(v=nm)
    ch3.series[-1].graphicalProperties.line.solidFill = color
    ch3.series[-1].graphicalProperties.line.width = 28000

ch3.set_categories(cats)
ch3.legend.position = 'b'
ws6.add_chart(ch3, "J45")

# ═══════════ FINALIZE ═══════════
# Move Dashboard to first position
wb.move_sheet("Dashboard", offset=-5)

# Freeze panes for easy scrolling
ws1.freeze_panes = "B2"
ws5.freeze_panes = "B4"
ws6.freeze_panes = "C5"

# Print settings — Dashboard fits on one landscape page
ws6.sheet_properties.pageSetUpPr = None
ws6.page_setup.orientation = "landscape"
ws6.page_setup.fitToWidth = 1
ws6.page_setup.fitToHeight = 1

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v3.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print("Dashboard: presentation-style layout with table LEFT + charts RIGHT")
print("Sheets: Dashboard, Personnel Cost, Infrastructure & Equipment, Manufacturing Cost, Revenue, Profitability Analysis")
