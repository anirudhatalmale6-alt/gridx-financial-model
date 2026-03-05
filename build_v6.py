"""
GRIDx Financial Dashboard v6 — Polished Visual Dashboard
Key improvements over v5:
  - Clear tile borders around each panel section
  - Better card styling with proper spacing
  - Consistent row heights for grid alignment
  - Chart tiles with bordered frames
  - Cleaner parameter cards with visual hierarchy
  - Bottom-right: Margin legend + Payback calendar matching reference
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.series import SeriesLabel, DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
SCENARIOS = ["3,000 Meters", "5,000 Meters", "10,000 Meters", "20,000 Meters", "50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]
NS = 5

# ── Shared styles ──
NAVY = "1B2A4A"; DARK_BLUE = "2C3E6B"; TEAL = "2E75B6"
hdr_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor="D6E4F0")
green_fill = PatternFill("solid", fgColor="E2EFDA")
orange_fill = PatternFill("solid", fgColor="FCE4D6")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
bold11 = Font(bold=True, size=11, name="Calibri")
label_font = Font(size=10, name="Calibri")
input_font = Font(color="0000FF", size=10, name="Calibri")
green_font = Font(color="008000", size=10, name="Calibri")
title_font = Font(bold=True, color=NAVY, size=14, name="Calibri")
navy_bold = Font(bold=True, color=NAVY, size=12, name="Calibri")

thin = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
total_bdr = Border(top=Side("medium", color=NAVY), bottom=Side("double", color=NAVY))

NUM = '#,##0'; PCT = '0.0%'; CURR = 'N$#,##0'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=r, column=cs+i, value=l)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin

def wl(ws, r, c, t, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=t); cell.font = f
    if fi: cell.fill = fi
    cell.border = thin; return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(row=r, column=c, value=v); cell.font = input_font; cell.fill = yellow_fill; cell.number_format = fmt; cell.border = thin; return cell

def wf(ws, r, c, formula, fmt=NUM, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = f
    if fi: cell.fill = fi
    cell.number_format = fmt; cell.border = thin; return cell

def fill_row(ws, r, cs, ce, fi):
    for c in range(cs, ce+1):
        ws.cell(r, c).fill = fi; ws.cell(r, c).border = thin

# ═══════════════════════════════════════════════════════════════
# SHEET 1: PERSONNEL COST
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Personnel Cost"; ws1.sheet_properties.tabColor = NAVY
ws1.column_dimensions['A'].width = 36
for i in range(NS): ws1.column_dimensions[get_column_letter(2+i)].width = 16
ws1.column_dimensions['H'].width = 30; ws1.column_dimensions['I'].width = 18; ws1.column_dimensions['J'].width = 18

wh(ws1, 1, ["Resource Type"] + SCENARIOS)
ws1.cell(1,1).alignment = Alignment(horizontal="left")
wl(ws1, 1, 8, "Position", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 9, "Monthly Salary (N$)", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 10, "Annual Salary (N$)", f=hdr_font, fi=hdr_fill)

departments = [
    ("LEADERSHIP & MANAGEMENT", [
        ("Chief Executive Officer (CEO)", [1,1,1,1,1], 35000),
        ("Chief Operations Officer (COO)", [0,0,1,1,1], 28000),
        ("Chief Financial Officer (CFO)", [0,1,1,1,1], 28000),
        ("Chief Technology Officer (CTO)", [0,1,1,1,1], 30000),
        ("Head of Business Development", [0,0,1,1,1], 25000),
        ("Head of Human Resources", [0,0,0,1,1], 22000),
        ("Legal & Compliance Officer", [0,0,0,1,1], 22000),
    ]),
    ("OPERATIONS & FIELD SERVICES", [
        ("Operations Manager", [0,1,1,1,1], 20000),
        ("Field Support Technicians", [2,3,4,5,8], 9000),
        ("Installation Team Lead", [0,1,1,2,3], 11000),
        ("Logistics Coordinator", [0,0,1,1,2], 8000),
    ]),
    ("ENGINEERING & QUALITY", [
        ("Engineering Manager", [0,0,1,1,1], 22000),
        ("Quality Assurance Engineers", [1,1,1,3,5], 12000),
        ("Firmware/Software Engineers", [0,1,2,3,5], 15000),
        ("Hardware Engineers", [0,0,1,2,3], 15000),
    ]),
    ("CUSTOMER SUCCESS & SUPPORT", [
        ("Customer Support Manager", [0,0,1,1,1], 18000),
        ("Customer Support Agents", [1,2,2,4,8], 7200),
        ("Technical Support Specialists", [0,0,1,2,3], 9000),
    ]),
    ("SALES & MARKETING", [
        ("Sales Manager", [0,0,1,1,1], 20000),
        ("Sales Representatives", [0,1,2,3,5], 10000),
        ("Marketing Coordinator", [0,0,1,1,2], 9000),
    ]),
    ("FINANCE & ADMINISTRATION", [
        ("Finance Manager", [0,0,1,1,1], 20000),
        ("Accountant", [0,1,1,2,3], 11000),
        ("Administrative Assistant", [0,0,1,2,3], 7000),
    ]),
]

r = 2; sal_row = 2
subtotal_rows = []; position_rows = {}
for dept_name, positions in departments:
    wl(ws1, r, 1, dept_name, f=bold11, fi=light_fill); fill_row(ws1, r, 2, 6, light_fill); r += 1
    start = r
    for name, counts, salary in positions:
        wl(ws1, r, 1, name)
        for i, cnt in enumerate(counts): wi(ws1, r, 2+i, cnt, NUM)
        position_rows[name] = r
        wl(ws1, sal_row, 8, name); wi(ws1, sal_row, 9, salary, CURR); wf(ws1, sal_row, 10, f"=I{sal_row}*12", CURR)
        sal_row += 1; r += 1
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, f"Subtotal – {short}", f=bold_font)
    for i in range(NS):
        col = get_column_letter(2+i)
        ws1.cell(r, 2+i).value = f"=SUM({col}{start}:{col}{r-1})"
        ws1.cell(r, 2+i).font = bold_font
    subtotal_rows.append(r); r += 2

wl(ws1, r, 1, "TOTAL PERSONNEL", f=navy_bold, fi=green_fill); fill_row(ws1, r, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{col}{sr}" for sr in subtotal_rows])
    cell = wf(ws1, r, 2+i, f"={refs}", NUM, navy_bold, green_fill); cell.border = total_bdr
TOTAL_PERS_ROW = r

r += 3; wl(ws1, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=navy_bold); r += 1
wh(ws1, r, ["Position", "Unit Cost (N$/year)"] + SCENARIOS); r += 1
cost_subtotal_rows = []
for dept_name, positions in departments:
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, dept_name, f=bold_font, fi=light_fill); fill_row(ws1, r, 2, 2+NS, light_fill); r += 1
    dept_start = r
    for name, counts, salary in positions:
        hc_row = position_rows[name]; annual = salary * 12
        wl(ws1, r, 1, name); wi(ws1, r, 2, annual, CURR)
        for i in range(NS): wf(ws1, r, 3+i, f"={get_column_letter(2+i)}{hc_row}*$B{r}", CURR, green_font)
        r += 1
    wl(ws1, r, 1, f"{short} Subtotal", f=bold_font)
    for i in range(NS): wf(ws1, r, 3+i, f"=SUM({get_column_letter(3+i)}{dept_start}:{get_column_letter(3+i)}{r-1})", CURR, bold_font)
    cost_subtotal_rows.append(r); r += 2

wl(ws1, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=navy_bold, fi=green_fill); fill_row(ws1, r, 2, 2+NS, green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    refs = "+".join([f"{col}{sr}" for sr in cost_subtotal_rows])
    cell = wf(ws1, r, 3+i, f"={refs}", CURR, navy_bold, green_fill); cell.border = total_bdr
TOTAL_COST_ROW = r
print(f"Personnel: headcount={TOTAL_PERS_ROW}, cost={TOTAL_COST_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: INFRASTRUCTURE & EQUIPMENT
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Infrastructure & Equipment"); ws2.sheet_properties.tabColor = TEAL
ws2.column_dimensions['A'].width = 36; ws2.column_dimensions['B'].width = 18
for i in range(NS): ws2.column_dimensions[get_column_letter(3+i)].width = 18
wl(ws2, 1, 1, "Infrastructure Costs (Annual)", f=title_font)
wh(ws2, 3, ["Cost Category", "Unit Cost"] + SCENARIOS)
infra = [("Data Storage & Transmission","N$48/meter",[144000,240000,480000,960000,2400000]),
    ("Office Space Rental","N$200/sq m",[60000,120000,240000,360000,600000]),
    ("Utilities & Connectivity","Lump sum",[36000,60000,120000,240000,480000])]
r = 4
for name, unit, vals in infra:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Annual Infrastructure Total", f=bold_font, fi=green_fill); wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}4:{col}6)", CURR, bold_font, green_fill); cell.border = total_bdr
INFRA_TOT = r
r += 2; wl(ws2, r, 1, "Equipment Costs (One-Time Investment)", f=title_font); r += 2
wh(ws2, r, ["Equipment Type", "Unit Cost"] + SCENARIOS); r += 1
equip = [("Service Vehicles","N$250,000",[250000,500000,750000,750000,1250000]),
    ("Field Test Equipment Kits","N$85,000",[85000,170000,340000,680000,1530000]),
    ("Spare Parts Inventory","N$4,495 × 2%",[269700,449500,899000,1798000,4495000]),
    ("Workstations","N$15,000",[75000,210000,420000,630000,930000]),
    ("Network Monitoring Tools","N$120,000",[120000,120000,240000,360000,600000]),
    ("Office Equipment & Furniture","Per employee",[50000,140000,280000,420000,620000]),
    ("Server/Cloud Infrastructure","Lump sum",[100000,150000,250000,400000,800000])]
eq_start = r
for name, unit, vals in equip:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Equipment Investment", f=bold_font, fi=green_fill); wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}{eq_start}:{col}{r-1})", CURR, bold_font, green_fill); cell.border = total_bdr
EQUIP_TOT = r
print(f"Infra: {INFRA_TOT}, Equip: {EQUIP_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: MANUFACTURING COST
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Manufacturing Cost"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 28; ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 20; ws3.column_dimensions['D'].width = 36
wl(ws3, 1, 1, "Volume Discount Structure", f=title_font)
wh(ws3, 3, ["Volume Tier", "Discount", "Cost per Meter (N$)", "Calculation Basis"])
for i, (name, disc, cost, basis) in enumerate([
    ("Base (0-9,999 meters)",0,2000,"N$2,000 per meter"),("10,000-19,999 meters",0.15,1700,"15% reduction"),
    ("20,000-49,999 meters",0.30,1400,"Additional 15%"),("50,000+ meters",0.45,1100,"Additional 15%")]):
    wl(ws3, 4+i, 1, name); wi(ws3, 4+i, 2, disc, PCT); wi(ws3, 4+i, 3, cost, CURR); wl(ws3, 4+i, 4, basis)
wl(ws3, 10, 1, "Manufacturing Cost by Scenario", f=title_font)
wh(ws3, 12, ["Scenario", "Meters", "Cost per Meter (N$)", "Total Manufacturing Cost (N$)"])
mfg_data = [(3000,2000),(5000,2000),(10000,1700),(20000,1700),(50000,1462)]
MFG_S = 13
for i, (m, c) in enumerate(mfg_data):
    r = MFG_S + i
    wl(ws3, r, 1, SCENARIOS[i]); wi(ws3, r, 2, m, NUM); wi(ws3, r, 3, c, CURR); wf(ws3, r, 4, f"=B{r}*C{r}", CURR, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 4: REVENUE
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Revenue"); ws4.sheet_properties.tabColor = "ED7D31"
ws4.column_dimensions['A'].width = 36; ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 22; ws4.column_dimensions['D'].width = 36; ws4.column_dimensions['E'].width = 24
wl(ws4, 1, 1, "Revenue Model Assumptions", f=title_font)
wh(ws4, 3, ["Revenue Stream","Rate","Adoption Rate","Calculation Basis","Annual Amount per Meter (N$)"])
rev_a = [(4,"Meter Hardware Sale","N$5,500 (one-time)",1.0,"One-time revenue per installed meter",5500),
    (5,"Meter Installation Fee","N$700 (one-time)",1.0,"One-time fee per new meter installation",700),
    (6,"Meter Maintenance","N$500 / year",0.05,"Annual maintenance fee on installed meters",25),
    (7,"Wi-Fi Subscription","N$25 / month","30% of installed base","Recurring",90),
    (8,"SMS Notifications","N$2.50 / SMS","50% adoption, 2 SMS/month","Recurring",30),
    (9,"Mobile App Subscription","N$65 / month","80% of installed base","Recurring",624),
    (10,"Real Estate Management Platform","N$3,200 / block / month","1 block per 500 meters","Recurring",76.80)]
for r, name, rate, adoption, basis, annual in rev_a:
    wl(ws4, r, 1, name); wl(ws4, r, 2, rate)
    if isinstance(adoption, float): wi(ws4, r, 3, adoption, PCT)
    else: wl(ws4, r, 3, adoption)
    wl(ws4, r, 4, basis); wi(ws4, r, 5, annual, 'N$#,##0.00')
wl(ws4, 11, 1, "Total Year 1 Revenue per Meter", f=bold_font, fi=green_fill)
cell = wf(ws4, 11, 5, "=SUM(E4:E10)", 'N$#,##0.00', bold_font, green_fill); cell.border = total_bdr
wl(ws4, 14, 1, "Revenue by Year", f=title_font); wh(ws4, 15, ["Year"] + SCENARIOS)
wl(ws4, 16, 1, "Year 1")
for i in range(NS): wf(ws4, 16, 2+i, f"={MC[i]}*E11", CURR, green_font)
for yr in range(2, 6):
    wl(ws4, 15+yr, 1, f"Year {yr}")
    for i in range(NS): wf(ws4, 15+yr, 2+i, f"={MC[i]}*(E6+E7+E8+E9+E10)", CURR, green_font)
wl(ws4, 21, 1, "5-Year Total", f=bold_font, fi=green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws4, 21, 2+i, f"=SUM({col}16:{col}20)", CURR, bold_font, green_fill); cell.border = total_bdr

# ═══════════════════════════════════════════════════════════════
# SHEET 5: PROFITABILITY ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Profitability Analysis"); ws5.sheet_properties.tabColor = "C00000"
ws5.column_dimensions['A'].width = 42
for i in range(NS): ws5.column_dimensions[get_column_letter(2+i)].width = 20
wl(ws5, 1, 1, "Revenue & Profitability Analysis (Year 1)", f=title_font)
wh(ws5, 3, ["Metric"] + SCENARIOS)
wl(ws5, 4, 1, "REVENUE (N$)", f=bold11, fi=green_fill); fill_row(ws5, 4, 2, 6, green_fill)
wl(ws5, 5, 1, "Hardware Sales")
for i in range(NS): wf(ws5, 5, 2+i, f"={MC[i]}*Revenue!E4", CURR, green_font)
wl(ws5, 6, 1, "Installation Fees")
for i in range(NS): wf(ws5, 6, 2+i, f"={MC[i]}*Revenue!E5", CURR, green_font)
wl(ws5, 7, 1, "Service Revenues")
for i in range(NS): wf(ws5, 7, 2+i, f"={MC[i]}*(Revenue!E6+Revenue!E7+Revenue!E8+Revenue!E9+Revenue!E10)", CURR, green_font)
wl(ws5, 8, 1, "Total Revenue", f=bold11, fi=green_fill); fill_row(ws5, 8, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 8, 2+i, f"=SUM({col}5:{col}7)", CURR, bold11, green_fill); cell.border = total_bdr
wl(ws5, 9, 1, "MANUFACTURING COST (N$)", f=bold11, fi=orange_fill); fill_row(ws5, 9, 2, 6, orange_fill)
wl(ws5, 10, 1, "Cost per Meter")
for i in range(NS): wf(ws5, 10, 2+i, f"='Manufacturing Cost'!C{MFG_S+i}", CURR, green_font)
wl(ws5, 11, 1, "Total Manufacturing Cost")
for i in range(NS): wf(ws5, 11, 2+i, f"='Manufacturing Cost'!D{MFG_S+i}", CURR, green_font)
wl(ws5, 12, 1, "OPERATING COSTS (N$)", f=bold11, fi=orange_fill)
for i in range(NS):
    pc = get_column_letter(3+i)
    cell = wf(ws5, 12, 2+i, f"='Personnel Cost'!{pc}{TOTAL_COST_ROW}+'Infrastructure & Equipment'!{pc}{INFRA_TOT}", CURR, green_font, orange_fill)
wl(ws5, 13, 1, "EQUIPMENT INVESTMENT (N$)")
for i in range(NS): wf(ws5, 13, 2+i, f"='Infrastructure & Equipment'!{get_column_letter(3+i)}{EQUIP_TOT}", CURR, green_font)
wl(ws5, 14, 1, "PROFITABILITY (N$)", f=bold11, fi=light_fill); fill_row(ws5, 14, 2, 6, light_fill)
wl(ws5, 15, 1, "Gross Profit", f=bold_font)
for i in range(NS): wf(ws5, 15, 2+i, f"={get_column_letter(2+i)}8-{get_column_letter(2+i)}11", CURR, bold_font)
wl(ws5, 16, 1, "Operating Profit (EBITDA)", f=bold_font)
for i in range(NS): wf(ws5, 16, 2+i, f"={get_column_letter(2+i)}15-{get_column_letter(2+i)}12-{get_column_letter(2+i)}13", CURR, bold_font)
wl(ws5, 17, 1, "Net Profit", f=navy_bold)
for i in range(NS):
    cell = wf(ws5, 17, 2+i, f"={get_column_letter(2+i)}16", CURR, navy_bold); cell.border = total_bdr
wl(ws5, 18, 1, "MARGINS", f=bold11, fi=light_fill); fill_row(ws5, 18, 2, 6, light_fill)
for rr, label, num_row in [(19,"Gross Margin %",15),(20,"Operating Margin %",16),(21,"Net Margin %",17)]:
    wl(ws5, rr, 1, label)
    for i in range(NS): wf(ws5, rr, 2+i, f"=IF({get_column_letter(2+i)}8<>0,{get_column_letter(2+i)}{num_row}/{get_column_letter(2+i)}8,0)", PCT, bold_font)
wl(ws5, 23, 1, "5-YEAR PROJECTIONS", f=title_font); wh(ws5, 24, ["Metric"] + SCENARIOS)
wl(ws5, 25, 1, "5-Year Total Revenue", f=bold_font)
for i in range(NS): wf(ws5, 25, 2+i, f"=Revenue!{get_column_letter(2+i)}21", CURR, green_font)
wl(ws5, 26, 1, "5-Year Total Costs")
for i in range(NS): wf(ws5, 26, 2+i, f"={get_column_letter(2+i)}11+{get_column_letter(2+i)}12*5+{get_column_letter(2+i)}13", CURR)
wl(ws5, 27, 1, "5-Year Net Profit", f=navy_bold)
for i in range(NS):
    cell = wf(ws5, 27, 2+i, f"={get_column_letter(2+i)}25-{get_column_letter(2+i)}26", CURR, navy_bold); cell.border = total_bdr
wl(ws5, 28, 1, "5-Year ROI %", f=bold_font)
for i in range(NS): wf(ws5, 28, 2+i, f"=IF({get_column_letter(2+i)}26<>0,{get_column_letter(2+i)}27/{get_column_letter(2+i)}26,0)", PCT, bold_font)
wl(ws5, 29, 1, "5-Year Net Margin %", f=bold_font)
for i in range(NS): wf(ws5, 29, 2+i, f"=IF({get_column_letter(2+i)}25<>0,{get_column_letter(2+i)}27/{get_column_letter(2+i)}25,0)", PCT, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 6: VISUAL DASHBOARD (POLISHED)
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Dashboard"); ws6.sheet_properties.tabColor = "002060"

# ── Column widths (carefully matched to reference) ──
ws6.column_dimensions['A'].width = 1.5     # Left margin
ws6.column_dimensions['B'].width = 18      # Labels
ws6.column_dimensions['C'].width = 14      # Values
ws6.column_dimensions['D'].width = 1.5     # Gap between panels
# Chart area left (E-I = 5 cols)
for ch in 'EFGHI':
    ws6.column_dimensions[ch].width = 8.5
# Chart area right (J-N = 5 cols)
for ch in 'JKLMN':
    ws6.column_dimensions[ch].width = 8.5
ws6.column_dimensions['O'].width = 1       # Right margin

# ── Consistent row heights ──
ws6.row_dimensions[1].height = 6           # Top margin
ws6.row_dimensions[2].height = 35          # Header row
ws6.row_dimensions[3].height = 6           # Spacing
for rr in range(4, 35):
    ws6.row_dimensions[rr].height = 18     # Standard rows

# ── Dashboard styles ──
PANEL_BG  = PatternFill("solid", fgColor="F2F6FA")
CARD_BG   = PatternFill("solid", fgColor="FFFFFF")
HDR_BG    = PatternFill("solid", fgColor="1B3A5C")
DROP_BG   = PatternFill("solid", fgColor="1565C0")
TILE_BG   = PatternFill("solid", fgColor="FFFFFF")
MARGIN_GR = PatternFill("solid", fgColor="2E7D32")
MARGIN_YL = PatternFill("solid", fgColor="F9A825")
MARGIN_BL = PatternFill("solid", fgColor="1565C0")
PB_BG     = PatternFill("solid", fgColor="E3F2FD")
PB_INNER  = PatternFill("solid", fgColor="BBDEFB")

# Fonts
f_title   = Font(bold=True, size=15, name="Calibri", color="FFFFFF")
f_rtitle  = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
f_label   = Font(bold=True, size=10, name="Calibri", color="37474F")
f_value   = Font(bold=True, size=16, name="Calibri", color="1B3A5C")
f_unit    = Font(size=9, name="Calibri", color="90A4AE")
f_drop    = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
f_sub     = Font(size=10, name="Calibri", color="607D8B")
f_mhead   = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
f_mlab    = Font(size=11, name="Calibri", color="455A64")
f_mpct    = Font(bold=True, size=12, name="Calibri", color="1B3A5C")
f_pbnum   = Font(bold=True, size=32, name="Calibri", color="1565C0")
f_pbday   = Font(bold=True, size=14, name="Calibri", color="455A64")
f_pblbl   = Font(size=10, name="Calibri", color="78909C")

# Borders
bdr_panel = Border(
    left=Side("medium", color="B0BEC5"), right=Side("medium", color="B0BEC5"),
    top=Side("medium", color="B0BEC5"), bottom=Side("medium", color="B0BEC5"))
bdr_tile = Border(
    left=Side("thin", color="CFD8DC"), right=Side("thin", color="CFD8DC"),
    top=Side("thin", color="CFD8DC"), bottom=Side("thin", color="CFD8DC"))
bdr_card = Border(
    left=Side("thin", color="E0E0E0"), right=Side("thin", color="E0E0E0"),
    top=Side("thin", color="E0E0E0"), bottom=Side("thin", color="E0E0E0"))
no_bdr = Border()

IDX = "$C$52"
def idx(rr):
    return f"INDEX($C${rr}:$G${rr},1,{IDX})"

# ═══════════════════════════════════════════════════════════════
# HELPER: Draw bordered panel background
# ═══════════════════════════════════════════════════════════════
def draw_panel(ws, r1, c1, r2, c2, fill, border):
    """Fill a rectangular area with background color and proper borders."""
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c)
            cell.fill = fill
            # Apply border edges
            left = border.left if c == c1 else Side()
            right = border.right if c == c2 else Side()
            top = border.top if r == r1 else Side()
            bottom = border.bottom if r == r2 else Side()
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# ═══════════════════════════════════════════════════════════════
# LEFT PANEL: INPUT PARAMETERS (B2:C26)
# ═══════════════════════════════════════════════════════════════

# Panel background
draw_panel(ws6, 2, 2, 26, 3, PANEL_BG, bdr_panel)

# Header: "Input Parameters"
ws6.merge_cells('B2:C2')
c = ws6['B2']; c.value = "Input Parameters"
c.font = f_title; c.fill = HDR_BG
c.alignment = Alignment(horizontal="center", vertical="center")
# Re-apply header border after merge
for col in [2, 3]:
    ws6.cell(2, col).fill = HDR_BG
    ws6.cell(2, col).border = Border(
        left=Side("medium", color="B0BEC5") if col == 2 else Side(),
        right=Side("medium", color="B0BEC5") if col == 3 else Side(),
        top=Side("medium", color="B0BEC5"), bottom=Side())

# Row 3: spacing (already set small)

# Row 4: "Meter Rollout:" label
ws6['B4'].value = "Meter Rollout:"
ws6['B4'].font = f_label; ws6['B4'].fill = PANEL_BG

# Row 5: Dropdown (blue)
ws6.merge_cells('B5:C5')
ws6['B5'].value = "10,000 Meters"
ws6['B5'].font = f_drop; ws6['B5'].fill = DROP_BG
ws6['B5'].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[5].height = 28
for col in [2, 3]:
    ws6.cell(5, col).fill = DROP_BG
    ws6.cell(5, col).border = Border(
        left=Side("medium", color="0D47A1") if col == 2 else Side(),
        right=Side("medium", color="0D47A1") if col == 3 else Side(),
        top=Side("medium", color="0D47A1"), bottom=Side("medium", color="0D47A1"))

# Row 6: spacer

# ── Parameter cards: each takes 3 rows (label, value, spacer) ──
def make_card(start_row, label, value_or_formula, unit_text, is_formula=False):
    """Create a parameter card at start_row using 3 rows."""
    # Label row
    ws6.cell(start_row, 2).value = label
    ws6.cell(start_row, 2).font = f_label
    ws6.cell(start_row, 2).fill = PANEL_BG
    ws6.cell(start_row, 3).value = unit_text
    ws6.cell(start_row, 3).font = f_unit
    ws6.cell(start_row, 3).fill = PANEL_BG
    ws6.cell(start_row, 3).alignment = Alignment(horizontal="right")

    # Value row — white card background with border
    ws6.merge_cells(f'B{start_row+1}:C{start_row+1}')
    vc = ws6.cell(start_row+1, 2)
    vc.value = value_or_formula
    vc.font = f_value
    vc.fill = CARD_BG
    vc.number_format = '"N$ "#,##0'
    vc.alignment = Alignment(horizontal="left", vertical="center")
    ws6.row_dimensions[start_row+1].height = 26
    for col in [2, 3]:
        ws6.cell(start_row+1, col).fill = CARD_BG
        ws6.cell(start_row+1, col).border = bdr_card

# Cards
make_card(7, "Hardware Price", 5500, "Per Unit")
make_card(10, "Installation Fee", 700, "Per Unit")
make_card(13, "Manufacturing Cost", f"={idx(54)}", "Per Unit", True)
make_card(16, "Personnel Cost", f"={idx(56)}", "/ Year", True)
make_card(19, "Infrastructure Cost", f"={idx(57)}", "/ Year", True)
make_card(22, "Service Revenue", f"={idx(53)}*SUM(Revenue!$E$6:$E$10)", "/ Year", True)
make_card(25, "Equipment", f"={idx(58)}", "One-Time", True)

# ═══════════════════════════════════════════════════════════════
# RIGHT PANEL HEADER (E2:N2)
# ═══════════════════════════════════════════════════════════════
ws6.merge_cells('E2:N2')
c = ws6['E2']
c.value = '="5-Year Financial Overview ("&$B$5&")"'
c.font = f_rtitle; c.fill = HDR_BG
c.alignment = Alignment(horizontal="center", vertical="center")
for col in range(5, 15):  # E=5 to N=14
    ws6.cell(2, col).fill = HDR_BG

# ═══════════════════════════════════════════════════════════════
# TILE BACKGROUNDS (4 quadrants with thin borders)
# ═══════════════════════════════════════════════════════════════
# Top-left tile (E4:I16) — Annual Revenue & Profit
draw_panel(ws6, 4, 5, 16, 9, TILE_BG, bdr_tile)
# Top-right tile (J4:N16) — Profit Margins Donut
draw_panel(ws6, 4, 10, 16, 14, TILE_BG, bdr_tile)
# Bottom-left tile (E17:I27) — Cumulative Cash Flow
draw_panel(ws6, 17, 5, 27, 9, TILE_BG, bdr_tile)
# Bottom-right tile (J17:N27) — Margins + Payback
draw_panel(ws6, 17, 10, 27, 14, TILE_BG, bdr_tile)

# ═══════════════════════════════════════════════════════════════
# HIDDEN DATA AREA (row 50+)
# ═══════════════════════════════════════════════════════════════
ref_font = Font(size=7, color="BDBDBD", name="Calibri")

ws6.cell(50, 2, "DATA AREA").font = ref_font
for i, s in enumerate(SCENARIOS):
    ws6.cell(51, 3+i, s).font = ref_font

ws6.cell(52, 2, "Index").font = ref_font
ws6.cell(52, 3).value = '=MATCH($B$5,$C$51:$G$51,0)'
ws6.cell(52, 3).font = ref_font

for i, m in enumerate(MC):
    ws6.cell(53, 3+i, m).font = ref_font

# Mfg cost/meter
for i in range(NS):
    ws6.cell(54, 3+i).value = f"='Manufacturing Cost'!C{MFG_S+i}"; ws6.cell(54, 3+i).font = ref_font

# Total mfg
for i in range(NS):
    ws6.cell(55, 3+i).value = f"='Manufacturing Cost'!D{MFG_S+i}"; ws6.cell(55, 3+i).font = ref_font

# Personnel cost
for i in range(NS):
    ws6.cell(56, 3+i).value = f"='Personnel Cost'!{get_column_letter(3+i)}{TOTAL_COST_ROW}"; ws6.cell(56, 3+i).font = ref_font

# Infrastructure cost
for i in range(NS):
    ws6.cell(57, 3+i).value = f"='Infrastructure & Equipment'!{get_column_letter(3+i)}{INFRA_TOT}"; ws6.cell(57, 3+i).font = ref_font

# Equipment cost
for i in range(NS):
    ws6.cell(58, 3+i).value = f"='Infrastructure & Equipment'!{get_column_letter(3+i)}{EQUIP_TOT}"; ws6.cell(58, 3+i).font = ref_font

# ── Year-by-year data (rows 61-70) ──
for i, yr in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    ws6.cell(61, 4+i, yr).font = ref_font

# Revenue
ws6.cell(62, 3, "Revenue").font = ref_font
ws6.cell(62, 4).value = f"={idx(53)}*Revenue!$E$11"
for yc in range(5, 9):
    ws6.cell(62, yc).value = f"={idx(53)}*SUM(Revenue!$E$6:$E$10)"

# EBITDA
ws6.cell(63, 3, "EBITDA").font = ref_font
ws6.cell(63, 4).value = "=D62-D65-D66"
for yc in range(5, 9):
    ws6.cell(63, yc).value = f"={get_column_letter(yc)}62-{get_column_letter(yc)}66"

# Net Profit
ws6.cell(64, 3, "Net Profit").font = ref_font
ws6.cell(64, 4).value = "=D63-D67"
for yc in range(5, 9):
    ws6.cell(64, yc).value = f"={get_column_letter(yc)}63-{get_column_letter(yc)}67"

# Mfg cost (one-time Year 1)
ws6.cell(65, 3, "Mfg Cost").font = ref_font
ws6.cell(65, 4).value = f"={idx(55)}"
for yc in range(5, 9):
    ws6.cell(65, yc, 0).font = ref_font

# Operating costs (recurring every year)
ws6.cell(66, 3, "Oper Cost").font = ref_font
ws6.cell(66, 4).value = f"={idx(56)}+{idx(57)}"
for yc in range(5, 9):
    ws6.cell(66, yc).value = "=$D$66"

# Equipment (one-time Year 1)
ws6.cell(67, 3, "Equipment").font = ref_font
ws6.cell(67, 4).value = f"={idx(58)}"
for yc in range(5, 9):
    ws6.cell(67, yc, 0).font = ref_font

# Total costs
ws6.cell(68, 3, "Total Costs").font = ref_font
for yc in range(4, 9):
    cl = get_column_letter(yc)
    ws6.cell(68, yc).value = f"={cl}65+{cl}66+{cl}67"

# Cumulative revenue
ws6.cell(69, 3, "Cum Rev").font = ref_font
ws6.cell(69, 4).value = "=D62"
for yc in range(5, 9):
    ws6.cell(69, yc).value = f"={get_column_letter(yc-1)}69+{get_column_letter(yc)}62"

# Cumulative net profit
ws6.cell(70, 3, "Cum Profit").font = ref_font
ws6.cell(70, 4).value = "=D64"
for yc in range(5, 9):
    ws6.cell(70, yc).value = f"={get_column_letter(yc-1)}70+{get_column_letter(yc)}64"

# ── Donut data (row 72-73) ──
donut_labels = ["Net Profit", "Operating", "Manufacturing", "Equipment"]
donut_refs = ["=D64", "=D66", "=D65", "=D67"]
for i, (lab, ref) in enumerate(zip(donut_labels, donut_refs)):
    ws6.cell(72, 4+i, lab).font = ref_font
    ws6.cell(73, 4+i).value = ref; ws6.cell(73, 4+i).font = ref_font

# ── Metrics (row 75-78) ──
ws6.cell(75, 3, "Gross Margin").font = ref_font
ws6.cell(75, 4).value = "=IF(D62>0,(D62-D65)/D62,0)"
ws6.cell(76, 3, "Oper Margin").font = ref_font
ws6.cell(76, 4).value = "=IF(D62>0,D63/D62,0)"
ws6.cell(77, 3, "Net Margin").font = ref_font
ws6.cell(77, 4).value = "=IF(D62>0,D64/D62,0)"
ws6.cell(78, 3, "Payback Days").font = ref_font
ws6.cell(78, 4).value = "=ROUND((D65+D67)/(D62/365),0)"

# ── Chart display rows in MILLIONS (row 80-88) ──
ws6.cell(80, 3, "CHART DATA (Millions)").font = ref_font
for i, yr in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    ws6.cell(81, 4+i, yr).font = ref_font

# Revenue in millions
ws6.cell(82, 3, "Revenue M").font = ref_font
for yc in range(4, 9):
    ws6.cell(82, yc).value = f"={get_column_letter(yc)}62/1000000"

# EBITDA in millions
ws6.cell(83, 3, "EBITDA M").font = ref_font
for yc in range(4, 9):
    ws6.cell(83, yc).value = f"={get_column_letter(yc)}63/1000000"

# Net Profit in millions
ws6.cell(84, 3, "Net Profit M").font = ref_font
for yc in range(4, 9):
    ws6.cell(84, yc).value = f"={get_column_letter(yc)}64/1000000"

# Cumulative Revenue in millions
ws6.cell(85, 3, "Cum Rev M").font = ref_font
for yc in range(4, 9):
    ws6.cell(85, yc).value = f"={get_column_letter(yc)}69/1000000"

# Cumulative Profit in millions
ws6.cell(86, 3, "Cum Profit M").font = ref_font
for yc in range(4, 9):
    ws6.cell(86, yc).value = f"={get_column_letter(yc)}70/1000000"

# ═══════════════════════════════════════════════════════════════
# CHART 1: Annual Revenue & Profit (TOP LEFT — E4:I16)
# ═══════════════════════════════════════════════════════════════
cats = Reference(ws6, min_col=4, min_row=81, max_col=8, max_row=81)

bar1 = BarChart()
bar1.type = "col"; bar1.grouping = "clustered"
bar1.title = "Annual Revenue & Profit (N$ Millions)"
bar1.style = 10
bar1.y_axis.title = "N$ Millions"
bar1.y_axis.numFmt = '0'

# Revenue bars (dark blue) — from millions row
d_rev = Reference(ws6, min_col=4, min_row=82, max_col=8, max_row=82)
bar1.add_data(d_rev, from_rows=True)
bar1.series[0].title = SeriesLabel(v="Revenue")
bar1.series[0].graphicalProperties.solidFill = "1B3A5C"

# EBITDA bars (medium blue)
d_ebitda = Reference(ws6, min_col=4, min_row=83, max_col=8, max_row=83)
bar1.add_data(d_ebitda, from_rows=True)
bar1.series[1].title = SeriesLabel(v="EBITDA")
bar1.series[1].graphicalProperties.solidFill = "2E75B6"

# Net Profit bars (light blue)
d_np = Reference(ws6, min_col=4, min_row=84, max_col=8, max_row=84)
bar1.add_data(d_np, from_rows=True)
bar1.series[2].title = SeriesLabel(v="Net Profit")
bar1.series[2].graphicalProperties.solidFill = "90CAF9"

# Line overlay (green trend)
line1 = LineChart()
d_trend = Reference(ws6, min_col=4, min_row=82, max_col=8, max_row=82)
line1.add_data(d_trend, from_rows=True)
line1.series[0].title = SeriesLabel(v="Revenue Trend")
line1.series[0].graphicalProperties.line.solidFill = "4CAF50"
line1.series[0].graphicalProperties.line.width = 25000
line1.y_axis.axId = 200

bar1 += line1
bar1.set_categories(cats)
bar1.legend.position = 'b'
bar1.width = 17; bar1.height = 12
ws6.add_chart(bar1, "E4")

# ═══════════════════════════════════════════════════════════════
# CHART 2: Profit Margins Donut (TOP RIGHT — J4:N16)
# ═══════════════════════════════════════════════════════════════
ch2 = DoughnutChart()
ch2.title = "Profit Margins"
ch2.style = 10

donut_data = Reference(ws6, min_col=4, min_row=73, max_col=7, max_row=73)
donut_cats = Reference(ws6, min_col=4, min_row=72, max_col=7, max_row=72)
ch2.add_data(donut_data, from_rows=True)
ch2.set_categories(donut_cats)

# Color slices: green (profit), orange (oper), red (mfg), gray (equip)
donut_colors = ["4CAF50", "FF9800", "E53935", "78909C"]
for i, color in enumerate(donut_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    ch2.series[0].data_points.append(pt)

ch2.legend.position = 'b'
ch2.width = 14; ch2.height = 12
ws6.add_chart(ch2, "J4")

# ═══════════════════════════════════════════════════════════════
# CHART 3: Cumulative Cash Flow (BOTTOM LEFT — E17:I27)
# ═══════════════════════════════════════════════════════════════
# Year labels for bottom charts
cum_cats = Reference(ws6, min_col=4, min_row=81, max_col=8, max_row=81)

bar3 = BarChart()
bar3.type = "col"; bar3.grouping = "clustered"
bar3.title = "Cumulative Cash Flow (N$ Millions)"
bar3.style = 10
bar3.y_axis.title = "N$ Millions"
bar3.y_axis.numFmt = '0'

d_cr = Reference(ws6, min_col=4, min_row=85, max_col=8, max_row=85)
bar3.add_data(d_cr, from_rows=True)
bar3.series[0].title = SeriesLabel(v="Cumulative Revenue")
bar3.series[0].graphicalProperties.solidFill = "2E7D32"

d_cp = Reference(ws6, min_col=4, min_row=86, max_col=8, max_row=86)
bar3.add_data(d_cp, from_rows=True)
bar3.series[1].title = SeriesLabel(v="Cumulative Profit")
bar3.series[1].graphicalProperties.solidFill = "81C784"

bar3.set_categories(cum_cats)
bar3.legend.position = 'b'
bar3.width = 17; bar3.height = 11
ws6.add_chart(bar3, "E17")

# ═══════════════════════════════════════════════════════════════
# BOTTOM RIGHT TILE: Profit Margins + Payback Period (J17:N27)
# ═══════════════════════════════════════════════════════════════

# Section title: "Profit Margins"
ws6.merge_cells('J17:N17')
ws6['J17'].value = "Profit Margins"
ws6['J17'].font = f_mhead
ws6['J17'].alignment = Alignment(horizontal="center", vertical="center")
ws6['J17'].fill = TILE_BG
ws6.row_dimensions[17].height = 22

# Margin rows with colored indicator dots
margin_items = [
    (18, "Gross Margin", MARGIN_GR, "=D75"),
    (19, "Operating Margin", MARGIN_YL, "=D76"),
    (20, "Net Margin", MARGIN_BL, "=D77"),
]
for mr, lab, dot_fill, formula in margin_items:
    ws6.row_dimensions[mr].height = 22
    # Color indicator (col J)
    ws6.cell(mr, 10).fill = dot_fill
    ws6.cell(mr, 10).border = Border(
        left=Side("thin", color="CFD8DC"),
        right=Side(), top=Side(), bottom=Side())
    # Label (cols K-L)
    ws6.merge_cells(f'K{mr}:L{mr}')
    ws6.cell(mr, 11).value = lab
    ws6.cell(mr, 11).font = f_mlab
    ws6.cell(mr, 11).fill = TILE_BG
    ws6.cell(mr, 11).alignment = Alignment(vertical="center")
    # Percentage value (cols M-N)
    ws6.merge_cells(f'M{mr}:N{mr}')
    c = ws6.cell(mr, 13)
    c.value = formula
    c.font = f_mpct
    c.fill = TILE_BG
    c.number_format = '0.0%'
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws6.cell(mr, 14).fill = TILE_BG
    ws6.cell(mr, 14).border = Border(right=Side("thin", color="CFD8DC"))

# Row 21: Spacer
ws6.row_dimensions[21].height = 8

# Section title: "Payback Period"
ws6.merge_cells('J22:N22')
ws6['J22'].value = "Payback Period"
ws6['J22'].font = f_mhead
ws6['J22'].alignment = Alignment(horizontal="center", vertical="center")
ws6['J22'].fill = TILE_BG
ws6.row_dimensions[22].height = 22

# Payback display box (rows 23-26)
for r in range(23, 27):
    for c in range(10, 15):  # J=10 to N=14
        ws6.cell(r, c).fill = PB_BG
# Inner box for the number
for r in range(23, 26):
    for c in range(11, 14):  # K=11 to M=13
        ws6.cell(r, c).fill = PB_INNER
        # Box border
        left = Side("medium", color="90CAF9") if c == 11 else Side()
        right = Side("medium", color="90CAF9") if c == 13 else Side()
        top = Side("medium", color="90CAF9") if r == 23 else Side()
        bottom = Side("medium", color="90CAF9") if r == 25 else Side()
        ws6.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)

# Big number
ws6.merge_cells('K23:M24')
ws6['K23'].value = "=D78"
ws6['K23'].font = f_pbnum
ws6['K23'].fill = PB_INNER
ws6['K23'].alignment = Alignment(horizontal="center", vertical="center")
ws6['K23'].number_format = '0'
ws6.row_dimensions[23].height = 30
ws6.row_dimensions[24].height = 20

# "Days" label
ws6.merge_cells('K25:M25')
ws6['K25'].value = "Days"
ws6['K25'].font = f_pbday
ws6['K25'].fill = PB_INNER
ws6['K25'].alignment = Alignment(horizontal="center", vertical="top")
ws6.row_dimensions[25].height = 20

# "Payback Period" sub-label
ws6.merge_cells('J26:N26')
ws6['J26'].value = "Payback Period"
ws6['J26'].font = f_pblbl
ws6['J26'].fill = PB_BG
ws6['J26'].alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════════════
# DATA VALIDATION (Dropdown)
# ═══════════════════════════════════════════════════════════════
dv = DataValidation(type="list", formula1="=$C$51:$G$51", allow_blank=False)
dv.prompt = "Select a rollout scenario"
dv.promptTitle = "Meter Rollout"
ws6.add_data_validation(dv)
dv.add(ws6['B5'])

# ═══════════════════════════════════════════════════════════════
# FINALIZE
# ═══════════════════════════════════════════════════════════════
wb.move_sheet("Dashboard", offset=-5)

# Freeze panes on data sheets
ws1.freeze_panes = "B2"
ws5.freeze_panes = "B4"

# Print settings
ws6.page_setup.orientation = "landscape"
ws6.page_setup.fitToWidth = 1
ws6.page_setup.fitToHeight = 1
ws6.print_area = 'A1:O27'

# Hide gridlines on Dashboard
ws6.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v6.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print("v6: Polished tiles, clear panel borders, organized layout")
