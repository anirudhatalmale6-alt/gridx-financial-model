"""
GRIDx Financial Dashboard v11 — Client's v6 data + v10 dark dashboard design
Uses the EXACT data structure and numbers from client's GRIDx_Financial_Dashboard_v6.xlsx
Tabs: Dashboard, Revenue, Personnel Cost, Infrastructure & Equipment,
      Manufacturing Cost, Profitability Analysis
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Data from client's v6 ──
SCENARIOS = ["3,000 Meters","5,000 Meters","10,000 Meters","20,000 Meters","50,000 Meters"]
SC_SHORT = ["3K","5K","10K","20K","50K"]
MC = [3000, 5000, 10000, 20000, 50000]; NS = 5

# Revenue per meter
HW = 5500       # one-time
INST = 700      # one-time
MAINT = 25      # N$500/yr x 5% adoption
WIFI = 90       # N$25/mo x 30%
SMS_R = 30      # N$2.50 x 50% x 2/mo
APP = 624       # N$65/mo x 80%
RE = 76.80      # N$3200/blk/mo, 1 blk/500m
SVC_ANNUAL = MAINT + WIFI + SMS_R + APP + RE  # 845.80
Y1_PM = HW + INST + SVC_ANNUAL  # 7045.80

# Manufacturing
MFG_U = [2000, 2000, 1700, 1700, 1462]
MFG_TOT_V = [6000000, 10000000, 17000000, 34000000, 73100000]

# Personnel annual cost by scenario (from client's v6)
PERS_COST = [866400, 2560800, 5140800, 7173600, 9571200]

# Infrastructure annual by scenario
INFRA_COST = [240000, 420000, 840000, 1560000, 3480000]
INFRA_DETAIL = {
    "Data Storage & Transmission": [144000, 240000, 480000, 960000, 2400000],
    "Office Space Rental": [60000, 120000, 240000, 360000, 600000],
    "Utilities & Connectivity": [36000, 60000, 120000, 240000, 480000],
}

# Equipment one-time by scenario
EQUIP_DETAIL = {
    "Service Vehicles": [250000, 500000, 750000, 750000, 1250000],
    "Field Test Equipment Kits": [85000, 170000, 340000, 680000, 1530000],
    "Spare Parts Inventory": [269700, 449500, 899000, 1798000, 4495000],
    "Workstations": [75000, 210000, 420000, 630000, 930000],
    "Network Monitoring Tools": [120000, 120000, 240000, 360000, 600000],
    "Office Equipment & Furniture": [50000, 140000, 280000, 420000, 620000],
    "Server/Cloud Infrastructure": [100000, 150000, 250000, 400000, 800000],
}
EQUIP_TOT_V = [949700, 1739500, 3179000, 5038000, 10225000]

# Operating cost = Personnel + Infrastructure
OPER_COST = [p + i for p, i in zip(PERS_COST, INFRA_COST)]  # annual

# Year 1 revenue by scenario
Y1_REV = [round(m * Y1_PM) for m in MC]
# Year 2-5 annual service revenue
ANN_SVC = [round(m * SVC_ANNUAL) for m in MC]
# 5-Year total revenue
REV_5Y = [y1 + ann * 4 for y1, ann in zip(Y1_REV, ANN_SVC)]

# Profitability
GROSS_PROF = [y1 - mfg for y1, mfg in zip(Y1_REV, MFG_TOT_V)]
EBITDA = [gp - oc for gp, oc in zip(GROSS_PROF, OPER_COST)]
# Net profit = EBITDA (no depreciation/tax in model)
NET_PROF_Y1 = EBITDA[:]

# 5-Year costs
COSTS_5Y = [mfg + oper * 5 + equip for mfg, oper, equip in zip(MFG_TOT_V, OPER_COST, EQUIP_TOT_V)]
PROFIT_5Y = [rev - cost for rev, cost in zip(REV_5Y, COSTS_5Y)]

# Margins
GROSS_MARGIN = [gp / y1 for gp, y1 in zip(GROSS_PROF, Y1_REV)]
OPER_MARGIN = [eb / y1 for eb, y1 in zip(EBITDA, Y1_REV)]
NET_MARGIN = OPER_MARGIN[:]

# Payback (days)
PAYBACK_DAYS = [round(cost_y1 / (y1/365)) for cost_y1, y1 in
                zip([mfg + oc + eq for mfg, oc, eq in zip(MFG_TOT_V, OPER_COST, EQUIP_TOT_V)], Y1_REV)]

# Headcount
POSITIONS = [
    ("Chief Executive Officer (CEO)", 35000, [1,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Chief Operations Officer (COO)", 28000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Chief Financial Officer (CFO)", 28000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Chief Technology Officer (CTO)", 30000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of Business Development", 25000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of Human Resources", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Legal & Compliance Officer", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Operations Manager", 20000, [0,1,1,1,1], "OPERATIONS & FIELD SERVICES"),
    ("Field Support Technicians", 9000, [2,3,4,5,8], "OPERATIONS & FIELD SERVICES"),
    ("Installation Team Lead", 11000, [0,1,1,2,3], "OPERATIONS & FIELD SERVICES"),
    ("Logistics Coordinator", 8000, [0,0,1,1,2], "OPERATIONS & FIELD SERVICES"),
    ("Engineering Manager", 22000, [0,0,1,1,1], "ENGINEERING & QUALITY"),
    ("Quality Assurance Engineers", 12000, [1,1,1,3,5], "ENGINEERING & QUALITY"),
    ("Firmware/Software Engineers", 15000, [0,1,2,3,5], "ENGINEERING & QUALITY"),
    ("Hardware Engineers", 15000, [0,0,1,2,3], "ENGINEERING & QUALITY"),
    ("Customer Support Manager", 18000, [0,0,1,1,1], "CUSTOMER SUCCESS & SUPPORT"),
    ("Customer Support Agents", 7200, [1,2,2,4,8], "CUSTOMER SUCCESS & SUPPORT"),
    ("Technical Support Specialists", 9000, [0,0,1,2,3], "CUSTOMER SUCCESS & SUPPORT"),
    ("Sales Manager", 20000, [0,0,1,1,1], "SALES & MARKETING"),
    ("Sales Representatives", 10000, [0,1,2,3,5], "SALES & MARKETING"),
    ("Marketing Coordinator", 9000, [0,0,1,1,2], "SALES & MARKETING"),
    ("Finance Manager", 20000, [0,0,1,1,1], "FINANCE & ADMINISTRATION"),
    ("Accountant", 11000, [0,1,1,2,3], "FINANCE & ADMINISTRATION"),
    ("Administrative Assistant", 7000, [0,0,1,2,3], "FINANCE & ADMINISTRATION"),
]
HEADCOUNTS = [sum(hc[i] for _, _, hc, _ in POSITIONS) for i in range(NS)]

# ── Styles ──
NV = "1B2A4A"
hdr_fill = PatternFill("solid", fgColor=NV)
grn_fill = PatternFill("solid", fgColor="E2EFDA")
org_fill = PatternFill("solid", fgColor="FCE4D6")
ylw_fill = PatternFill("solid", fgColor="FFF2CC")
lt_fill  = PatternFill("solid", fgColor="D6E4F0")
hdr_f = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bf = Font(bold=True, size=10, name="Calibri")
b11 = Font(bold=True, size=11, name="Calibri")
lbl = Font(size=10, name="Calibri")
inp = Font(color="0000FF", size=10, name="Calibri")
ttl = Font(bold=True, color=NV, size=14, name="Calibri")
dept_f = Font(bold=True, color="4472C4", size=10, name="Calibri")
note_f = Font(italic=True, size=10, color="666666", name="Calibri")
tb = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
            top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
tot_b = Border(top=Side("medium", color=NV), bottom=Side("double", color=NV))
NUM = '#,##0'; CURR = 'N$#,##0'; PCT = '0.0%'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(r, cs+i, l); c.font = hdr_f; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = tb

def wl(ws, r, c, t, f=lbl, fi=None):
    cell = ws.cell(r, c, t); cell.font = f; cell.border = tb
    if fi: cell.fill = fi
    return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(r, c, v); cell.font = inp; cell.fill = ylw_fill
    cell.number_format = fmt; cell.border = tb; return cell

def wf(ws, r, c, formula, fmt=NUM, f=lbl, fi=None):
    cell = ws.cell(r, c, formula); cell.font = f; cell.number_format = fmt; cell.border = tb
    if fi: cell.fill = fi
    return cell

def fr(ws, r, cs, ce, fi):
    for c in range(cs, ce+1): ws.cell(r, c).fill = fi; ws.cell(r, c).border = tb

def scols(ws, lw=42):
    ws.column_dimensions['A'].width = lw
    for i in range(NS): ws.column_dimensions[get_column_letter(2+i)].width = 18

# ═══════════════════════════════════════════════════════════════
# SHEET 1: REVENUE
# ═══════════════════════════════════════════════════════════════
ws_r = wb.active; ws_r.title = "Revenue"; ws_r.sheet_properties.tabColor = "ED7D31"
ws_r.column_dimensions['A'].width = 40
for c in ['B','C','D','E']: ws_r.column_dimensions[c].width = 22

wl(ws_r, 1, 1, "Revenue Model Assumptions", f=ttl)
wh(ws_r, 3, ["Revenue Stream", "Rate", "Adoption Rate", "Calculation Basis", "Annual Amount per Meter (N$)"])
for r, nm, rate, ad, bas, val in [
    (4, "Meter Hardware Sale", "N$5,500 (one-time)", "100%", "One-time revenue per installed meter", HW),
    (5, "Meter Installation Fee", "N$700 (one-time)", "100%", "One-time fee per new meter installation", INST),
    (6, "Meter Maintenance", "N$500 / year", "5%", "Annual maintenance fee on installed meters", MAINT),
    (7, "Wi-Fi Subscription", "N$25 / month", "30%", "Recurring", WIFI),
    (8, "SMS Notifications", "N$2.50 / SMS", "50%, 2 SMS/month", "Recurring", SMS_R),
    (9, "Mobile App Subscription", "N$65 / month", "80%", "Recurring", APP),
    (10, "Real Estate Management Platform", "N$3,200 / block / month", "1 block per 500 meters", "Recurring", RE),
]:
    wl(ws_r, r, 1, nm); wl(ws_r, r, 2, rate); wl(ws_r, r, 3, ad)
    wl(ws_r, r, 4, bas); wi(ws_r, r, 5, val, 'N$#,##0.00')

wl(ws_r, 11, 1, "Total Year 1 Revenue per Meter", f=bf, fi=grn_fill); fr(ws_r, 11, 2, 5, grn_fill)
wf(ws_r, 11, 5, "=SUM(E4:E10)", 'N$#,##0.00', bf, grn_fill).border = tot_b

# Revenue by Year
wl(ws_r, 14, 1, "Revenue by Year", f=ttl)
wh(ws_r, 15, ["Year"] + SCENARIOS)
wl(ws_r, 16, 1, "Year 1")
for i in range(NS): wi(ws_r, 16, 2+i, Y1_REV[i], CURR)
for yr in range(2, 6):
    wl(ws_r, 14+yr, 1, f"Year {yr}")
    for i in range(NS): wi(ws_r, 14+yr, 2+i, ANN_SVC[i], CURR)
wl(ws_r, 21, 1, "5-Year Total", f=b11, fi=grn_fill); fr(ws_r, 21, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_r, 21, 2+i, f"=SUM({col}16:{col}20)", CURR, b11, grn_fill).border = tot_b
REV_5Y_ROW = 21

# Revenue chart
wl(ws_r, 23, 1, "VISUAL ANALYSIS", f=ttl)
rev_ch = BarChart(); rev_ch.type = "col"; rev_ch.grouping = "clustered"
rev_ch.title = "5-Year Total Revenue by Scenario"
rev_ch.y_axis.numFmt = '#,##0,,"M"'; rev_ch.y_axis.title = "N$ Millions"
cats = Reference(ws_r, min_col=2, min_row=15, max_col=6, max_row=15)
vals = Reference(ws_r, min_col=2, min_row=21, max_col=6, max_row=21)
rev_ch.add_data(vals, from_rows=True); rev_ch.set_categories(cats)
rev_ch.series[0].title = SeriesLabel(v="5-Year Revenue")
rev_ch.series[0].graphicalProperties.solidFill = "3B82F6"
rev_ch.width = 22; rev_ch.height = 12; rev_ch.legend = None
ws_r.add_chart(rev_ch, "A24")
ws_r.freeze_panes = "B4"
print(f"Revenue: 5Y row={REV_5Y_ROW}, Y1 10K={Y1_REV[2]}, 5Y 10K={REV_5Y[2]}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: PERSONNEL COST
# ═══════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Personnel Cost"); ws_p.sheet_properties.tabColor = "70AD47"
ws_p.column_dimensions['A'].width = 35
for i in range(NS): ws_p.column_dimensions[get_column_letter(2+i)].width = 16
ws_p.column_dimensions['H'].width = 35; ws_p.column_dimensions['I'].width = 16; ws_p.column_dimensions['J'].width = 16

wh(ws_p, 1, ["Resource Type"] + SCENARIOS)
wh(ws_p, 1, ["Position", "Monthly Salary (N$)", "Annual Salary (N$)"], cs=8)

r = 2; prev_dept = None
for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 6, lt_fill)
        prev_dept = dept; r += 1
    wl(ws_p, r, 1, nm)
    for i in range(NS): wi(ws_p, r, 2+i, hc[i], NUM)
    wl(ws_p, r, 8, nm); wi(ws_p, r, 9, monthly, CURR); wi(ws_p, r, 10, monthly*12, CURR)
    r += 1

# Subtotals by department
dept_start = {dept: None for _, _, _, dept in POSITIONS}
# Just add total
wl(ws_p, r+1, 1, "TOTAL PERSONNEL", f=b11, fi=grn_fill); fr(ws_p, r+1, 2, 6, grn_fill)
for i in range(NS):
    wi(ws_p, r+1, 2+i, HEADCOUNTS[i], NUM).font = b11
    ws_p.cell(r+1, 2+i).fill = grn_fill
HC_ROW = r+1

# Annual cost section
r = HC_ROW + 3
wl(ws_p, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=ttl)
r += 1
wh(ws_p, r, ["Position", "Unit Cost (N$/year)"] + SCENARIOS)
cost_hdr = r; r += 1; cost_start = r; prev_dept = None

for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 7, lt_fill)
        prev_dept = dept; r += 1
    annual = monthly * 12
    wl(ws_p, r, 1, nm); wi(ws_p, r, 2, annual, CURR)
    for i in range(NS): wi(ws_p, r, 3+i, hc[i] * annual, CURR)
    r += 1

wl(ws_p, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=b11, fi=grn_fill); fr(ws_p, r, 2, 7, grn_fill)
wl(ws_p, r, 2, "", f=b11, fi=grn_fill)
for i in range(NS):
    wi(ws_p, r, 3+i, PERS_COST[i], CURR).font = b11
    ws_p.cell(r, 3+i).fill = grn_fill; ws_p.cell(r, 3+i).border = tot_b
PERS_TOT_ROW = r

# Personnel chart
r += 2
wl(ws_p, r, 1, "VISUAL ANALYSIS", f=ttl)
pc = BarChart(); pc.type = "col"; pc.grouping = "clustered"
pc.title = "Headcount & Annual Personnel Cost by Scenario"
cats = Reference(ws_p, min_col=2, min_row=1, max_col=6, max_row=1)
vals = Reference(ws_p, min_col=2, min_row=HC_ROW, max_col=6, max_row=HC_ROW)
pc.add_data(vals, from_rows=True); pc.set_categories(cats)
pc.series[0].title = SeriesLabel(v="Headcount"); pc.series[0].graphicalProperties.solidFill = "70AD47"
pc.width = 22; pc.height = 12; pc.legend = None
ws_p.add_chart(pc, f"A{r+1}")
ws_p.freeze_panes = "B2"
print(f"Personnel: PERS_TOT_ROW={PERS_TOT_ROW}, HC_ROW={HC_ROW}, 10K={PERS_COST[2]}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: INFRASTRUCTURE & EQUIPMENT
# ═══════════════════════════════════════════════════════════════
ws_ie = wb.create_sheet("Infrastructure & Equipment"); ws_ie.sheet_properties.tabColor = "4472C4"
ws_ie.column_dimensions['A'].width = 35; ws_ie.column_dimensions['B'].width = 16
for i in range(NS): ws_ie.column_dimensions[get_column_letter(3+i)].width = 18

wl(ws_ie, 1, 1, "Infrastructure Costs (Annual)", f=ttl)
wh(ws_ie, 3, ["Cost Category", "Unit Cost"] + SCENARIOS)
r = 4
for cat, vals in INFRA_DETAIL.items():
    unit = {"Data Storage & Transmission": "N$48/meter", "Office Space Rental": "N$200/sq m",
            "Utilities & Connectivity": "Lump sum"}.get(cat, "")
    wl(ws_ie, r, 1, cat); wl(ws_ie, r, 2, unit)
    for i in range(NS): wi(ws_ie, r, 3+i, vals[i], CURR)
    r += 1
wl(ws_ie, r, 1, "Annual Infrastructure Total", f=b11, fi=grn_fill); fr(ws_ie, r, 2, 7, grn_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    wf(ws_ie, r, 3+i, f"=SUM({col}4:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
INFRA_TOT_ROW = r

r += 2
wl(ws_ie, r, 1, "Equipment Costs (One-Time Investment)", f=ttl)
r += 2
wh(ws_ie, r, ["Equipment Type", "Unit Cost"] + SCENARIOS)
eq_hdr = r; r += 1; eq_start = r
unit_costs = {"Service Vehicles": "N$250,000", "Field Test Equipment Kits": "N$85,000",
              "Spare Parts Inventory": "N$4,495 x 2%", "Workstations": "N$15,000",
              "Network Monitoring Tools": "N$120,000", "Office Equipment & Furniture": "Per employee",
              "Server/Cloud Infrastructure": "Lump sum"}
for cat, vals in EQUIP_DETAIL.items():
    wl(ws_ie, r, 1, cat); wl(ws_ie, r, 2, unit_costs.get(cat, ""))
    for i in range(NS): wi(ws_ie, r, 3+i, vals[i], CURR)
    r += 1
wl(ws_ie, r, 1, "Total Equipment Investment", f=b11, fi=grn_fill); fr(ws_ie, r, 2, 7, grn_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    wf(ws_ie, r, 3+i, f"=SUM({col}{eq_start}:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
EQUIP_TOT_ROW = r

# Chart
r += 2
wl(ws_ie, r, 1, "VISUAL ANALYSIS", f=ttl)
ec = BarChart(); ec.type = "col"; ec.grouping = "stacked"
ec.title = "Equipment Investment by Category"
colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "FF6384", "9966FF"]
cats = Reference(ws_ie, min_col=3, min_row=eq_hdr, max_col=7, max_row=eq_hdr)
for idx, cat in enumerate(EQUIP_DETAIL):
    vals = Reference(ws_ie, min_col=3, min_row=eq_start+idx, max_col=7, max_row=eq_start+idx)
    ec.add_data(vals, from_rows=True)
    ec.series[-1].title = SeriesLabel(v=cat)
    ec.series[-1].graphicalProperties.solidFill = colors[idx % len(colors)]
ec.set_categories(cats); ec.legend.position = 'b'
ec.width = 22; ec.height = 14
ws_ie.add_chart(ec, f"A{r+1}")
ws_ie.freeze_panes = "C4"
print(f"Infra & Equipment: INFRA_TOT={INFRA_TOT_ROW}, EQUIP_TOT={EQUIP_TOT_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 4: MANUFACTURING COST
# ═══════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("Manufacturing Cost"); ws_m.sheet_properties.tabColor = "C00000"
ws_m.column_dimensions['A'].width = 30; ws_m.column_dimensions['B'].width = 14
ws_m.column_dimensions['C'].width = 20; ws_m.column_dimensions['D'].width = 25

wl(ws_m, 1, 1, "Volume Discount Structure", f=ttl)
wh(ws_m, 3, ["Volume Tier", "Discount", "Cost per Meter (N$)", "Calculation Basis"])
for r, tier, disc, cost, bas in [
    (4, "Base (0-9,999 meters)", "0%", 2000, "N$2,000 per meter"),
    (5, "10,000-19,999 meters", "15%", 1700, "15% reduction"),
    (6, "20,000-49,999 meters", "30%", 1400, "Additional 15%"),
    (7, "50,000+ meters", "45%", 1100, "Additional 15%"),
]:
    wl(ws_m, r, 1, tier); wl(ws_m, r, 2, disc); wi(ws_m, r, 3, cost, CURR); wl(ws_m, r, 4, bas)

wl(ws_m, 10, 1, "Manufacturing Cost by Scenario", f=ttl)
wh(ws_m, 12, ["Scenario", "Meters", "Cost per Meter (N$)", "Total Manufacturing Cost (N$)"])
for i in range(NS):
    wl(ws_m, 13+i, 1, SCENARIOS[i])
    wi(ws_m, 13+i, 2, MC[i], NUM)
    wi(ws_m, 13+i, 3, MFG_U[i], CURR)
    wi(ws_m, 13+i, 4, MFG_TOT_V[i], CURR)

# Chart
wl(ws_m, 19, 1, "VISUAL ANALYSIS", f=ttl)
mc = BarChart(); mc.type = "col"; mc.grouping = "clustered"
mc.title = "Manufacturing Cost by Scenario"
mc.y_axis.numFmt = '#,##0,,"M"'; mc.y_axis.title = "N$ Millions"
cats = Reference(ws_m, min_col=1, min_row=13, max_row=17)
vals = Reference(ws_m, min_col=4, min_row=13, max_row=17)
mc.add_data(vals); mc.set_categories(cats)
mc.series[0].title = SeriesLabel(v="Manufacturing Cost"); mc.series[0].graphicalProperties.solidFill = "C00000"
mc.width = 20; mc.height = 12; mc.legend = None
ws_m.add_chart(mc, "A20")
print(f"Manufacturing: 10K cost={MFG_TOT_V[2]}")

# ═══════════════════════════════════════════════════════════════
# SHEET 5: PROFITABILITY ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws_pa = wb.create_sheet("Profitability Analysis"); ws_pa.sheet_properties.tabColor = "002060"; scols(ws_pa)
wl(ws_pa, 1, 1, "Revenue & Profitability Analysis (Year 1)", f=ttl)
wh(ws_pa, 3, ["Metric"] + SCENARIOS)

wl(ws_pa, 4, 1, "REVENUE (N$)", f=bf, fi=lt_fill); fr(ws_pa, 4, 2, 6, lt_fill)
wl(ws_pa, 5, 1, "Hardware Sales")
for i in range(NS): wi(ws_pa, 5, 2+i, MC[i]*HW, CURR)
wl(ws_pa, 6, 1, "Installation Fees")
for i in range(NS): wi(ws_pa, 6, 2+i, MC[i]*INST, CURR)
wl(ws_pa, 7, 1, "Service Revenues")
for i in range(NS): wi(ws_pa, 7, 2+i, ANN_SVC[i], CURR)
wl(ws_pa, 8, 1, "Total Revenue", f=b11, fi=grn_fill); fr(ws_pa, 8, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 8, 2+i, f"=SUM({col}5:{col}7)", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 9, 1, "MANUFACTURING COST (N$)", f=bf, fi=org_fill); fr(ws_pa, 9, 2, 6, org_fill)
wl(ws_pa, 10, 1, "Cost per Meter")
for i in range(NS): wi(ws_pa, 10, 2+i, MFG_U[i], CURR)
wl(ws_pa, 11, 1, "Total Manufacturing Cost")
for i in range(NS): wi(ws_pa, 11, 2+i, MFG_TOT_V[i], CURR)

wl(ws_pa, 12, 1, "OPERATING COSTS (N$)")
for i in range(NS): wi(ws_pa, 12, 2+i, OPER_COST[i], CURR)
wl(ws_pa, 13, 1, "EQUIPMENT INVESTMENT (N$)")
for i in range(NS): wi(ws_pa, 13, 2+i, EQUIP_TOT_V[i], CURR)

wl(ws_pa, 14, 1, "PROFITABILITY (N$)", f=bf, fi=lt_fill); fr(ws_pa, 14, 2, 6, lt_fill)
wl(ws_pa, 15, 1, "Gross Profit")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 15, 2+i, f"={col}8-{col}11", CURR)
wl(ws_pa, 16, 1, "Operating Profit (EBITDA)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 16, 2+i, f"={col}15-{col}12", CURR)
wl(ws_pa, 17, 1, "Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 17, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 17, 2+i, f"={col}16", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 18, 1, "MARGINS", f=bf, fi=lt_fill); fr(ws_pa, 18, 2, 6, lt_fill)
wl(ws_pa, 19, 1, "Gross Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 19, 2+i, f"=IF({col}8>0,{col}15/{col}8,0)", PCT)
wl(ws_pa, 20, 1, "Operating Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 20, 2+i, f"=IF({col}8>0,{col}16/{col}8,0)", PCT)
wl(ws_pa, 21, 1, "Net Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 21, 2+i, f"=IF({col}8>0,{col}17/{col}8,0)", PCT)

# 5-Year Projections
wl(ws_pa, 23, 1, "5-YEAR PROJECTIONS", f=b11, fi=lt_fill); fr(ws_pa, 23, 2, 6, lt_fill)
wh(ws_pa, 24, ["Metric"] + SCENARIOS)
wl(ws_pa, 25, 1, "5-Year Total Revenue")
for i in range(NS): wi(ws_pa, 25, 2+i, REV_5Y[i], CURR)
wl(ws_pa, 26, 1, "5-Year Total Costs")
for i in range(NS): wi(ws_pa, 26, 2+i, COSTS_5Y[i], CURR)
wl(ws_pa, 27, 1, "5-Year Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 27, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 27, 2+i, f"={col}25-{col}26", CURR, b11, grn_fill).border = tot_b
wl(ws_pa, 28, 1, "5-Year ROI %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 28, 2+i, f"=IF({col}26>0,{col}27/{col}26,0)", PCT)
wl(ws_pa, 29, 1, "5-Year Net Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 29, 2+i, f"=IF({col}25>0,{col}27/{col}25,0)", PCT)

# Profitability chart
wl(ws_pa, 31, 1, "VISUAL ANALYSIS", f=ttl)
prc = BarChart(); prc.type = "col"; prc.grouping = "clustered"
prc.title = "Year 1: Revenue vs Costs vs Profit"
prc.y_axis.numFmt = '#,##0,,"M"'; prc.y_axis.title = "N$ Millions"
cats = Reference(ws_pa, min_col=2, min_row=3, max_col=6, max_row=3)
for rr, label, color in [(8, "Revenue", "3B82F6"), (11, "Manufacturing", "EF4444"), (17, "Net Profit", "22C55E")]:
    vals = Reference(ws_pa, min_col=2, min_row=rr, max_col=6, max_row=rr)
    prc.add_data(vals, from_rows=True)
    prc.series[-1].title = SeriesLabel(v=label); prc.series[-1].graphicalProperties.solidFill = color
prc.set_categories(cats); prc.legend.position = 'b'
prc.width = 22; prc.height = 14
ws_pa.add_chart(prc, "A32")
ws_pa.freeze_panes = "B4"
print(f"Profitability: Y1 Rev 10K={Y1_REV[2]}, 5Y Rev 10K={REV_5Y[2]}, 5Y Profit 10K={PROFIT_5Y[2]}")

# ═══════════════════════════════════════════════════════════════
# SHEET 6: DASHBOARD (DARK NAVY + COMPREHENSIVE CHARTS)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "002060"

BG1="0F1B2D"; BG2="162236"; BG3="1D2B42"; BG4="243B55"
ACC_BLUE="3B82F6"; ACC_GREEN="22C55E"; ACC_RED="EF4444"; GOLD="F59E0B"
WHITE="FFFFFF"; LGRAY="94A3B8"; BDR_C="2D3F59"

bg1=PatternFill("solid",fgColor=BG1); bg2=PatternFill("solid",fgColor=BG2)
bg3=PatternFill("solid",fgColor=BG3); bg4=PatternFill("solid",fgColor=BG4)
acc_bl=PatternFill("solid",fgColor=ACC_BLUE); kpi_bg=PatternFill("solid",fgColor="1E3A5F")
tbl_h=PatternFill("solid",fgColor=BG4); tbl1=PatternFill("solid",fgColor=BG2); tbl2=PatternFill("solid",fgColor=BG3)

ft=Font(bold=True,size=16,name="Calibri",color=WHITE)
fsec=Font(bold=True,size=12,name="Calibri",color=ACC_BLUE)
fw=Font(size=10,name="Calibri",color=WHITE)
fwb=Font(bold=True,size=10,name="Calibri",color=WHITE)
fdr=Font(bold=True,size=12,name="Calibri",color=WHITE)
fkl=Font(size=9,name="Calibri",color=LGRAY)
fg=Font(bold=True,size=10,name="Calibri",color=ACC_GREEN)
frd=Font(size=10,name="Calibri",color=ACC_RED)
fgo=Font(bold=True,size=10,name="Calibri",color=GOLD)
ftn=Font(size=7,name="Calibri",color=BG1)
db=Border(left=Side("thin",color=BDR_C),right=Side("thin",color=BDR_C),
          top=Side("thin",color=BDR_C),bottom=Side("thin",color=BDR_C))

for c,w in [('A',1.5),('B',18),('C',14),('D',14),('E',1),('F',12),('G',12),
            ('H',12),('I',12),('J',12),('K',1),('L',12),('M',12),('N',12),('O',12),('P',1.5)]:
    ws.column_dimensions[c].width = w

for r in range(1, 80):
    ws.row_dimensions[r].height = 18
    for c in range(1, 17): ws.cell(r, c).fill = bg1

# ── HIDDEN DATA (rows 90+) ──
for i, s in enumerate(SCENARIOS): ws.cell(90, 3+i, s).font = ftn
ws.cell(91, 3).value = '=MATCH($N$6,$C$90:$G$90,0)'; ws.cell(91, 3).font = ftn
IDX = "$C$91"
def didx(rr): return f"INDEX($C${rr}:$G${rr},1,{IDX})"

for i in range(NS): ws.cell(92, 3+i, MC[i]).font = ftn       # meters
for i in range(NS): ws.cell(93, 3+i, MFG_U[i]).font = ftn    # mfg unit cost
for i in range(NS): ws.cell(94, 3+i, MFG_TOT_V[i]).font = ftn  # mfg total
for i in range(NS): ws.cell(95, 3+i, PERS_COST[i]).font = ftn  # personnel
for i in range(NS): ws.cell(96, 3+i, INFRA_COST[i]).font = ftn  # infra
for i in range(NS): ws.cell(97, 3+i, EQUIP_TOT_V[i]).font = ftn  # equipment
# Operating = personnel + infra
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(98, 3+i).value = f"={c}95+{c}96"; ws.cell(98, 3+i).font = ftn  # annual oper
# Y1 revenue
for i in range(NS): ws.cell(99, 3+i, Y1_REV[i]).font = ftn
# Annual service revenue
for i in range(NS): ws.cell(100, 3+i, ANN_SVC[i]).font = ftn
# 5Y revenue
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(101, 3+i).value = f"={c}99+{c}100*4"; ws.cell(101, 3+i).font = ftn
# 5Y costs
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(102, 3+i).value = f"={c}94+{c}98*5+{c}97"; ws.cell(102, 3+i).font = ftn
# 5Y profit
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(103, 3+i).value = f"={c}101-{c}102"; ws.cell(103, 3+i).font = ftn
# Gross margin (Y1)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(104, 3+i).value = f"=({c}99-{c}94)/{c}99"; ws.cell(104, 3+i).font = ftn
# Operating margin (Y1)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(105, 3+i).value = f"=({c}99-{c}94-{c}98)/{c}99"; ws.cell(105, 3+i).font = ftn
# Net margin
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(106, 3+i).value = f"={c}105"; ws.cell(106, 3+i).font = ftn
# Payback (days)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(107, 3+i).value = f"=({c}94+{c}98+{c}97)/({c}99/365)"; ws.cell(107, 3+i).font = ftn

# ── CHART DATA: Annual Rev/Costs/Profit (rows 110-115) ──
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(110, 3+i, yr).font = ftn
ws.cell(111, 2, "Revenue").font = ftn
ws.cell(111, 3).value = f"={didx(99)}/1000000"; ws.cell(111, 3).font = ftn
for yc in range(4, 8): ws.cell(111, yc).value = f"={didx(100)}/1000000"; ws.cell(111, yc).font = ftn
ws.cell(112, 2, "Total Costs").font = ftn
ws.cell(112, 3).value = f"=({didx(94)}+{didx(98)}+{didx(97)})/1000000"; ws.cell(112, 3).font = ftn
for yc in range(4, 8): ws.cell(112, yc).value = f"={didx(98)}/1000000"; ws.cell(112, yc).font = ftn
ws.cell(113, 2, "Net Profit").font = ftn
for yc in range(3, 8):
    ws.cell(113, yc).value = f"={get_column_letter(yc)}111-{get_column_letter(yc)}112"
    ws.cell(113, yc).font = ftn
# Cumulative
ws.cell(114, 2, "Cum Revenue").font = ftn
ws.cell(114, 3).value = "=C111"; ws.cell(114, 3).font = ftn
for yc in range(4, 8):
    ws.cell(114, yc).value = f"={get_column_letter(yc-1)}114+{get_column_letter(yc)}111"
    ws.cell(114, yc).font = ftn
ws.cell(115, 2, "Cum Profit").font = ftn
ws.cell(115, 3).value = "=C113"; ws.cell(115, 3).font = ftn
for yc in range(4, 8):
    ws.cell(115, yc).value = f"={get_column_letter(yc-1)}115+{get_column_letter(yc)}113"
    ws.cell(115, yc).font = ftn

# Cost breakdown for donut (row 120-121)
cost_cats = ["Net Profit", "Operating", "Manufacturing", "Equipment"]
for i, cat in enumerate(cost_cats): ws.cell(120, 3+i, cat).font = ftn
ws.cell(121, 3).value = f"={didx(99)}-{didx(94)}-{didx(98)}-{didx(97)}"; ws.cell(121, 3).font = ftn  # Net Profit Y1
ws.cell(121, 4).value = f"={didx(98)}"; ws.cell(121, 4).font = ftn  # Operating
ws.cell(121, 5).value = f"={didx(94)}"; ws.cell(121, 5).font = ftn  # Manufacturing
ws.cell(121, 6).value = f"={didx(97)}"; ws.cell(121, 6).font = ftn  # Equipment

# Margin data for donut (row 125-126)
ws.cell(125, 3, "Gross Margin").font = ftn; ws.cell(125, 4, "Operating Margin").font = ftn; ws.cell(125, 5, "Net Margin").font = ftn
ws.cell(126, 3).value = f"={didx(104)}"; ws.cell(126, 3).font = ftn
ws.cell(126, 4).value = f"={didx(105)}"; ws.cell(126, 4).font = ftn
ws.cell(126, 5).value = f"={didx(106)}"; ws.cell(126, 5).font = ftn

# Scenario comparison (row 130-133)
for i, sc in enumerate(SC_SHORT): ws.cell(130, 3+i, sc).font = ftn
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(131, 3+i).value = f"={c}101/1000000"; ws.cell(131, 3+i).font = ftn  # Rev M
    ws.cell(132, 3+i).value = f"={c}102/1000000"; ws.cell(132, 3+i).font = ftn  # Costs M
    ws.cell(133, 3+i).value = f"={c}103/1000000"; ws.cell(133, 3+i).font = ftn  # Profit M

# ══════════════════════════════════════════════════════════════
# DASHBOARD VISIBLE LAYOUT
# ══════════════════════════════════════════════════════════════

# Title bar
ws.row_dimensions[2].height = 32
ws.merge_cells('B2:O2')
ws['B2'] = "GRIDx Smart Meter  -  5-Year Financial Model"; ws['B2'].font = ft; ws['B2'].fill = bg2
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 16): ws.cell(2, c).fill = bg2

ws.row_dimensions[3].height = 24
ws.merge_cells('B3:O3')
ws['B3'] = "Scenarios: 3,000  |  5,000  |  10,000  |  20,000  |  50,000 Meters  |  Currency: N$ (Namibian Dollar)"
ws['B3'].font = Font(size=9, color=LGRAY, name="Calibri")
ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

# KPI Cards
ws.row_dimensions[5].height = 36; ws.row_dimensions[6].height = 16
kpis = [
    (5, 2, 3, "5-Year Revenue", f'="N$"&TEXT({didx(101)}/1000000,"#,##0")&"M"', ACC_GREEN),
    (5, 6, 7, "5-Year Net Profit", f'="N$"&TEXT({didx(103)}/1000000,"#,##0")&"M"', ACC_GREEN),
    (5, 9, 10, "Gross Margin", f'=TEXT({didx(104)}*100,"0.0")&"%"', GOLD),
    (5, 12, 13, "Payback Period", f'=TEXT(ROUND({didx(107)},0),"#,##0")&" days"', ACC_BLUE),
]
for kr, kc1, kc2, label, formula, acc in kpis:
    for r in range(kr, kr+2):
        for c in range(kc1, kc2+1): ws.cell(r, c).fill = kpi_bg; ws.cell(r, c).border = db
    ws.merge_cells(start_row=kr, start_column=kc1, end_row=kr, end_column=kc2)
    ws.cell(kr, kc1).value = formula
    ws.cell(kr, kc1).font = Font(bold=True, size=18, name="Calibri", color=acc)
    ws.cell(kr, kc1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=kr+1, start_column=kc1, end_row=kr+1, end_column=kc2)
    ws.cell(kr+1, kc1).value = label; ws.cell(kr+1, kc1).font = fkl
    ws.cell(kr+1, kc1).alignment = Alignment(horizontal="center")

# Dropdown
ws.merge_cells('N5:O5')
ws.cell(5, 14).value = "SELECT SCENARIO"; ws.cell(5, 14).font = Font(bold=True, size=9, color=LGRAY, name="Calibri")
ws.cell(5, 14).alignment = Alignment(horizontal="center", vertical="center")
for c in [14,15]: ws.cell(5, c).fill = kpi_bg; ws.cell(5, c).border = db
ws.merge_cells('N6:O6')
ws['N6'] = "10,000 Meters"; ws['N6'].font = fdr; ws['N6'].fill = acc_bl
ws['N6'].alignment = Alignment(horizontal="center", vertical="center")
for c in [14,15]: ws.cell(6, c).fill = acc_bl; ws.cell(6, c).border = db

# Input Parameters
ws.row_dimensions[8].height = 22
ws.merge_cells('B8:D8'); ws['B8'] = "INPUT PARAMETERS"; ws['B8'].font = fsec
for pr, label, val, unit in [
    (9, "Meters:", f'=TEXT({didx(92)},"#,##0")', "units"),
    (10, "Hardware:", "N$5,500", "one-time/unit"),
    (11, "Installation:", "N$700", "one-time/unit"),
    (12, "Mfg Cost/Meter:", f'="N$"&TEXT({didx(93)},"#,##0")', "per unit"),
    (13, "Personnel Cost:", f'="N$"&TEXT({didx(95)},"#,##0")', "/year"),
    (14, "Infrastructure:", f'="N$"&TEXT({didx(96)},"#,##0")', "/year"),
    (15, "Service Revenue:", f'="N$"&TEXT({didx(100)},"#,##0")', "/year"),
]:
    ws.cell(pr, 2).value = label; ws.cell(pr, 2).font = fw
    ws.cell(pr, 3).value = val; ws.cell(pr, 3).font = fg
    ws.cell(pr, 4).value = unit; ws.cell(pr, 4).font = Font(size=8, color=LGRAY, name="Calibri")

# Financial Summary
ws.row_dimensions[17].height = 22
ws.merge_cells('B17:D17'); ws['B17'] = "5-YEAR FINANCIAL SUMMARY"; ws['B17'].font = fsec
for sr, label, formula, font in [
    (18, "Total Revenue:", f'="N$"&TEXT({didx(101)},"#,##0")', fg),
    (19, "Total Costs:", f'="N$"&TEXT({didx(102)},"#,##0")', frd),
    (20, "Net Profit:", f'="N$"&TEXT({didx(103)},"#,##0")', fgo),
    (21, "ROI:", f'=TEXT(IF({didx(102)}>0,{didx(103)}/{didx(102)},0)*100,"#,##0")&"%"', fg),
]:
    ws.cell(sr, 2).value = label; ws.cell(sr, 2).font = fw
    ws.cell(sr, 3).value = formula; ws.cell(sr, 3).font = font

# Profit Margins
ws.row_dimensions[22].height = 4  # spacer
ws.merge_cells('B23:D23'); ws['B23'] = "PROFIT MARGINS"; ws['B23'].font = fsec
for sr, label, formula in [
    (24, "Gross Margin:", f'=TEXT({didx(104)}*100,"0.0")&"%"'),
    (25, "Operating Margin:", f'=TEXT({didx(105)}*100,"0.0")&"%"'),
    (26, "Net Margin:", f'=TEXT({didx(106)}*100,"0.0")&"%"'),
]:
    ws.cell(sr, 2).value = label; ws.cell(sr, 2).font = fw
    ws.cell(sr, 3).value = formula; ws.cell(sr, 3).font = fgo

# ── CHART 1: Annual Revenue & Profit bar chart ──
ch1 = BarChart(); ch1.type = "col"; ch1.grouping = "clustered"
ch1.title = "Annual Revenue & Profit (N$ Millions)"
ch1.y_axis.numFmt = '0.0'; ch1.y_axis.title = "N$ Millions"
cats = Reference(ws, min_col=3, min_row=110, max_col=7, max_row=110)
d_rev = Reference(ws, min_col=3, min_row=111, max_col=7, max_row=111)
ch1.add_data(d_rev, from_rows=True)
ch1.series[0].title = SeriesLabel(v="Revenue"); ch1.series[0].graphicalProperties.solidFill = "3B82F6"
d_cost = Reference(ws, min_col=3, min_row=112, max_col=7, max_row=112)
ch1.add_data(d_cost, from_rows=True)
ch1.series[1].title = SeriesLabel(v="Costs"); ch1.series[1].graphicalProperties.solidFill = "EF4444"
# Profit as line
ln1 = LineChart()
d_prof = Reference(ws, min_col=3, min_row=113, max_col=7, max_row=113)
ln1.add_data(d_prof, from_rows=True)
ln1.series[0].title = SeriesLabel(v="Net Profit"); ln1.series[0].graphicalProperties.line.solidFill = "22C55E"
ln1.series[0].graphicalProperties.line.width = 28000
ln1.y_axis.numFmt = '0.0'
ch1.y_axis.crosses = "min"
ch1 += ln1
ch1.set_categories(cats); ch1.legend.position = 'b'
ch1.width = 18; ch1.height = 12
ws.add_chart(ch1, "F8")

# ── CHART 2: Profit Margins Doughnut ──
ch2 = DoughnutChart()
ch2.title = "Profit Margins"
cats2 = Reference(ws, min_col=3, min_row=125, max_col=5, max_row=125)
vals2 = Reference(ws, min_col=3, min_row=126, max_col=5, max_row=126)
ch2.add_data(vals2, from_rows=True); ch2.set_categories(cats2)
ch2.series[0].title = SeriesLabel(v="Margins")
ch2.width = 14; ch2.height = 12
ws.add_chart(ch2, "L8")

# ── CHART 3: Cumulative Cash Flow ──
ch3 = BarChart(); ch3.type = "col"; ch3.grouping = "clustered"
ch3.title = "Cumulative Cash Flow (N$ Millions)"
ch3.y_axis.numFmt = '0'; ch3.y_axis.title = "N$ Millions"
cats3 = Reference(ws, min_col=3, min_row=110, max_col=7, max_row=110)
d_cr = Reference(ws, min_col=3, min_row=114, max_col=7, max_row=114)
ch3.add_data(d_cr, from_rows=True)
ch3.series[0].title = SeriesLabel(v="Cum Revenue"); ch3.series[0].graphicalProperties.solidFill = "3B82F6"
d_cp = Reference(ws, min_col=3, min_row=115, max_col=7, max_row=115)
ch3.add_data(d_cp, from_rows=True)
ch3.series[1].title = SeriesLabel(v="Cum Profit"); ch3.series[1].graphicalProperties.solidFill = "22C55E"
ch3.set_categories(cats3); ch3.legend.position = 'b'
ch3.width = 14; ch3.height = 12
ws.add_chart(ch3, "L21")

# ── CHART 4: Year 1 Cost Breakdown Doughnut ──
ws.row_dimensions[28].height = 22
ws.merge_cells('B28:D28'); ws['B28'] = "COST ANALYSIS"; ws['B28'].font = fsec

ch4 = DoughnutChart()
ch4.title = "Year 1 Breakdown"
cats4 = Reference(ws, min_col=3, min_row=120, max_col=6, max_row=120)
vals4 = Reference(ws, min_col=3, min_row=121, max_col=6, max_row=121)
ch4.add_data(vals4, from_rows=True); ch4.set_categories(cats4)
ch4.series[0].title = SeriesLabel(v="Y1 Breakdown")
ch4.width = 14; ch4.height = 12
ws.add_chart(ch4, "B29")

# ── CHART 5: Operating Cost Breakdown ──
# Operating cost items for selected scenario
op_items = list(INFRA_DETAIL.keys()) + ["Personnel"]
for i, item in enumerate(op_items):
    ws.cell(140, 3+i, item).font = ftn
for i, (cat, vals) in enumerate(INFRA_DETAIL.items()):
    ws.cell(141, 3+i).value = f"=INDEX($C$96:$G$96,1,{IDX})*{vals[2]}/{INFRA_COST[2]}/1000000"  # proportion of infra for 10K
    ws.cell(141, 3+i).font = ftn
# Actually simpler: just put the values directly indexed
for i in range(len(INFRA_DETAIL)):
    cat = list(INFRA_DETAIL.keys())[i]
    detail_vals = list(INFRA_DETAIL.values())[i]
    # Store all 5 scenario values and INDEX them
    for j in range(NS):
        ws.cell(145+i, 3+j, detail_vals[j]/1000000).font = ftn
    ws.cell(141, 3+i).value = f"=INDEX($C${145+i}:$G${145+i},1,{IDX})"; ws.cell(141, 3+i).font = ftn
# Personnel in millions
for j in range(NS): ws.cell(148, 3+j, PERS_COST[j]/1000000).font = ftn
ws.cell(141, 3+len(INFRA_DETAIL)).value = f"=INDEX($C$148:$G$148,1,{IDX})"; ws.cell(141, 3+len(INFRA_DETAIL)).font = ftn

ch5 = BarChart(); ch5.type = "bar"; ch5.grouping = "clustered"
ch5.title = "Operating Costs (Selected Scenario)"
ch5.x_axis.numFmt = '0.0'; ch5.x_axis.title = "N$ Millions"
cats5 = Reference(ws, min_col=3, min_row=140, max_col=6+len(INFRA_DETAIL), max_row=140)
vals5 = Reference(ws, min_col=3, min_row=141, max_col=6+len(INFRA_DETAIL), max_row=141)
ch5.add_data(vals5, from_rows=True); ch5.set_categories(cats5)
ch5.series[0].title = SeriesLabel(v="Cost"); ch5.series[0].graphicalProperties.solidFill = "ED7D31"
ch5.width = 14; ch5.height = 12; ch5.legend = None
ws.add_chart(ch5, "F29")

# ── CHART 6: Scenario Comparison ──
ws.row_dimensions[42].height = 22
ws.merge_cells('B42:O42'); ws['B42'] = "SCENARIO COMPARISON"; ws['B42'].font = fsec

# Scenario table
r = 43
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = fwb; ws.cell(r, 2).fill = tbl_h
for c in [2,3,4]: ws.cell(r, c).fill = tbl_h; ws.cell(r, c).border = db
sc_cols = [6,8,10,12,14]
for i, label in enumerate(SC_SHORT):
    ws.cell(r, sc_cols[i]).value = label; ws.cell(r, sc_cols[i]).font = fwb
    ws.cell(r, sc_cols[i]).fill = tbl_h; ws.cell(r, sc_cols[i]).border = db
    ws.cell(r, sc_cols[i]).alignment = Alignment(horizontal="center")

for ri, (metric, row_src, fmt, positive) in enumerate([
    ("5-Year Revenue", 101, '"N$"#,##0.0,,"M"', True),
    ("5-Year Costs", 102, '"N$"#,##0.0,,"M"', False),
    ("Net Profit", 103, '"N$"#,##0.0,,"M"', True),
    ("Gross Margin", 104, '0.0%', True),
    ("Payback", 107, '0" days"', True),
    ("ROI", None, '0.0%', True),
]):
    r = 44 + ri
    row_fill = tbl1 if ri % 2 == 0 else tbl2
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = metric; ws.cell(r, 2).font = fw
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db
    for i, sc in enumerate(sc_cols):
        c_letter = get_column_letter(3+i)
        if metric == "ROI":
            ws.cell(r, sc).value = f"=IF({c_letter}102>0,{c_letter}103/{c_letter}102,0)"
        else:
            ws.cell(r, sc).value = f"={c_letter}{row_src}"
        ws.cell(r, sc).font = fg if positive else frd
        ws.cell(r, sc).number_format = fmt; ws.cell(r, sc).fill = row_fill; ws.cell(r, sc).border = db
        ws.cell(r, sc).alignment = Alignment(horizontal="center")
        if i == 2:
            ws.cell(r, sc).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.cell(r, sc).font = Font(bold=True, size=10, name="Calibri", color=ACC_GREEN if positive else ACC_RED)

# ── CHART 7: Scenario bar chart ──
ws.row_dimensions[51].height = 22
ws.merge_cells('B51:O51'); ws['B51'] = "SCENARIO ANALYSIS CHARTS"; ws['B51'].font = fsec

ch7 = BarChart(); ch7.type = "col"; ch7.grouping = "clustered"
ch7.title = "5-Year Revenue, Costs & Profit by Scenario"
ch7.y_axis.numFmt = '0'; ch7.y_axis.title = "N$ Millions"
cats7 = Reference(ws, min_col=3, min_row=130, max_col=7, max_row=130)
for rr, label, color in [(131, "Revenue", "3B82F6"), (132, "Costs", "EF4444"), (133, "Profit", "22C55E")]:
    vals = Reference(ws, min_col=3, min_row=rr, max_col=7, max_row=rr)
    ch7.add_data(vals, from_rows=True)
    ch7.series[-1].title = SeriesLabel(v=label); ch7.series[-1].graphicalProperties.solidFill = color
ch7.set_categories(cats7); ch7.legend.position = 'b'
ch7.width = 18; ch7.height = 12
ws.add_chart(ch7, "B52")

# ── CHART 8: Headcount by scenario ──
for i in range(NS): ws.cell(135, 3+i, HEADCOUNTS[i]).font = ftn
ch8 = BarChart(); ch8.type = "col"; ch8.grouping = "clustered"
ch8.title = "Personnel Headcount by Scenario"
ch8.y_axis.title = "Employees"
cats8 = Reference(ws, min_col=3, min_row=130, max_col=7, max_row=130)
vals8 = Reference(ws, min_col=3, min_row=135, max_col=7, max_row=135)
ch8.add_data(vals8, from_rows=True); ch8.set_categories(cats8)
ch8.series[0].title = SeriesLabel(v="Headcount"); ch8.series[0].graphicalProperties.solidFill = "70AD47"
ch8.width = 14; ch8.height = 12; ch8.legend = None
ws.add_chart(ch8, "L52")

# Footer
ws.merge_cells('B66:O66')
ws['B66'] = "GRIDx Smart Energy Solutions  |  Confidential Financial Model  |  All figures in Namibian Dollars (N$)"
ws['B66'].font = Font(italic=True, size=8, color=LGRAY, name="Calibri")
ws['B66'].alignment = Alignment(horizontal="center")

# Data validation
dv = DataValidation(type="list", formula1="=$C$90:$G$90", allow_blank=False)
dv.prompt = "Select scenario"; dv.promptTitle = "Rollout"
ws.add_data_validation(dv); dv.add(ws['N6'])

# ── Finalize ──
wb.move_sheet("Dashboard", offset=-5)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v11.xlsx"
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print("v11: Client's v6 data + v10 dark dashboard with all charts")
