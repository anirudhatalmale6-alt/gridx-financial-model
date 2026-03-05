"""
GRIDx Financial Dashboard v9 — Dark Navy Dashboard + Detailed Derivation Tabs
Tabs: Dashboard, Revenue, Manufacturing, Personnel, Operating Costs,
      Implementation, Equipment, Development, Financing, Profitability
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Constants ──
SCENARIOS = ["3,000 Meters","5,000 Meters","10,000 Meters","20,000 Meters","50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]; NS = 5
HW=5500; INST=700; MAINT=300; WIFI=90; SMS=30; APP=624; RE=76.80
ANN_REC = HW+MAINT+WIFI+SMS+APP+RE  # 6620.80
Y1_PM = ANN_REC + INST  # 7320.80
MFG_U = [2000, 2000, 1700, 1700, 1462]
SP_R = 4495 * 0.02
VH=[2,2,3,5,10]; TK=[2,3,4,6,12]; WK=[3,5,7,12,25]; NT=[1,1,2,3,5]
CLIENT_PB = [7.6, 7.2, 4.2, 3.9, 3.5]
CLIENT_DCF = [65500000, 119300000, 259647182, 538200000, 1390000000]

# Personnel data (from v4)
POSITIONS = [
    ("CEO", 35000, [1,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("COO", 28000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("CFO", 28000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("CTO", 30000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of Business Dev", 25000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of HR", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Legal & Compliance", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Operations Manager", 20000, [0,1,1,1,1], "OPERATIONS & FIELD"),
    ("Field Support Techs", 9000, [2,3,4,5,8], "OPERATIONS & FIELD"),
    ("Installation Team Lead", 11000, [0,1,1,2,3], "OPERATIONS & FIELD"),
    ("Logistics Coordinator", 8000, [0,0,1,1,2], "OPERATIONS & FIELD"),
    ("Engineering Manager", 22000, [0,0,1,1,1], "ENGINEERING & QA"),
    ("QA Engineers", 12000, [1,1,1,3,5], "ENGINEERING & QA"),
    ("Firmware/Software Eng", 15000, [0,1,2,3,5], "ENGINEERING & QA"),
    ("Hardware Engineers", 15000, [0,0,1,2,3], "ENGINEERING & QA"),
    ("Customer Support Mgr", 18000, [0,0,1,1,1], "CUSTOMER SUCCESS"),
    ("Customer Support Agents", 7200, [1,2,2,4,8], "CUSTOMER SUCCESS"),
    ("Technical Support", 9000, [0,0,1,2,3], "CUSTOMER SUCCESS"),
    ("Sales Manager", 20000, [0,0,1,1,1], "SALES & MARKETING"),
    ("Sales Representatives", 10000, [0,1,2,3,5], "SALES & MARKETING"),
    ("Marketing Coordinator", 9000, [0,0,1,1,2], "SALES & MARKETING"),
    ("Finance Manager", 20000, [0,0,1,1,1], "FINANCE & ADMIN"),
    ("Accountant", 11000, [0,1,1,2,3], "FINANCE & ADMIN"),
    ("Admin Assistant", 7000, [0,0,1,2,3], "FINANCE & ADMIN"),
]

# ── Styles ──
NV = "1B2A4A"
hdr_fill = PatternFill("solid", fgColor=NV)
grn_fill = PatternFill("solid", fgColor="E2EFDA")
org_fill = PatternFill("solid", fgColor="FCE4D6")
ylw_fill = PatternFill("solid", fgColor="FFF2CC")
lt_fill  = PatternFill("solid", fgColor="D6E4F0")
hdr_f = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bf    = Font(bold=True, size=10, name="Calibri")
b11   = Font(bold=True, size=11, name="Calibri")
lbl   = Font(size=10, name="Calibri")
inp   = Font(color="0000FF", size=10, name="Calibri")
ttl   = Font(bold=True, color=NV, size=14, name="Calibri")
nb12  = Font(bold=True, color=NV, size=12, name="Calibri")
dept_f = Font(bold=True, color="4472C4", size=10, name="Calibri")
note_f = Font(italic=True, size=10, color="666666", name="Calibri")
warn_f = Font(italic=True, size=10, color="C00000", name="Calibri")
tb = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
            top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
tot_b = Border(top=Side("medium", color=NV), bottom=Side("double", color=NV))
NUM = '#,##0'; CURR = 'N$#,##0'; PCT = '0.0%'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(r, cs+i, l); c.font = hdr_f; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = tb

def wl(ws, r, c, t, f=lbl, fi=None):
    cell = ws.cell(r, c, t); cell.font = f; cell.border = tb
    if fi: cell.fill = fi
    return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(r, c, v); cell.font = inp; cell.fill = ylw_fill
    cell.number_format = fmt; cell.border = tb; return cell

def wf(ws, r, c, formula, fmt=NUM, f=lbl, fi=None):
    cell = ws.cell(r, c, formula); cell.font = f; cell.number_format = fmt; cell.border = tb
    if fi: cell.fill = fi
    return cell

def fr(ws, r, cs, ce, fi):
    for c in range(cs, ce+1): ws.cell(r, c).fill = fi; ws.cell(r, c).border = tb

def scols(ws, lw=42):
    ws.column_dimensions['A'].width = lw
    for i in range(NS): ws.column_dimensions[get_column_letter(2+i)].width = 18

# ═══════════════════════════════════════════════════════════════
# SHEET 1: REVENUE
# ═══════════════════════════════════════════════════════════════
ws_r = wb.active; ws_r.title = "Revenue"; ws_r.sheet_properties.tabColor = "ED7D31"
ws_r.column_dimensions['A'].width = 40
for c in ['B','C','D','E']: ws_r.column_dimensions[c].width = 22

wl(ws_r, 1, 1, "Revenue Model — Assumptions & Derivation", f=ttl)
wh(ws_r, 3, ["Revenue Stream", "Rate", "Adoption", "Basis", "Annual/Meter (N$)"])
for r, nm, rate, ad, bas, val in [
    (4, "Meter Hardware Sale", "N$5,500/yr", "100%", "Annual recurring", HW),
    (5, "Installation Fee", "N$700 one-time", "100%", "One-time", INST),
    (6, "Maintenance", "N$300/yr", "100%", "Annual", MAINT),
    (7, "Wi-Fi Subscription", "N$25/mo", "30%", "30% x 25 x 12", WIFI),
    (8, "SMS Notifications", "N$2.50/SMS", "50%", "50% x 2 x 2.50 x 12", SMS),
    (9, "Mobile App", "N$65/mo", "80%", "80% x 65 x 12", APP),
    (10, "Real Estate Mgmt", "N$3,200/blk/mo", "1/500m", "(m/500) x 3200 x 12 / m", RE),
]:
    wl(ws_r, r, 1, nm); wl(ws_r, r, 2, rate); wl(ws_r, r, 3, ad)
    wl(ws_r, r, 4, bas); wi(ws_r, r, 5, val, 'N$#,##0.00')

wl(ws_r, 11, 1, "Annual Recurring per Meter", f=bf, fi=grn_fill); fr(ws_r, 11, 2, 5, grn_fill)
wf(ws_r, 11, 5, "=E4+E6+E7+E8+E9+E10", 'N$#,##0.00', bf, grn_fill).border = tot_b
wl(ws_r, 12, 1, "Year 1 Total per Meter (incl Install)", f=bf, fi=grn_fill); fr(ws_r, 12, 2, 5, grn_fill)
wf(ws_r, 12, 5, "=E11+E5", 'N$#,##0.00', bf, grn_fill).border = tot_b

# Revenue by Year
wl(ws_r, 14, 1, "Revenue by Year & Scenario", f=ttl)
wh(ws_r, 16, ["Year"] + SCENARIOS)
wl(ws_r, 17, 1, "Year 1")
for i in range(NS): wi(ws_r, 17, 2+i, round(MC[i] * Y1_PM), CURR)
for yr in range(2, 6):
    wl(ws_r, 15+yr, 1, f"Year {yr}")
    for i in range(NS): wi(ws_r, 15+yr, 2+i, round(MC[i] * ANN_REC), CURR)
wl(ws_r, 22, 1, "5-Year Total", f=b11, fi=grn_fill); fr(ws_r, 22, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_r, 22, 2+i, f"=SUM({col}17:{col}21)", CURR, b11, grn_fill).border = tot_b

# 5-Year by Stream
wl(ws_r, 24, 1, "5-Year Revenue Breakdown by Stream", f=ttl)
wh(ws_r, 26, ["Revenue Stream"] + SCENARIOS)
for rr, nm, rate, yrs in [
    (27, "Hardware Sales (N$5,500/yr x 5)", HW, 5),
    (28, "Installation Fees (N$700 one-time)", INST, 1),
    (29, "Maintenance (N$300/yr x 5)", MAINT, 5),
    (30, "Wi-Fi (N$90/yr x 5)", WIFI, 5),
    (31, "SMS (N$30/yr x 5)", SMS, 5),
    (32, "Mobile App (N$624/yr x 5)", APP, 5),
    (33, "Real Estate (N$76.80/yr x 5)", RE, 5),
]:
    wl(ws_r, rr, 1, nm)
    for i in range(NS): wi(ws_r, rr, 2+i, round(MC[i] * rate * yrs), CURR)

wl(ws_r, 34, 1, "Total 5-Year Revenue", f=b11, fi=grn_fill); fr(ws_r, 34, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_r, 34, 2+i, f"=SUM({col}27:{col}33)", CURR, b11, grn_fill).border = tot_b
REV_TOT = 34
ws_r.freeze_panes = "B4"
print(f"Revenue: REV_TOT={REV_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: MANUFACTURING
# ═══════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("Manufacturing"); ws_m.sheet_properties.tabColor = "4472C4"; scols(ws_m)
wl(ws_m, 1, 1, "Manufacturing Costs — Derivation", f=ttl)

# Volume discounts
wl(ws_m, 3, 1, "VOLUME DISCOUNT STRUCTURE", f=b11, fi=lt_fill); fr(ws_m, 3, 2, 5, lt_fill)
wh(ws_m, 5, ["Volume Tier", "Discount", "Cost/Meter (N$)", "Calculation Basis"])
for r, tier, disc, cost, bas in [
    (6, "Base (0-9,999)", "0%", 2000, "N$2,000 per meter"),
    (7, "10,000 - 19,999", "15%", 1700, "15% reduction"),
    (8, "20,000 - 49,999", "30%", 1400, "30% reduction"),
    (9, "50,000+", "~27%", 1462, "Adjusted to N$1,462"),
]:
    wl(ws_m, r, 1, tier); wl(ws_m, r, 2, disc); wi(ws_m, r, 3, cost, CURR); wl(ws_m, r, 4, bas)

# Cost by scenario
wl(ws_m, 11, 1, "MANUFACTURING COST BY SCENARIO", f=b11, fi=org_fill); fr(ws_m, 11, 2, 6, org_fill)
wh(ws_m, 13, ["Cost Item"] + SCENARIOS)
wl(ws_m, 14, 1, "Meter Hardware (Volume x Unit Cost)")
for i in range(NS): wi(ws_m, 14, 2+i, MC[i] * MFG_U[i], CURR)
wl(ws_m, 15, 1, "Spare Parts (2% x N$4,495/meter)")
for i in range(NS): wi(ws_m, 15, 2+i, round(MC[i] * SP_R), CURR)
wl(ws_m, 16, 1, "TOTAL MANUFACTURING", f=b11, fi=grn_fill); fr(ws_m, 16, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_m, 16, 2+i, f"=SUM({col}14:{col}15)", CURR, b11, grn_fill).border = tot_b
MFG_TOT = 16

# Unit summary
wl(ws_m, 18, 1, "UNIT COST SUMMARY", f=b11, fi=lt_fill); fr(ws_m, 18, 2, 6, lt_fill)
wh(ws_m, 19, ["Metric"] + SCENARIOS)
wl(ws_m, 20, 1, "Meters Deployed")
for i in range(NS): wi(ws_m, 20, 2+i, MC[i], NUM)
wl(ws_m, 21, 1, "Cost per Meter (N$)")
for i in range(NS): wi(ws_m, 21, 2+i, MFG_U[i], CURR)
wl(ws_m, 22, 1, "Discount Applied")
for i in range(NS): wl(ws_m, 22, 2+i, f"{round((1 - MFG_U[i]/2000) * 100)}%")
ws_m.freeze_panes = "B4"
print(f"Manufacturing: MFG_TOT={MFG_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: PERSONNEL (Reference from v4)
# ═══════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Personnel"); ws_p.sheet_properties.tabColor = "70AD47"
ws_p.column_dimensions['A'].width = 30
for i in range(NS): ws_p.column_dimensions[get_column_letter(2+i)].width = 14
ws_p.column_dimensions['H'].width = 30; ws_p.column_dimensions['I'].width = 14; ws_p.column_dimensions['J'].width = 14

wl(ws_p, 1, 1, "Personnel — Detailed Staffing Plan", f=ttl)
wl(ws_p, 2, 1, "Headcount by position per deployment scenario", f=note_f)

# Headcount table
wh(ws_p, 4, ["Position"] + SCENARIOS)
wh(ws_p, 4, ["Position", "Monthly (N$)", "Annual (N$)"], cs=8)

r = 5; prev_dept = None
for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 6, lt_fill)
        prev_dept = dept; r += 1
    wl(ws_p, r, 1, nm)
    for i in range(NS): wi(ws_p, r, 2+i, hc[i], NUM)
    wl(ws_p, r, 8, nm); wi(ws_p, r, 9, monthly, CURR)
    wf(ws_p, r, 10, f"=I{r}*12", CURR)
    r += 1

# Total headcount
wl(ws_p, r, 1, "TOTAL HEADCOUNT", f=b11, fi=grn_fill); fr(ws_p, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_p, r, 2+i, f"=SUMPRODUCT(({col}5:{col}{r-1})*({col}5:{col}{r-1}<>0)*1)", NUM, b11, grn_fill)
HC_TOT_ROW = r

# Annual cost section
r += 2
wl(ws_p, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=ttl)
r += 1
wh(ws_p, r, ["Position", "Annual Salary"] + SCENARIOS)
cost_hdr = r; r += 1; cost_start = r; prev_dept = None

for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 7, lt_fill)
        prev_dept = dept; r += 1
    annual = monthly * 12
    wl(ws_p, r, 1, nm); wi(ws_p, r, 2, annual, CURR)
    for i in range(NS): wi(ws_p, r, 3+i, hc[i] * annual, CURR)
    r += 1

wl(ws_p, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=b11, fi=grn_fill); fr(ws_p, r, 2, 7, grn_fill)
wl(ws_p, r, 2, "", f=b11, fi=grn_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    wf(ws_p, r, 3+i, f"=SUMPRODUCT(({col}{cost_start}:{col}{r-1})*({col}{cost_start}:{col}{r-1}<>0)*1)", CURR, b11, grn_fill).border = tot_b
PER_TOT = r

r += 2
wl(ws_p, r, 1, "Note: The financial model uses a consolidated annual personnel figure of N$748,800/yr.", f=warn_f)
wl(ws_p, r+1, 1, "This sheet shows the recommended full-scale staffing plan for reference.", f=note_f)
ws_p.freeze_panes = "B5"
print(f"Personnel: PER_TOT={PER_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 4: OPERATING COSTS
# ═══════════════════════════════════════════════════════════════
ws_o = wb.create_sheet("Operating Costs"); ws_o.sheet_properties.tabColor = "ED7D31"; scols(ws_o)
wl(ws_o, 1, 1, "Operating Costs — 5-Year Total", f=ttl)
wl(ws_o, 2, 1, "Annual recurring costs x 5 years", f=note_f)
wh(ws_o, 4, ["Cost Item"] + SCENARIOS)

oper_items = [
    ("Personnel (N$748,800/yr x 5)", [748800*5]*5),
    ("Data Storage (N$48/meter/yr x 5)", [m*48*5 for m in MC]),
    ("Vehicle Maintenance", [v*30000*5 for v in VH]),
    ("Field Equipment Upkeep", [k*10000*5 for k in TK]),
    ("Office & Utilities (N$240,000/yr x 5)", [240000*5]*5),
    ("Software & Tools (N$120,000/yr x 5)", [120000*5]*5),
    ("Marketing & Branding (N$200,000/yr x 5)", [200000*5]*5),
    ("Regulatory Compliance (N$100,000/yr x 5)", [100000*5]*5),
    ("Insurance & Risk (N$150,000/yr x 5)", [150000*5]*5),
]
r = 5
for nm, vals in oper_items:
    wl(ws_o, r, 1, nm)
    for i in range(NS): wi(ws_o, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_o, r, 1, "TOTAL OPERATING (5-Year)", f=b11, fi=grn_fill); fr(ws_o, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_o, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
OPER_TOT = r
ws_o.freeze_panes = "B5"
print(f"Operating: OPER_TOT={OPER_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 5: IMPLEMENTATION
# ═══════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("Implementation"); ws_i.sheet_properties.tabColor = "FFC000"; scols(ws_i)
wl(ws_i, 1, 1, "Implementation Costs — One-Time", f=ttl)
wl(ws_i, 2, 1, "Deployment and setup costs (Year 1 only)", f=note_f)
wh(ws_i, 4, ["Cost Item"] + SCENARIOS)

impl_items = [
    ("Project Management", [500000]*5),
    ("Site Surveys & Assessment", [300000]*5),
    ("Onboarding & Training", [250000]*5),
    ("System Integration", [450000]*5),
]
r = 5
for nm, vals in impl_items:
    wl(ws_i, r, 1, nm)
    for i in range(NS): wi(ws_i, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_i, r, 1, "TOTAL IMPLEMENTATION", f=b11, fi=grn_fill); fr(ws_i, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_i, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
IMPL_TOT = r
ws_i.freeze_panes = "B5"
print(f"Implementation: IMPL_TOT={IMPL_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 6: EQUIPMENT
# ═══════════════════════════════════════════════════════════════
ws_e = wb.create_sheet("Equipment"); ws_e.sheet_properties.tabColor = "4472C4"; scols(ws_e)
wl(ws_e, 1, 1, "Equipment Investment — One-Time", f=ttl)
wl(ws_e, 2, 1, "Capital equipment purchases (Year 1 only)", f=note_f)
wh(ws_e, 4, ["Cost Item"] + SCENARIOS)

equip_items = [
    ("Service Vehicles (N$250,000 each)", [250000*v for v in VH]),
    ("Test Equipment Kits (N$85,000 each)", [85000*k for k in TK]),
    ("Workstations (N$15,000 each)", [15000*w for w in WK]),
    ("Network Monitoring (N$120,000 each)", [120000*n for n in NT]),
    ("Office Furniture", [50000 + round(m/10000*230000) for m in MC]),
    ("Server/Cloud Infrastructure", [100000 + round(m/10000*150000) for m in MC]),
]
r = 5
for nm, vals in equip_items:
    wl(ws_e, r, 1, nm)
    for i in range(NS): wi(ws_e, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_e, r, 1, "TOTAL EQUIPMENT", f=b11, fi=grn_fill); fr(ws_e, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_e, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
EQUIP_TOT = r

# Quantity summary
r += 2
wl(ws_e, r, 1, "EQUIPMENT QUANTITIES", f=b11, fi=lt_fill); fr(ws_e, r, 2, 6, lt_fill)
wh(ws_e, r+1, ["Item"] + SCENARIOS)
for rr, nm, cts in [
    (r+2, "Service Vehicles", VH), (r+3, "Test Equipment Kits", TK),
    (r+4, "Workstations", WK), (r+5, "Network Monitoring Tools", NT),
]:
    wl(ws_e, rr, 1, nm)
    for i in range(NS): wi(ws_e, rr, 2+i, cts[i], NUM)
ws_e.freeze_panes = "B5"
print(f"Equipment: EQUIP_TOT={EQUIP_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 7: DEVELOPMENT
# ═══════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("Development"); ws_d.sheet_properties.tabColor = "7030A0"; scols(ws_d)
wl(ws_d, 1, 1, "Development Costs — Pre-Launch", f=ttl)
wl(ws_d, 2, 1, "R&D and product development (one-time investment)", f=note_f)
wh(ws_d, 4, ["Cost Item"] + SCENARIOS)

dev_items = [
    ("Electronics Manufacturing Dev", [312000]*5),
    ("Software R&D", [777747]*5),
    ("Mobile App Development", [300000]*5),
    ("Firmware Development", [2160000]*5),
    ("Enclosure Manufacturing", [1403850]*5),
    ("Marketing/Branding Launch", [325000]*5),
    ("Regulatory & Legal", [1060000]*5),
    ("Office/Logistics Setup", [897400]*5),
    ("Development Contingency (5%)", [615650]*5),
]
r = 5
for nm, vals in dev_items:
    wl(ws_d, r, 1, nm)
    for i in range(NS): wi(ws_d, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_d, r, 1, "TOTAL DEVELOPMENT", f=b11, fi=grn_fill); fr(ws_d, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_d, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
DEV_TOT = r
ws_d.freeze_panes = "B5"
print(f"Development: DEV_TOT={DEV_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 8: FINANCING & CONTINGENCY
# ═══════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("Financing"); ws_f.sheet_properties.tabColor = "C00000"; scols(ws_f)
wl(ws_f, 1, 1, "Financing & Contingency", f=ttl)

# Financing
wl(ws_f, 3, 1, "FINANCING COSTS", f=b11, fi=org_fill); fr(ws_f, 3, 2, 6, org_fill)
wh(ws_f, 4, ["Cost Item"] + SCENARIOS)
wl(ws_f, 5, 1, "Interest on Working Capital")
for i in range(NS): wi(ws_f, 5, 2+i, 500000, CURR)
wl(ws_f, 6, 1, "Bank Fees & Charges")
for i in range(NS): wi(ws_f, 6, 2+i, 200000, CURR)
wl(ws_f, 7, 1, "TOTAL FINANCING", f=b11, fi=grn_fill); fr(ws_f, 7, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 7, 2+i, f"=SUM({col}5:{col}6)", CURR, b11, grn_fill).border = tot_b
FIN_TOT = 7

# Contingency
wl(ws_f, 9, 1, "CONTINGENCY RESERVES", f=b11, fi=org_fill); fr(ws_f, 9, 2, 6, org_fill)
wh(ws_f, 10, ["Cost Item"] + SCENARIOS)
wl(ws_f, 11, 1, "Operating Contingency (10% of 5Y Operating)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 11, 2+i, f"='Operating Costs'!{col}{OPER_TOT}*0.1", CURR)
wl(ws_f, 12, 1, "Implementation Contingency (10%)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 12, 2+i, f"=Implementation!{col}{IMPL_TOT}*0.1", CURR)
wl(ws_f, 13, 1, "TOTAL CONTINGENCY", f=b11, fi=grn_fill); fr(ws_f, 13, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 13, 2+i, f"=SUM({col}11:{col}12)", CURR, b11, grn_fill).border = tot_b
CONT_TOT = 13
ws_f.freeze_panes = "B4"
print(f"Financing: FIN_TOT={FIN_TOT}, CONT_TOT={CONT_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 9: PROFITABILITY ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws_pa = wb.create_sheet("Profitability"); ws_pa.sheet_properties.tabColor = "002060"; scols(ws_pa)
wl(ws_pa, 1, 1, "Profitability Analysis — All Scenarios", f=ttl)

# Year 1 Analysis
wl(ws_pa, 3, 1, "YEAR 1 ANALYSIS", f=b11, fi=lt_fill); fr(ws_pa, 3, 2, 6, lt_fill)
wh(ws_pa, 5, ["Metric"] + SCENARIOS)

# Year 1 Revenue
wl(ws_pa, 6, 1, "Year 1 Revenue", f=bf, fi=grn_fill); fr(ws_pa, 6, 2, 6, grn_fill)
for i in range(NS): wi(ws_pa, 6, 2+i, round(MC[i] * Y1_PM), CURR)

wl(ws_pa, 7, 1, "YEAR 1 COSTS:", f=bf)
# Year 1 costs (one-time + 1yr operating)
y1_cost_refs = [
    (8, "Manufacturing (one-time)", "Manufacturing", MFG_TOT, 1),
    (9, "Operating (1 year)", "'Operating Costs'", OPER_TOT, 0.2),
    (10, "Implementation (one-time)", "Implementation", IMPL_TOT, 1),
    (11, "Equipment (one-time)", "Equipment", EQUIP_TOT, 1),
    (12, "Development (one-time)", "Development", DEV_TOT, 1),
    (13, "Financing", "Financing", FIN_TOT, 1),
    (14, "Contingency (1 year)", "Financing", CONT_TOT, 0.2),
]
for rr, label, sheet, tot_row, mult in y1_cost_refs:
    wl(ws_pa, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        if mult == 1:
            wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}", CURR)
        else:
            wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}*{mult}", CURR)

wl(ws_pa, 15, 1, "Total Year 1 Costs", f=b11, fi=org_fill); fr(ws_pa, 15, 2, 6, org_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 15, 2+i, f"=SUM({col}8:{col}14)", CURR, b11, org_fill).border = tot_b

wl(ws_pa, 16, 1, "Year 1 Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 16, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 16, 2+i, f"={col}6-{col}15", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 17, 1, "Year 1 Net Margin")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 17, 2+i, f"=IF({col}6>0,{col}16/{col}6,0)", PCT)

# 5-Year Projections
wl(ws_pa, 19, 1, "5-YEAR PROJECTIONS", f=b11, fi=lt_fill); fr(ws_pa, 19, 2, 6, lt_fill)
wh(ws_pa, 21, ["Metric"] + SCENARIOS)

wl(ws_pa, 22, 1, "5-Year Total Revenue")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 22, 2+i, f"=Revenue!{col}{REV_TOT}", CURR)

# 5-Year costs = sum of all cost sheet totals
cost_sheet_refs = [
    ("Manufacturing", MFG_TOT), ("'Operating Costs'", OPER_TOT),
    ("Implementation", IMPL_TOT), ("Equipment", EQUIP_TOT),
    ("Development", DEV_TOT), ("Financing", FIN_TOT), ("Financing", CONT_TOT),
]
wl(ws_pa, 23, 1, "5-Year Total Costs")
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{sn}!{col}{sr}" for sn, sr in cost_sheet_refs])
    wf(ws_pa, 23, 2+i, f"={refs}", CURR)

wl(ws_pa, 24, 1, "5-Year Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 24, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 24, 2+i, f"={col}22-{col}23", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 25, 1, "5-Year ROI %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 25, 2+i, f"=IF({col}23>0,{col}24/{col}23,0)", PCT)

wl(ws_pa, 26, 1, "5-Year Net Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 26, 2+i, f"=IF({col}22>0,{col}24/{col}22,0)", PCT)

wl(ws_pa, 27, 1, "Payback Period (months)")
for i in range(NS): wi(ws_pa, 27, 2+i, CLIENT_PB[i], '0.0')

wl(ws_pa, 28, 1, "DCF / Net Present Value")
for i in range(NS): wi(ws_pa, 28, 2+i, CLIENT_DCF[i], CURR)

# Cost breakdown
wl(ws_pa, 30, 1, "5-YEAR COST BREAKDOWN", f=b11, fi=lt_fill); fr(ws_pa, 30, 2, 6, lt_fill)
wh(ws_pa, 32, ["Cost Category"] + SCENARIOS)
for rr, label, sheet, tot_row in [
    (33, "Manufacturing", "Manufacturing", MFG_TOT),
    (34, "Operating Costs (5Y)", "'Operating Costs'", OPER_TOT),
    (35, "Implementation", "Implementation", IMPL_TOT),
    (36, "Equipment", "Equipment", EQUIP_TOT),
    (37, "Development", "Development", DEV_TOT),
    (38, "Financing", "Financing", FIN_TOT),
    (39, "Contingency", "Financing", CONT_TOT),
]:
    wl(ws_pa, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}", CURR)

wl(ws_pa, 40, 1, "TOTAL 5-YEAR COSTS", f=b11, fi=grn_fill); fr(ws_pa, 40, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 40, 2+i, f"=SUM({col}33:{col}39)", CURR, b11, grn_fill).border = tot_b

ws_pa.freeze_panes = "B5"
print(f"Profitability: complete")

# ═══════════════════════════════════════════════════════════════
# SHEET 10: DASHBOARD (DARK NAVY THEME)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "002060"

# Dark theme colors
BG1="0F1B2D"; BG2="162236"; BG3="1D2B42"; BG4="243B55"
ACC_BLUE="3B82F6"; ACC_GREEN="22C55E"; ACC_RED="EF4444"; GOLD="F59E0B"
WHITE="FFFFFF"; LGRAY="94A3B8"; BDR_C="2D3F59"

bg1=PatternFill("solid",fgColor=BG1); bg2=PatternFill("solid",fgColor=BG2)
bg3=PatternFill("solid",fgColor=BG3); bg4=PatternFill("solid",fgColor=BG4)
acc_bl=PatternFill("solid",fgColor=ACC_BLUE); kpi_bg=PatternFill("solid",fgColor="1E3A5F")
tbl_h=PatternFill("solid",fgColor=BG4); tbl1=PatternFill("solid",fgColor=BG2); tbl2=PatternFill("solid",fgColor=BG3)

ft=Font(bold=True,size=16,name="Calibri",color=WHITE)
fs=Font(bold=True,size=11,name="Calibri",color=LGRAY)
fsec=Font(bold=True,size=11,name="Calibri",color=ACC_BLUE)
fw=Font(size=10,name="Calibri",color=WHITE)
fwb=Font(bold=True,size=10,name="Calibri",color=WHITE)
fdr=Font(bold=True,size=12,name="Calibri",color=WHITE)
fkl=Font(size=9,name="Calibri",color=LGRAY)
fg=Font(bold=True,size=10,name="Calibri",color=ACC_GREEN)
frd=Font(size=10,name="Calibri",color=ACC_RED)
fgo=Font(bold=True,size=10,name="Calibri",color=GOLD)
ftn=Font(size=7,name="Calibri",color=BG2)
db=Border(left=Side("thin",color=BDR_C),right=Side("thin",color=BDR_C),
          top=Side("thin",color=BDR_C),bottom=Side("thin",color=BDR_C))

# Column widths
for c,w in [('A',1.5),('B',16),('C',14),('D',14),('E',1),('F',12),('G',12),
            ('H',12),('I',12),('J',12),('K',1),('L',12),('M',12),('N',12),('O',12),('P',1.5)]:
    ws.column_dimensions[c].width = w

# Fill background
for r in range(1, 50):
    ws.row_dimensions[r].height = 18
    for c in range(1, 17): ws.cell(r, c).fill = bg1

# Title bar
ws.row_dimensions[2].height = 32
ws.merge_cells('B2:O2')
ws['B2'] = "GRIDx Smart Meter - 5-Year Financial Model"
ws['B2'].font = ft; ws['B2'].fill = bg2
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 16): ws.cell(2, c).fill = bg2

# Scenario selector
ws.row_dimensions[3].height = 28
ws.merge_cells('F3:I3')
ws['F3'] = "10,000 Meters"; ws['F3'].font = fdr; ws['F3'].fill = acc_bl
ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(6, 10): ws.cell(3, c).fill = acc_bl; ws.cell(3, c).border = db
ws.merge_cells('L3:O3')
ws['L3'] = "Recommended"
ws['L3'].font = Font(italic=True, size=10, name="Calibri", color=GOLD)
ws['L3'].alignment = Alignment(horizontal="center", vertical="center")

# Input Parameters
ws.merge_cells('B5:D5')
ws['B5'] = "1. Input Parameters"; ws['B5'].font = fsec
ws['B6'] = "Rollout Scenario:"; ws['B6'].font = fw
ws.merge_cells('C6:D6')
ws['C6'] = "10,000 Meters"; ws['C6'].font = fdr
ws['C6'].fill = PatternFill("solid", fgColor="2563EB")
ws['C6'].alignment = Alignment(horizontal="center"); ws['C6'].border = db
ws['D6'].fill = PatternFill("solid", fgColor="2563EB"); ws['D6'].border = db

IDX = "$C$51"
def idx(rr): return f"INDEX($C${rr}:$G${rr},1,{IDX})"

params = [
    (8, "Hardware Price:", "N$5,500", "/unit/yr", None),
    (9, "Installation:", "N$700", "one-time", None),
    (10, "Mfg Cost/Meter:", None, "per unit", f'="N$"&TEXT({idx(53)},"#,##0")'),
    (11, "Personnel:", "N$748,800", "/year", None),
    (12, "Annual Revenue:", None, "/year", f'="N$"&TEXT({idx(52)}*{ANN_REC},"#,##0")'),
    (13, "Total 5Y Revenue:", None, "5-year", f'="N$"&TEXT({idx(63)},"#,##0")'),
    (14, "Total 5Y Costs:", None, "5-year", f'="N$"&TEXT({idx(61)},"#,##0")'),
    (15, "Net Profit:", None, "5-year", f'="N$"&TEXT({idx(62)},"#,##0")'),
]
for pr, label, val, unit, formula in params:
    ws.cell(pr, 2).value = label; ws.cell(pr, 2).font = fw
    if val: ws.cell(pr, 3).value = val; ws.cell(pr, 3).font = fg
    elif formula: ws.cell(pr, 3).value = formula; ws.cell(pr, 3).font = fg if pr != 14 else frd
    if pr == 15 and formula: ws.cell(pr, 3).font = fgo
    ws.cell(pr, 4).value = unit; ws.cell(pr, 4).font = Font(size=8, color=LGRAY, name="Calibri")

# KPI cards
kpis = [
    (5, 12, 13, "Total Revenue", f'="N$"&TEXT({idx(63)}/1000000,"#,##0")&"M"', kpi_bg, ACC_GREEN),
    (5, 14, 15, "Net Profit", f'="N$"&TEXT({idx(62)}/1000000,"#,##0")&"M"', kpi_bg, ACC_GREEN),
    (7, 12, 13, "Gross Margin", f'=TEXT({idx(64)}*100,"0.0")&"%"', kpi_bg, GOLD),
    (7, 14, 15, "Payback Period", f'=TEXT({idx(65)},"0.0")&" months"', kpi_bg, ACC_BLUE),
]
for kr, kc1, kc2, label, formula, bgc, acc in kpis:
    for r in range(kr, kr+2):
        for c in range(kc1, kc2+1): ws.cell(r, c).fill = bgc; ws.cell(r, c).border = db
    ws.merge_cells(start_row=kr, start_column=kc1, end_row=kr, end_column=kc2)
    ws.cell(kr, kc1).value = formula
    ws.cell(kr, kc1).font = Font(bold=True, size=18, name="Calibri", color=acc)
    ws.cell(kr, kc1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=kr+1, start_column=kc1, end_row=kr+1, end_column=kc2)
    ws.cell(kr+1, kc1).value = label
    ws.cell(kr+1, kc1).font = fkl
    ws.cell(kr+1, kc1).alignment = Alignment(horizontal="center")

# ── Hidden data (rows 50+) ──
for i, s in enumerate(SCENARIOS): ws.cell(50, 3+i, s).font = ftn
ws.cell(51, 3).value = '=MATCH($C$6,$C$50:$G$50,0)'; ws.cell(51, 3).font = ftn
for i, m in enumerate(MC): ws.cell(52, 3+i, m).font = ftn
for i in range(NS): ws.cell(53, 3+i, MFG_U[i]).font = ftn

# Cost totals from individual sheets
sheet_refs = [
    (54, "Manufacturing", MFG_TOT),
    (55, "'Operating Costs'", OPER_TOT),
    (56, "Implementation", IMPL_TOT),
    (57, "Equipment", EQUIP_TOT),
    (58, "Development", DEV_TOT),
    (59, "Financing", FIN_TOT),
    (60, "Financing", CONT_TOT),
]
for dr, sname, srow in sheet_refs:
    for i in range(NS):
        col = get_column_letter(2+i)
        ws.cell(dr, 3+i).value = f"={sname}!{col}{srow}"; ws.cell(dr, 3+i).font = ftn

# Grand total, profit, revenue, margin
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(61, 3+i).value = f"=SUM({c}54:{c}60)"; ws.cell(61, 3+i).font = ftn
    ws.cell(62, 3+i).value = f"={c}63-{c}61"; ws.cell(62, 3+i).font = ftn
    col = get_column_letter(2+i)
    ws.cell(63, 3+i).value = f"=Revenue!{col}{REV_TOT}"; ws.cell(63, 3+i).font = ftn
    ws.cell(64, 3+i).value = f"=IF({c}63>0,{c}62/{c}63,0)"; ws.cell(64, 3+i).font = ftn

for i in range(NS):
    ws.cell(65, 3+i, CLIENT_PB[i]).font = ftn
    ws.cell(66, 3+i, CLIENT_DCF[i]).font = ftn

# Chart data (millions)
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(70, 4+i, yr).font = ftn
ws.cell(71, 3, "RevM").font = ftn
ws.cell(71, 4).value = f"={idx(52)}*{Y1_PM}/1000000"
for yc in range(5, 9): ws.cell(71, yc).value = f"={idx(52)}*{ANN_REC}/1000000"
ws.cell(72, 3, "CostM").font = ftn
ws.cell(72, 4).value = f"=({idx(54)}+{idx(55)}/5+{idx(56)}+{idx(57)}+{idx(58)}+{idx(59)}+{idx(60)}/5)/1000000"
for yc in range(5, 9): ws.cell(72, yc).value = f"=({idx(55)}/5+{idx(60)}/5)/1000000"
ws.cell(73, 3, "ProfM").font = ftn
for yc in range(4, 9): ws.cell(73, yc).value = f"={get_column_letter(yc)}71-{get_column_letter(yc)}72"

# Bar chart
cats = Reference(ws, min_col=4, min_row=70, max_col=8, max_row=70)
bar1 = BarChart(); bar1.type = "col"; bar1.grouping = "clustered"
bar1.title = "Revenue vs. Costs (N$ Millions)"; bar1.style = 10
bar1.y_axis.numFmt = '0'; bar1.y_axis.title = "N$ Millions"
d_rev = Reference(ws, min_col=4, min_row=71, max_col=8, max_row=71)
bar1.add_data(d_rev, from_rows=True)
bar1.series[0].title = SeriesLabel(v="Revenue"); bar1.series[0].graphicalProperties.solidFill = "3B82F6"
d_cost = Reference(ws, min_col=4, min_row=72, max_col=8, max_row=72)
bar1.add_data(d_cost, from_rows=True)
bar1.series[1].title = SeriesLabel(v="Costs"); bar1.series[1].graphicalProperties.solidFill = "22C55E"
bar1.set_categories(cats); bar1.legend.position = 'b'
bar1.width = 18; bar1.height = 12
ws.add_chart(bar1, "F5")

# Financial table
r = 19
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = fwb; ws.cell(r, 2).fill = tbl_h
for c in [2,3,4]: ws.cell(r, c).fill = tbl_h; ws.cell(r, c).border = db
yr_cols = [6,7,8,9,10]
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(r, yr_cols[i]).value = yr; ws.cell(r, yr_cols[i]).font = fwb
    ws.cell(r, yr_cols[i]).fill = tbl_h; ws.cell(r, yr_cols[i]).border = db
    ws.cell(r, yr_cols[i]).alignment = Alignment(horizontal="center")
ws.merge_cells(f'L{r}:O{r}')
ws.cell(r, 12).value = "5-Year Total"; ws.cell(r, 12).font = fwb
ws.cell(r, 12).fill = tbl_h; ws.cell(r, 12).border = db
ws.cell(r, 12).alignment = Alignment(horizontal="center")
for c in [13,14,15]: ws.cell(r, c).fill = tbl_h; ws.cell(r, c).border = db

table_data = [
    ("Hardware Sales", [f"={idx(52)}*{HW}"]*5, f"={idx(52)}*{HW}*5", True, False),
    ("Installation Fees", [f"={idx(52)}*{INST}","0","0","0","0"], f"={idx(52)}*{INST}", True, False),
    ("Service Revenue", [f"={idx(52)}*{MAINT+WIFI+SMS+APP+RE}"]*5, f"={idx(52)}*{MAINT+WIFI+SMS+APP+RE}*5", True, False),
    ("Manufacturing Costs", [f"=-{idx(54)}","0","0","0","0"], f"=-{idx(54)}", False, True),
    ("Operating Costs", [f"=-{idx(55)}/5"]*5, f"=-{idx(55)}", False, True),
    ("Net Profit", None, None, True, False),
]
for ri, (label, yr_formulas, total_formula, is_pos, is_cost) in enumerate(table_data):
    r = 20 + ri
    row_fill = tbl1 if ri % 2 == 0 else tbl2
    val_f = fg if is_pos else frd
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = label; ws.cell(r, 2).font = fwb if label == "Net Profit" else fw
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db

    if label == "Net Profit":
        for ci, yc in enumerate(yr_cols):
            ws.cell(r, yc).value = f"={get_column_letter(yc)}20+{get_column_letter(yc)}21+{get_column_letter(yc)}22+{get_column_letter(yc)}23+{get_column_letter(yc)}24"
            ws.cell(r, yc).font = fgo; ws.cell(r, yc).number_format = '"N$"#,##0.0,,"M"'
            ws.cell(r, yc).fill = row_fill; ws.cell(r, yc).border = db
            ws.cell(r, yc).alignment = Alignment(horizontal="center")
        ws.merge_cells(f'L{r}:O{r}')
        ws.cell(r, 12).value = f"={idx(62)}"
        ws.cell(r, 12).font = Font(bold=True, size=12, color=GOLD, name="Calibri")
        ws.cell(r, 12).number_format = '"N$"#,##0.0,,"M"'
        ws.cell(r, 12).fill = row_fill; ws.cell(r, 12).border = db
        ws.cell(r, 12).alignment = Alignment(horizontal="center")
        for c in [13,14,15]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db
    else:
        for ci, yc in enumerate(yr_cols):
            ws.cell(r, yc).value = yr_formulas[ci]
            ws.cell(r, yc).font = val_f; ws.cell(r, yc).number_format = '"N$"#,##0.0,,"M"'
            ws.cell(r, yc).fill = row_fill; ws.cell(r, yc).border = db
            ws.cell(r, yc).alignment = Alignment(horizontal="center")
        ws.merge_cells(f'L{r}:O{r}')
        ws.cell(r, 12).value = total_formula
        ws.cell(r, 12).font = val_f; ws.cell(r, 12).number_format = '"N$"#,##0.0,,"M"'
        ws.cell(r, 12).fill = row_fill; ws.cell(r, 12).border = db
        ws.cell(r, 12).alignment = Alignment(horizontal="center")
        for c in [13,14,15]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db

# Scenario Comparison
r = 27
ws.merge_cells(f'B{r}:O{r}')
ws.cell(r, 2).value = "Scenario Comparison"; ws.cell(r, 2).font = fsec; ws.cell(r, 2).fill = bg1
for c in range(2, 16): ws.cell(r, c).fill = bg1

r = 28
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = fwb; ws.cell(r, 2).fill = tbl_h
for c in [2,3,4]: ws.cell(r, c).fill = tbl_h; ws.cell(r, c).border = db
sc_cols = [6,8,10,12,14]
for i, label in enumerate(["3K","5K","10K","20K","50K"]):
    ws.cell(r, sc_cols[i]).value = label; ws.cell(r, sc_cols[i]).font = fwb
    ws.cell(r, sc_cols[i]).fill = tbl_h; ws.cell(r, sc_cols[i]).border = db
    ws.cell(r, sc_cols[i]).alignment = Alignment(horizontal="center")

for ri, (metric, formulas, fmt, positive) in enumerate([
    ("5-Year Revenue", ["=C63","=D63","=E63","=F63","=G63"], '"N$"#,##0.0,,"M"', True),
    ("5-Year Costs", ["=C61","=D61","=E61","=F61","=G61"], '"N$"#,##0.0,,"M"', False),
    ("Net Profit", ["=C62","=D62","=E62","=F62","=G62"], '"N$"#,##0.0,,"M"', True),
    ("Gross Margin", ["=C64","=D64","=E64","=F64","=G64"], '0.0%', True),
    ("Payback", ["=C65","=D65","=E65","=F65","=G65"], '0.0" mo"', True),
]):
    r = 29 + ri
    row_fill = tbl1 if ri % 2 == 0 else tbl2
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = metric; ws.cell(r, 2).font = fw
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db
    for i, sc in enumerate(sc_cols):
        ws.cell(r, sc).value = formulas[i]
        ws.cell(r, sc).font = fg if positive else frd
        ws.cell(r, sc).number_format = fmt; ws.cell(r, sc).fill = row_fill; ws.cell(r, sc).border = db
        ws.cell(r, sc).alignment = Alignment(horizontal="center")
        if i == 2:  # highlight 10K
            ws.cell(r, sc).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.cell(r, sc).font = Font(bold=True, size=10, name="Calibri", color=ACC_GREEN if positive else ACC_RED)

# Data validation
dv = DataValidation(type="list", formula1="=$C$50:$G$50", allow_blank=False)
dv.prompt = "Select scenario"; dv.promptTitle = "Rollout"
ws.add_data_validation(dv); dv.add(ws['C6'])

# ── Finalize ──
wb.move_sheet("Dashboard", offset=-9)  # Move to first position
ws_r.freeze_panes = "B4"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v9.xlsx"
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print("v9: Dark navy dashboard + detailed derivation tabs")
