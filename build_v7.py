"""
GRIDx Financial Dashboard v7 — Complete Model with All Cost Categories
Based on client's detailed financial input with 7 cost categories:
  Manufacturing, Operating, Implementation, Equipment, Development, Financing, Contingency

Dashboard: Dark blue theme, KPI cards, bar charts, financial table, scenario comparison
Revenue: Hardware (annual), Installation (Y1), Maintenance, Wi-Fi, SMS, App, Real Estate
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.series import SeriesLabel, DataPoint
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# CONSTANTS & DATA
# ═══════════════════════════════════════════════════════════════
SCENARIOS = ["3,000 Meters","5,000 Meters","10,000 Meters","20,000 Meters","50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]
NS = 5

# Revenue per meter per year (from client's input)
HW_RATE   = 5500   # Hardware sales (annual recurring)
INST_RATE = 700    # Installation fee (Year 1 only)
MAINT_RATE= 300    # Maintenance (annual)
WIFI_RATE = 90     # Wi-Fi: 30% × N$25 × 12
SMS_RATE  = 30     # SMS: 50% × 2 × N$2.50 × 12
APP_RATE  = 624    # Mobile App: 80% × N$65 × 12
RE_RATE   = 76.80  # Real Estate: ÷500 × N$3,200 × 12

ANNUAL_RECURRING = HW_RATE + MAINT_RATE + WIFI_RATE + SMS_RATE + APP_RATE + RE_RATE  # 6620.80
YEAR1_PER_METER  = ANNUAL_RECURRING + INST_RATE  # 7320.80

# Manufacturing cost per meter by scenario
MFG_UNIT = [2000, 2000, 1700, 1700, 1462]
SPARE_RATE = 4495 * 0.02  # N$89.90 per meter

# Costs for 10K scenario (client's detailed breakdown) — used as reference
# Manufacturing: meters × mfg_unit + meters × spare_rate
# Operating (5-year): Fixed base + variable per meter
OPER_FIXED_ANNUAL = 748800 + 240000 + 120000 + 200000 + 100000 + 150000  # 1,558,800
OPER_VAR_PER_METER_ANNUAL = 48 + 9 + 4  # data storage + vehicles share + field equip share = 61

# One-time costs (scale by scenario factor)
IMPL_BASE = 1500000    # Implementation costs (relatively fixed)
DEV_BASE  = 7851647    # Development costs (relatively fixed)
FIN_BASE  = 700000     # Financing costs (relatively fixed)

# Equipment scales: reference 10K values
EQUIP_10K = 1965000

# Client's scenario totals for verification
CLIENT_REV_5Y  = [101412000, 169020000, 338040000, 676080000, 1690200000]
CLIENT_COST_5Y = [26700000, 33000000, 41994047, 62500000, 109200000]
CLIENT_PROFIT  = [74700000, 136000000, 296045953, 613600000, 1580000000]
CLIENT_MARGIN  = [0.737, 0.805, 0.876, 0.908, 0.935]
CLIENT_PAYBACK = [7.6, 7.2, 4.2, 3.9, 3.5]
CLIENT_DCF     = [65500000, 119300000, 259647182, 538200000, 1390000000]

# ── Shared styles ──
NAVY = "1B2A4A"; STEEL = "2C3E50"; TEAL_C = "2E75B6"
hdr_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor="D6E4F0")
green_fill = PatternFill("solid", fgColor="E2EFDA")
orange_fill = PatternFill("solid", fgColor="FCE4D6")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
bold11 = Font(bold=True, size=11, name="Calibri")
label_font = Font(size=10, name="Calibri")
input_font = Font(color="0000FF", size=10, name="Calibri")
green_font = Font(color="008000", size=10, name="Calibri")
red_font = Font(color="C00000", size=10, name="Calibri")
title_font = Font(bold=True, color=NAVY, size=14, name="Calibri")
navy_bold = Font(bold=True, color=NAVY, size=12, name="Calibri")

thin = Border(left=Side("thin",color="D9D9D9"),right=Side("thin",color="D9D9D9"),
    top=Side("thin",color="D9D9D9"),bottom=Side("thin",color="D9D9D9"))
total_bdr = Border(top=Side("medium",color=NAVY),bottom=Side("double",color=NAVY))

NUM='#,##0'; PCT='0.0%'; CURR='N$#,##0'; CURR_M='"N$"#,##0.0,,"M"'

def wh(ws, r, labels, cs=1):
    for i,l in enumerate(labels):
        c=ws.cell(r,cs+i,l); c.font=hdr_font; c.fill=hdr_fill
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin

def wl(ws,r,c,t,f=label_font,fi=None):
    cell=ws.cell(r,c,t); cell.font=f
    if fi: cell.fill=fi
    cell.border=thin; return cell

def wi(ws,r,c,v,fmt=NUM):
    cell=ws.cell(r,c,v); cell.font=input_font; cell.fill=yellow_fill
    cell.number_format=fmt; cell.border=thin; return cell

def wf(ws,r,c,formula,fmt=NUM,f=label_font,fi=None):
    cell=ws.cell(r,c,formula); cell.font=f
    if fi: cell.fill=fi
    cell.number_format=fmt; cell.border=thin; return cell

def fill_row(ws,r,cs,ce,fi):
    for c in range(cs,ce+1): ws.cell(r,c).fill=fi; ws.cell(r,c).border=thin

# ═══════════════════════════════════════════════════════════════
# SHEET 1: REVENUE MODEL
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Revenue"; ws1.sheet_properties.tabColor = "ED7D31"
ws1.column_dimensions['A'].width = 40; ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 16; ws1.column_dimensions['D'].width = 22; ws1.column_dimensions['E'].width = 22

wl(ws1, 1, 1, "Revenue Model — Per Meter Rates", f=title_font)
wh(ws1, 3, ["Revenue Stream", "Rate", "Adoption", "Basis", "Annual per Meter (N$)"])

rev_items = [
    (4, "Meter Hardware Sale", "N$5,500 (annual)", "100%", "Annual revenue per meter", HW_RATE),
    (5, "Meter Installation Fee", "N$700 (one-time)", "100%", "One-time per new meter", INST_RATE),
    (6, "Maintenance Fees", "N$300 / year", "100%", "Annual maintenance", MAINT_RATE),
    (7, "Wi-Fi Subscription", "N$25 / month", "30%", "30% × N$25 × 12", WIFI_RATE),
    (8, "SMS Notifications", "N$2.50 / SMS", "50%", "50% × 2 SMS × N$2.50 × 12", SMS_RATE),
    (9, "Mobile App Subscription", "N$65 / month", "80%", "80% × N$65 × 12", APP_RATE),
    (10, "Real Estate Management", "N$3,200/block/mo", "1 per 500m", "(meters÷500)×3200×12÷meters", RE_RATE),
]
for r, name, rate, adopt, basis, annual in rev_items:
    wl(ws1, r, 1, name); wl(ws1, r, 2, rate); wl(ws1, r, 3, adopt); wl(ws1, r, 4, basis)
    wi(ws1, r, 5, annual, 'N$#,##0.00')

wl(ws1, 11, 1, "Annual Recurring per Meter", f=bold_font, fi=green_fill)
fill_row(ws1, 11, 2, 5, green_fill)
cell = wf(ws1, 11, 5, "=E4+E6+E7+E8+E9+E10", 'N$#,##0.00', bold_font, green_fill)
cell.border = total_bdr

wl(ws1, 12, 1, "Year 1 Total per Meter (incl. installation)", f=bold_font, fi=green_fill)
fill_row(ws1, 12, 2, 5, green_fill)
cell = wf(ws1, 12, 5, "=E11+E5", 'N$#,##0.00', bold_font, green_fill)
cell.border = total_bdr

# Revenue by scenario (5-year)
r = 15; wl(ws1, r, 1, "5-Year Revenue by Scenario", f=title_font)
r = 17; wh(ws1, r, ["Revenue Stream"] + SCENARIOS)
rev_streams = ["Hardware Sales","Installation Fees","Maintenance Fees","Wi-Fi Subscriptions",
    "SMS Notifications","Mobile App Subscriptions","Real Estate Management"]
# Per-meter annual rates (all recurring except installation)
per_meter_annual = [HW_RATE, 0, MAINT_RATE, WIFI_RATE, SMS_RATE, APP_RATE, RE_RATE]
per_meter_y1_only = [0, INST_RATE, 0, 0, 0, 0, 0]

r = 18
for idx_s, (stream, annual, y1_only) in enumerate(zip(rev_streams, per_meter_annual, per_meter_y1_only)):
    wl(ws1, r, 1, stream)
    for i in range(NS):
        if y1_only > 0:
            val = MC[i] * y1_only  # one-time
        else:
            val = MC[i] * annual * 5  # 5-year recurring
        wi(ws1, r, 2+i, val, CURR)
    r += 1

wl(ws1, r, 1, "Total 5-Year Revenue", f=bold11, fi=green_fill)
fill_row(ws1, r, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws1, r, 2+i, f"=SUM({col}18:{col}24)", CURR, bold11, green_fill)
    cell.border = total_bdr
REV_TOTAL_ROW = r

# ═══════════════════════════════════════════════════════════════
# SHEET 2: COST DETAIL
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cost Detail"); ws2.sheet_properties.tabColor = "C00000"
ws2.column_dimensions['A'].width = 42
for i in range(NS): ws2.column_dimensions[get_column_letter(2+i)].width = 18

wl(ws2, 1, 1, "Complete Cost Structure — 5-Year Projections", f=title_font)

# --- MANUFACTURING COSTS ---
r = 3; wl(ws2, r, 1, "MANUFACTURING COSTS", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
wl(ws2, r, 1, "Meter Hardware")
for i in range(NS): wi(ws2, r, 2+i, MC[i] * MFG_UNIT[i], CURR)
r += 1
wl(ws2, r, 1, "Spare Parts Inventory (2%)")
for i in range(NS): wi(ws2, r, 2+i, round(MC[i] * SPARE_RATE), CURR)
r += 1
wl(ws2, r, 1, "Total Manufacturing", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{r-2}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
MFG_TOT_ROW = r

# --- OPERATING COSTS (5-YEAR) ---
r += 2; wl(ws2, r, 1, "OPERATING COSTS (5-Year)", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
oper_items = [
    ("Personnel Costs (N$748,800/yr)", [748800*5]*5),
    ("Data Storage & Transmission", [m*48*5 for m in MC]),
    ("Service Vehicles Maintenance", [v*30000*5 for v in [2, 2, 3, 5, 10]]),
    ("Field Equipment Maintenance", [k*10000*5 for k in [2, 3, 4, 6, 12]]),
    ("Office Rental & Utilities", [240000*5]*5),
    ("Software Licensing & Tools", [120000*5]*5),
    ("Marketing & Customer Acquisition", [200000*5]*5),
    ("Regulatory & Compliance", [100000*5]*5),
    ("Insurance", [150000*5]*5),
]
oper_start = r
for name, vals in oper_items:
    wl(ws2, r, 1, name)
    for i, v in enumerate(vals): wi(ws2, r, 2+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Operating Costs (5-Year)", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{oper_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
OPER_TOT_ROW = r

# --- IMPLEMENTATION COSTS ---
r += 2; wl(ws2, r, 1, "IMPLEMENTATION COSTS (One-Time)", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
impl_items = [
    ("Project Management & Planning", [500000]*5),
    ("Site Surveys & Assessments", [300000]*5),
    ("Customer Onboarding & Training", [250000]*5),
    ("Integration with Utility Systems", [450000]*5),
]
impl_start = r
for name, vals in impl_items:
    wl(ws2, r, 1, name)
    for i, v in enumerate(vals): wi(ws2, r, 2+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Implementation", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{impl_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
IMPL_TOT_ROW = r

# --- EQUIPMENT INVESTMENT ---
r += 2; wl(ws2, r, 1, "EQUIPMENT INVESTMENT (One-Time)", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
# Scale equipment based on scenario
equip_items = [
    ("Service Vehicles", [250000*v for v in [2, 2, 3, 5, 10]]),
    ("Field Test Equipment Kits", [85000*k for k in [2, 3, 4, 6, 12]]),
    ("Workstations", [15000*w for w in [3, 5, 7, 12, 25]]),
    ("Network Monitoring Systems", [120000*n for n in [1, 1, 2, 3, 5]]),
    ("Office Equipment & Furniture", [50000+round(m/10000*230000) for m in MC]),
    ("Server/Cloud Infrastructure", [100000+round(m/10000*150000) for m in MC]),
]
equip_start = r
for name, vals in equip_items:
    wl(ws2, r, 1, name)
    for i, v in enumerate(vals): wi(ws2, r, 2+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Equipment", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{equip_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
EQUIP_TOT_ROW = r

# --- DEVELOPMENT COSTS ---
r += 2; wl(ws2, r, 1, "DEVELOPMENT COSTS", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
dev_items = [
    ("Electronics Manufacturing Development", [312000]*5),
    ("Software Engineering (R&D)", [777747]*5),
    ("Mobile Application Development", [300000]*5),
    ("Hardware Firmware Development", [2160000]*5),
    ("Enclosure Manufacturing & Assembly", [1403850]*5),
    ("Marketing (Branding, Collateral)", [325000]*5),
    ("Regulatory & Legal (Certifications)", [1060000]*5),
    ("Office & Logistics Setup", [897400]*5),
    ("Development Contingency (5%)", [615650]*5),
]
dev_start = r
for name, vals in dev_items:
    wl(ws2, r, 1, name)
    for i, v in enumerate(vals): wi(ws2, r, 2+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Development", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{dev_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
DEV_TOT_ROW = r

# --- FINANCING COSTS ---
r += 2; wl(ws2, r, 1, "FINANCING COSTS", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
fin_items = [
    ("Interest on Working Capital", [500000]*5),
    ("Bank Fees & Transaction Costs", [200000]*5),
]
fin_start = r
for name, vals in fin_items:
    wl(ws2, r, 1, name)
    for i, v in enumerate(vals): wi(ws2, r, 2+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Financing", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{fin_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
FIN_TOT_ROW = r

# --- CONTINGENCY ---
r += 2; wl(ws2, r, 1, "CONTINGENCY", f=bold11, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
wh(ws2, r+1, ["Cost Item"] + SCENARIOS)
r += 2
cont_start = r
wl(ws2, r, 1, "Operations Contingency (10%)")
for i in range(NS):
    wf(ws2, r, 2+i, f"={get_column_letter(2+i)}{OPER_TOT_ROW}*0.1", CURR, label_font)
r += 1
wl(ws2, r, 1, "Implementation Contingency (10%)")
for i in range(NS):
    wf(ws2, r, 2+i, f"={get_column_letter(2+i)}{IMPL_TOT_ROW}*0.1", CURR, label_font)
r += 1
wl(ws2, r, 1, "Total Contingency", f=bold_font, fi=orange_fill)
fill_row(ws2, r, 2, 6, orange_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws2, r, 2+i, f"=SUM({col}{cont_start}:{col}{r-1})", CURR, bold_font, orange_fill)
    cell.border = total_bdr
CONT_TOT_ROW = r

# --- GRAND TOTAL ---
r += 2; wl(ws2, r, 1, "TOTAL 5-YEAR COSTS", f=navy_bold, fi=green_fill)
fill_row(ws2, r, 2, 6, green_fill)
cost_rows = [MFG_TOT_ROW, OPER_TOT_ROW, IMPL_TOT_ROW, EQUIP_TOT_ROW, DEV_TOT_ROW, FIN_TOT_ROW, CONT_TOT_ROW]
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{col}{cr}" for cr in cost_rows])
    cell = wf(ws2, r, 2+i, f"={refs}", CURR, navy_bold, green_fill)
    cell.border = total_bdr
GRAND_COST_ROW = r

print(f"Cost rows: Mfg={MFG_TOT_ROW}, Oper={OPER_TOT_ROW}, Impl={IMPL_TOT_ROW}, Equip={EQUIP_TOT_ROW}, Dev={DEV_TOT_ROW}, Fin={FIN_TOT_ROW}, Cont={CONT_TOT_ROW}, Grand={GRAND_COST_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: 5-YEAR MODEL (Year-by-year breakdown)
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("5 Year Model"); ws3.sheet_properties.tabColor = TEAL_C
ws3.column_dimensions['A'].width = 36
for i in range(6): ws3.column_dimensions[get_column_letter(2+i)].width = 18

wl(ws3, 1, 1, "5-Year Financial Model — Year-by-Year Projections", f=title_font)
wl(ws3, 2, 1, "Linked to scenario selected on Dashboard", f=Font(italic=True, size=10, color="666666", name="Calibri"))

wh(ws3, 4, ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "5-Year Total"])

# Revenue rows
r = 5
wl(ws3, r, 1, "REVENUE", f=bold11, fi=green_fill); fill_row(ws3, r, 2, 8, green_fill)

# Use Dashboard hidden data reference — will add IDX formulas
# For now, put placeholder formulas that reference Dashboard data area
rev_stream_data = [
    ("Hardware Sales", HW_RATE, True),      # annual recurring
    ("Installation Fees", INST_RATE, False), # Year 1 only
    ("Maintenance Fees", MAINT_RATE, True),
    ("Wi-Fi Subscriptions", WIFI_RATE, True),
    ("SMS Notifications", SMS_RATE, True),
    ("Mobile App Subscriptions", APP_RATE, True),
    ("Real Estate Management", RE_RATE, True),
]
r = 6
for stream, rate, recurring in rev_stream_data:
    wl(ws3, r, 1, stream)
    # Year 1
    ws3.cell(r, 2).value = f"=Dashboard!$C$52*{rate}"
    ws3.cell(r, 2).font = green_font; ws3.cell(r, 2).number_format = CURR
    if recurring:
        for yc in range(3, 7):  # Years 2-5
            ws3.cell(r, yc).value = f"=B{r}"
            ws3.cell(r, yc).font = green_font; ws3.cell(r, yc).number_format = CURR
    else:
        for yc in range(3, 7):
            ws3.cell(r, yc).value = 0
            ws3.cell(r, yc).font = green_font; ws3.cell(r, yc).number_format = CURR
    # 5-Year Total
    wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, green_font)
    r += 1

wl(ws3, r, 1, "Total Revenue", f=bold11, fi=green_fill)
fill_row(ws3, r, 2, 8, green_fill)
for c in range(2, 9):
    col = get_column_letter(c)
    cell = wf(ws3, r, c, f"=SUM({col}6:{col}12)", CURR, bold11, green_fill)
    cell.border = total_bdr
REV_TOT_5YR = r

# Cost rows
r += 2; wl(ws3, r, 1, "COSTS", f=bold11, fi=orange_fill); fill_row(ws3, r, 2, 8, orange_fill)
r += 1; cost_detail_start = r

# Manufacturing (Year 1 only)
wl(ws3, r, 1, "Manufacturing Costs")
ws3.cell(r, 2).value = f"=Dashboard!$C$54"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc, 0).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Operating costs (annual)
wl(ws3, r, 1, "Operating Costs")
ws3.cell(r, 2).value = f"=Dashboard!$C$55/5"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc).value = f"=B{r}"; ws3.cell(r, yc).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Implementation (Year 1)
wl(ws3, r, 1, "Implementation Costs")
ws3.cell(r, 2).value = f"=Dashboard!$C$56"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc, 0).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Equipment (Year 1)
wl(ws3, r, 1, "Equipment Investment")
ws3.cell(r, 2).value = f"=Dashboard!$C$57"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc, 0).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Development (Year 1)
wl(ws3, r, 1, "Development Costs")
ws3.cell(r, 2).value = f"=Dashboard!$C$58"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc, 0).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Financing (Year 1)
wl(ws3, r, 1, "Financing Costs")
ws3.cell(r, 2).value = f"=Dashboard!$C$59"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc, 0).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Contingency (spread over 5 years)
wl(ws3, r, 1, "Contingency")
ws3.cell(r, 2).value = f"=Dashboard!$C$60/5"; ws3.cell(r, 2).font = red_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7): ws3.cell(r, yc).value = f"=B{r}"; ws3.cell(r, yc).font = red_font; ws3.cell(r, yc).number_format = CURR
wf(ws3, r, 8, f"=SUM(B{r}:G{r})", CURR, red_font); r += 1

# Total costs
wl(ws3, r, 1, "Total Costs", f=bold11, fi=orange_fill)
fill_row(ws3, r, 2, 8, orange_fill)
for c in range(2, 9):
    col = get_column_letter(c)
    cell = wf(ws3, r, c, f"=SUM({col}{cost_detail_start}:{col}{r-1})", CURR, bold11, orange_fill)
    cell.border = total_bdr
COST_TOT_5YR = r

# Net Profit
r += 1; wl(ws3, r, 1, "NET PROFIT", f=navy_bold, fi=green_fill)
fill_row(ws3, r, 2, 8, green_fill)
for c in range(2, 9):
    col = get_column_letter(c)
    cell = wf(ws3, r, c, f"={col}{REV_TOT_5YR}-{col}{COST_TOT_5YR}", CURR, navy_bold, green_fill)
    cell.border = total_bdr
PROFIT_ROW = r

# Cumulative
r += 2; wl(ws3, r, 1, "Cumulative Revenue", f=bold_font)
ws3.cell(r, 2).value = f"=B{REV_TOT_5YR}"
ws3.cell(r, 2).font = green_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7):
    col = get_column_letter(yc); prev = get_column_letter(yc-1)
    ws3.cell(r, yc).value = f"={prev}{r}+{col}{REV_TOT_5YR}"
    ws3.cell(r, yc).font = green_font; ws3.cell(r, yc).number_format = CURR
CUM_REV_ROW = r

r += 1; wl(ws3, r, 1, "Cumulative Net Profit", f=bold_font)
ws3.cell(r, 2).value = f"=B{PROFIT_ROW}"
ws3.cell(r, 2).font = green_font; ws3.cell(r, 2).number_format = CURR
for yc in range(3, 7):
    col = get_column_letter(yc); prev = get_column_letter(yc-1)
    ws3.cell(r, yc).value = f"={prev}{r}+{col}{PROFIT_ROW}"
    ws3.cell(r, yc).font = green_font; ws3.cell(r, yc).number_format = CURR
CUM_PROF_ROW = r

ws3.freeze_panes = "B5"

# ═══════════════════════════════════════════════════════════════
# SHEET 4: DASHBOARD (Main Visual)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "002060"

# Column widths
ws.column_dimensions['A'].width = 1.5
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 1.5
for ch in 'EFGHIJKLMNOP':
    ws.column_dimensions[ch].width = 8
ws.column_dimensions['Q'].width = 1.5

# Row heights
ws.row_dimensions[1].height = 4
ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 4
for rr in range(4, 50):
    ws.row_dimensions[rr].height = 18

# Dashboard styles
HDR_BG    = PatternFill("solid", fgColor="1B3A5C")
PANEL_BG  = PatternFill("solid", fgColor="F2F6FA")
CARD_BG   = PatternFill("solid", fgColor="FFFFFF")
DROP_BG   = PatternFill("solid", fgColor="1565C0")
TILE_BG   = PatternFill("solid", fgColor="FFFFFF")
KPI_BG1   = PatternFill("solid", fgColor="E8F5E9")  # Revenue green
KPI_BG2   = PatternFill("solid", fgColor="E3F2FD")  # Profit blue
KPI_BG3   = PatternFill("solid", fgColor="FFF3E0")  # Margin orange
KPI_BG4   = PatternFill("solid", fgColor="F3E5F5")  # Payback purple
PB_BG     = PatternFill("solid", fgColor="E3F2FD")
MARGIN_GR = PatternFill("solid", fgColor="2E7D32")
MARGIN_YL = PatternFill("solid", fgColor="F9A825")
MARGIN_BL = PatternFill("solid", fgColor="1565C0")

# Fonts
f_banner  = Font(bold=True, size=16, name="Calibri", color="FFFFFF")
f_title   = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
f_label   = Font(bold=True, size=10, name="Calibri", color="37474F")
f_value   = Font(bold=True, size=16, name="Calibri", color="1B3A5C")
f_unit    = Font(size=9, name="Calibri", color="90A4AE")
f_drop    = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
f_kpi_val = Font(bold=True, size=18, name="Calibri", color="1B3A5C")
f_kpi_lbl = Font(bold=True, size=9, name="Calibri", color="546E7A")
f_mhead   = Font(bold=True, size=12, name="Calibri", color="1B3A5C")
f_mlab    = Font(size=10, name="Calibri", color="455A64")
f_mpct    = Font(bold=True, size=11, name="Calibri", color="1B3A5C")
f_pbnum   = Font(bold=True, size=28, name="Calibri", color="1565C0")
f_pbday   = Font(bold=True, size=12, name="Calibri", color="455A64")
ref_font  = Font(size=7, color="BDBDBD", name="Calibri")

bdr_panel = Border(left=Side("medium",color="B0BEC5"),right=Side("medium",color="B0BEC5"),
    top=Side("medium",color="B0BEC5"),bottom=Side("medium",color="B0BEC5"))
bdr_tile = Border(left=Side("thin",color="CFD8DC"),right=Side("thin",color="CFD8DC"),
    top=Side("thin",color="CFD8DC"),bottom=Side("thin",color="CFD8DC"))
bdr_card = Border(left=Side("thin",color="E0E0E0"),right=Side("thin",color="E0E0E0"),
    top=Side("thin",color="E0E0E0"),bottom=Side("thin",color="E0E0E0"))

IDX = "$C$51"
def idx(rr):
    return f"INDEX($C${rr}:$G${rr},1,{IDX})"

def draw_panel(ws, r1, c1, r2, c2, fill, border):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c); cell.fill = fill
            left = border.left if c == c1 else Side()
            right = border.right if c == c2 else Side()
            top = border.top if r == r1 else Side()
            bottom = border.bottom if r == r2 else Side()
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# ════════════ TITLE BAR ════════════
ws.merge_cells('B2:P2')
ws['B2'].value = "GRIDx Smart Meter — 5-Year Financial Model"
ws['B2'].font = f_banner; ws['B2'].fill = HDR_BG
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 17): ws.cell(2, c).fill = HDR_BG

# ════════════ LEFT PANEL: INPUT PARAMETERS ════════════
draw_panel(ws, 4, 2, 27, 3, PANEL_BG, bdr_panel)

ws.merge_cells('B4:C4')
ws['B4'].value = "Input Parameters"
ws['B4'].font = f_title; ws['B4'].fill = HDR_BG
ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
for c in [2, 3]: ws.cell(4, c).fill = HDR_BG

# Dropdown
ws['B5'] = "Meter Rollout:"; ws['B5'].font = f_label; ws['B5'].fill = PANEL_BG
ws.merge_cells('B6:C6')
ws['B6'] = "10,000 Meters"; ws['B6'].font = f_drop; ws['B6'].fill = DROP_BG
ws['B6'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[6].height = 26
for c in [2, 3]:
    ws.cell(6, c).fill = DROP_BG
    ws.cell(6, c).border = Border(left=Side("medium",color="0D47A1") if c==2 else Side(),
        right=Side("medium",color="0D47A1") if c==3 else Side(),
        top=Side("medium",color="0D47A1"), bottom=Side("medium",color="0D47A1"))

# Parameter cards
def make_card(sr, label, val, unit):
    ws.cell(sr, 2).value = label; ws.cell(sr, 2).font = f_label; ws.cell(sr, 2).fill = PANEL_BG
    ws.cell(sr, 3).value = unit; ws.cell(sr, 3).font = f_unit; ws.cell(sr, 3).fill = PANEL_BG
    ws.cell(sr, 3).alignment = Alignment(horizontal="right")
    ws.merge_cells(f'B{sr+1}:C{sr+1}')
    vc = ws.cell(sr+1, 2)
    vc.value = val; vc.font = f_value; vc.fill = CARD_BG; vc.number_format = '"N$ "#,##0'
    vc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[sr+1].height = 24
    for c in [2,3]: ws.cell(sr+1, c).fill = CARD_BG; ws.cell(sr+1, c).border = bdr_card

make_card(8, "Hardware Price", 5500, "Per Unit/Year")
make_card(11, "Installation Fee", 700, "One-Time")
make_card(14, "Manufacturing Cost", f"={idx(53)}", "Per Unit")
make_card(17, "Personnel (Annual)", f"=748800", "Per Year")
make_card(20, "Service Revenue", f"={idx(52)}*{ANNUAL_RECURRING}", "Per Year")
make_card(23, "Total 5Y Costs", f"={idx(61)}", "5-Year")
make_card(23, "Total 5Y Costs", f"={idx(61)}", "5-Year")
ws.cell(24, 2).number_format = '"N$ "#,##0'
make_card(26, "Net Profit", f"={idx(62)}", "5-Year")
ws.cell(27, 2).number_format = '"N$ "#,##0'

# ════════════ KPI CARDS (Row 4-6, cols E-P) ════════════
kpi_configs = [
    (5, 8, "Total Revenue (5Y)", f"={idx(63)}", '"N$"#,##0.0,,"M"', KPI_BG1),
    (8, 11, "Net Profit (5Y)", f"={idx(62)}", '"N$"#,##0.0,,"M"', KPI_BG2),
    (11, 14, "Gross Margin", f"={idx(64)}", '0.0%', KPI_BG3),
    (14, 17, "Payback Period", f"={idx(65)}", '0.0" months"', KPI_BG4),
]
for cs, ce, label, formula, fmt, bg in kpi_configs:
    # KPI card background
    for r in range(4, 7):
        for c in range(cs, ce):
            ws.cell(r, c).fill = bg
            ws.cell(r, c).border = bdr_card
    # Label
    ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce-1)
    ws.cell(4, cs).value = label
    ws.cell(4, cs).font = f_kpi_lbl
    ws.cell(4, cs).alignment = Alignment(horizontal="center", vertical="center")
    # Value
    ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce-1)
    ws.cell(5, cs).value = formula
    ws.cell(5, cs).font = f_kpi_val
    ws.cell(5, cs).number_format = fmt
    ws.cell(5, cs).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 30

# ════════════ HIDDEN DATA AREA (row 50+) ════════════
ws.cell(49, 2, "DATA AREA").font = ref_font
for i, s in enumerate(SCENARIOS):
    ws.cell(50, 3+i, s).font = ref_font

# Row 51: MATCH index
ws.cell(51, 2, "Index").font = ref_font
ws.cell(51, 3).value = '=MATCH($B$6,$C$50:$G$50,0)'; ws.cell(51, 3).font = ref_font

# Row 52: Meter counts
for i, m in enumerate(MC):
    ws.cell(52, 3+i, m).font = ref_font

# Row 53: Mfg cost per meter
for i in range(NS):
    ws.cell(53, 3+i, MFG_UNIT[i]).font = ref_font

# Row 54: Total manufacturing
for i in range(NS):
    ws.cell(54, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{MFG_TOT_ROW}"
    ws.cell(54, 3+i).font = ref_font

# Row 55: Total operating
for i in range(NS):
    ws.cell(55, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{OPER_TOT_ROW}"
    ws.cell(55, 3+i).font = ref_font

# Row 56: Total implementation
for i in range(NS):
    ws.cell(56, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{IMPL_TOT_ROW}"
    ws.cell(56, 3+i).font = ref_font

# Row 57: Total equipment
for i in range(NS):
    ws.cell(57, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{EQUIP_TOT_ROW}"
    ws.cell(57, 3+i).font = ref_font

# Row 58: Total development
for i in range(NS):
    ws.cell(58, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{DEV_TOT_ROW}"
    ws.cell(58, 3+i).font = ref_font

# Row 59: Total financing
for i in range(NS):
    ws.cell(59, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{FIN_TOT_ROW}"
    ws.cell(59, 3+i).font = ref_font

# Row 60: Total contingency
for i in range(NS):
    ws.cell(60, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{CONT_TOT_ROW}"
    ws.cell(60, 3+i).font = ref_font

# Row 61: Grand total costs
for i in range(NS):
    ws.cell(61, 3+i).value = f"='Cost Detail'!{get_column_letter(2+i)}{GRAND_COST_ROW}"
    ws.cell(61, 3+i).font = ref_font

# Row 62: Net Profit
for i in range(NS):
    ws.cell(62, 3+i).value = f"='Revenue'!{get_column_letter(2+i)}{REV_TOTAL_ROW}-{get_column_letter(3+i)}61"
    # Fix: Revenue uses col 2+i, costs use col 3+i
    # Actually Revenue sheet uses columns 2-6 (B-F), Cost Detail also uses 2-6
    # Dashboard data area uses columns 3-7 (C-G)
    # Let me fix this
ws.cell(62, 3).value = f"=C{REV_TOTAL_ROW+10}-C61"  # temp, will fix below

# Actually, let me recalculate. Revenue total is on Revenue sheet, row REV_TOTAL_ROW, columns B-F (2-6)
# Cost total is on Cost Detail sheet, row GRAND_COST_ROW, columns B-F (2-6)
# Dashboard data has these in rows 54-61, columns C-G (3-7)
# I need Revenue totals in the data area too

# Row 63: Total Revenue (5-year)
for i in range(NS):
    ws.cell(63, 3+i).value = f"='Revenue'!{get_column_letter(2+i)}{REV_TOTAL_ROW}"
    ws.cell(63, 3+i).font = ref_font

# Fix Row 62: Net Profit = Revenue - Costs
for i in range(NS):
    col = get_column_letter(3+i)
    ws.cell(62, 3+i).value = f"={col}63-{col}61"
    ws.cell(62, 3+i).font = ref_font

# Row 64: Gross Margin
for i in range(NS):
    col = get_column_letter(3+i)
    ws.cell(64, 3+i).value = f"=IF({col}63>0,{col}62/{col}63,0)"
    ws.cell(64, 3+i).font = ref_font

# Row 65: Payback months
for i in range(NS):
    ws.cell(65, 3+i, CLIENT_PAYBACK[i]).font = ref_font

# Row 66: DCF
for i in range(NS):
    ws.cell(66, 3+i, CLIENT_DCF[i]).font = ref_font

# Row 67: HW Profit Margin
for i in range(NS):
    ws.cell(67, 3+i, CLIENT_MARGIN[i]).font = ref_font  # Using client's HW margin

# ── Year-by-year chart data (rows 70+) ──
for i, yr in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    ws.cell(70, 4+i, yr).font = ref_font

# Revenue by year (in millions for charts)
ws.cell(71, 3, "Revenue M").font = ref_font
for yc in range(4, 9):
    yr_idx = yc - 4  # 0-4
    if yr_idx == 0:
        ws.cell(71, yc).value = f"={idx(52)}*{YEAR1_PER_METER}/1000000"
    else:
        ws.cell(71, yc).value = f"={idx(52)}*{ANNUAL_RECURRING}/1000000"

# Costs by year (in millions) — Year 1 has all one-time + operating share
ws.cell(72, 3, "Costs M").font = ref_font
ws.cell(72, 4).value = f"=({idx(54)}+{idx(55)}/5+{idx(56)}+{idx(57)}+{idx(58)}+{idx(59)}+{idx(60)}/5)/1000000"
for yc in range(5, 9):
    ws.cell(72, yc).value = f"=({idx(55)}/5+{idx(60)}/5)/1000000"

# Net Profit by year (in millions)
ws.cell(73, 3, "Profit M").font = ref_font
for yc in range(4, 9):
    cl = get_column_letter(yc)
    ws.cell(73, yc).value = f"={cl}71-{cl}72"

# Cumulative Revenue (millions)
ws.cell(74, 3, "Cum Rev M").font = ref_font
ws.cell(74, 4).value = "=D71"
for yc in range(5, 9):
    ws.cell(74, yc).value = f"={get_column_letter(yc-1)}74+{get_column_letter(yc)}71"

# Cumulative Profit (millions)
ws.cell(75, 3, "Cum Prof M").font = ref_font
ws.cell(75, 4).value = "=D73"
for yc in range(5, 9):
    ws.cell(75, yc).value = f"={get_column_letter(yc-1)}75+{get_column_letter(yc)}73"

# Cost breakdown data for donut (row 77-78)
cost_cats = ["Manufacturing", "Operating", "Implementation", "Equipment", "Development", "Financing", "Contingency"]
for i, cat in enumerate(cost_cats):
    ws.cell(77, 4+i, cat).font = ref_font
    ws.cell(78, 4+i).value = f"={idx(54+i)}"
    ws.cell(78, 4+i).font = ref_font

# ════════════ CHARTS ════════════

cats = Reference(ws, min_col=4, min_row=70, max_col=8, max_row=70)

# CHART 1: Revenue vs Costs Bar (TOP LEFT, E8:J20)
bar1 = BarChart()
bar1.type = "col"; bar1.grouping = "clustered"
bar1.title = "Annual Revenue & Profit (N$ Millions)"
bar1.style = 10; bar1.y_axis.title = "N$ Millions"; bar1.y_axis.numFmt = '0'

d_rev = Reference(ws, min_col=4, min_row=71, max_col=8, max_row=71)
bar1.add_data(d_rev, from_rows=True)
bar1.series[0].title = SeriesLabel(v="Revenue")
bar1.series[0].graphicalProperties.solidFill = "1B3A5C"

d_cost = Reference(ws, min_col=4, min_row=72, max_col=8, max_row=72)
bar1.add_data(d_cost, from_rows=True)
bar1.series[1].title = SeriesLabel(v="Costs")
bar1.series[1].graphicalProperties.solidFill = "E53935"

d_prof = Reference(ws, min_col=4, min_row=73, max_col=8, max_row=73)
bar1.add_data(d_prof, from_rows=True)
bar1.series[2].title = SeriesLabel(v="Net Profit")
bar1.series[2].graphicalProperties.solidFill = "4CAF50"

# Line overlay
line1 = LineChart()
d_trend = Reference(ws, min_col=4, min_row=71, max_col=8, max_row=71)
line1.add_data(d_trend, from_rows=True)
line1.series[0].title = SeriesLabel(v="Revenue Trend")
line1.series[0].graphicalProperties.line.solidFill = "FF9800"
line1.series[0].graphicalProperties.line.width = 25000
line1.y_axis.axId = 200
bar1 += line1

bar1.set_categories(cats)
bar1.legend.position = 'b'
bar1.width = 17; bar1.height = 13
ws.add_chart(bar1, "E8")

# CHART 2: Cost Breakdown Donut (TOP RIGHT, K8:P20)
ch2 = DoughnutChart()
ch2.title = "5-Year Cost Breakdown"
ch2.style = 10

donut_data = Reference(ws, min_col=4, min_row=78, max_col=10, max_row=78)
donut_cats = Reference(ws, min_col=4, min_row=77, max_col=10, max_row=77)
ch2.add_data(donut_data, from_rows=True)
ch2.set_categories(donut_cats)

donut_colors = ["1B3A5C", "FF9800", "4CAF50", "9C27B0", "2E75B6", "F44336", "FFC107"]
for i, color in enumerate(donut_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    ch2.series[0].data_points.append(pt)

ch2.legend.position = 'b'
ch2.width = 14; ch2.height = 13
ws.add_chart(ch2, "K8")

# CHART 3: Cumulative Cash Flow (BOTTOM LEFT, E22:J34)
bar3 = BarChart()
bar3.type = "col"; bar3.grouping = "clustered"
bar3.title = "Cumulative Cash Flow (N$ Millions)"
bar3.style = 10; bar3.y_axis.title = "N$ Millions"; bar3.y_axis.numFmt = '0'

d_cr = Reference(ws, min_col=4, min_row=74, max_col=8, max_row=74)
bar3.add_data(d_cr, from_rows=True)
bar3.series[0].title = SeriesLabel(v="Cumulative Revenue")
bar3.series[0].graphicalProperties.solidFill = "2E7D32"

d_cp = Reference(ws, min_col=4, min_row=75, max_col=8, max_row=75)
bar3.add_data(d_cp, from_rows=True)
bar3.series[1].title = SeriesLabel(v="Cumulative Profit")
bar3.series[1].graphicalProperties.solidFill = "81C784"

bar3.set_categories(cats)
bar3.legend.position = 'b'
bar3.width = 17; bar3.height = 12
ws.add_chart(bar3, "E22")

# BOTTOM RIGHT: Margins + Payback (K22:P34)
# Profit Margins section
ws.merge_cells('K22:P22')
ws['K22'].value = "Profit Margins"
ws['K22'].font = f_mhead
ws['K22'].alignment = Alignment(horizontal="center", vertical="center")

margin_items = [
    (23, "Gross Margin", MARGIN_GR, f"={idx(64)}"),
    (24, "Operating Margin", MARGIN_YL, f"=IF({idx(63)}>0,({idx(63)}-{idx(61)})/{idx(63)},0)"),
    (25, "Net Margin", MARGIN_BL, f"={idx(64)}"),
]
for mr, lab, dot_fill, formula in margin_items:
    ws.cell(mr, 11).fill = dot_fill
    ws.merge_cells(f'L{mr}:M{mr}')
    ws.cell(mr, 12).value = lab; ws.cell(mr, 12).font = f_mlab
    ws.merge_cells(f'N{mr}:P{mr}')
    c = ws.cell(mr, 14); c.value = formula; c.font = f_mpct; c.number_format = '0.0%'
    c.alignment = Alignment(horizontal="right")

# Payback Period
ws.merge_cells('K27:P27')
ws['K27'].value = "Payback Period"
ws['K27'].font = f_mhead
ws['K27'].alignment = Alignment(horizontal="center", vertical="center")

for r in range(28, 32):
    for c in range(11, 17): ws.cell(r, c).fill = PB_BG; ws.cell(r, c).border = bdr_card

ws.merge_cells('K28:P29')
ws['K28'].value = f"={idx(65)}"
ws['K28'].font = f_pbnum; ws['K28'].fill = PB_BG
ws['K28'].alignment = Alignment(horizontal="center", vertical="center")
ws['K28'].number_format = '0.0'
ws.row_dimensions[28].height = 28
ws.row_dimensions[29].height = 20

ws.merge_cells('K30:P30')
ws['K30'].value = "Months"
ws['K30'].font = f_pbday; ws['K30'].fill = PB_BG
ws['K30'].alignment = Alignment(horizontal="center")

ws.merge_cells('K31:P31')
ws['K31'].value = "Payback Period"
ws['K31'].font = Font(size=9, color="78909C", name="Calibri")
ws['K31'].fill = PB_BG; ws['K31'].alignment = Alignment(horizontal="center")

# ════════════ SCENARIO COMPARISON TABLE (rows 36+) ════════════
r = 36
ws.merge_cells(f'B{r}:P{r}')
ws.cell(r, 2).value = "Scenario Comparison"
ws.cell(r, 2).font = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
ws.cell(r, 2).fill = HDR_BG
ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 17): ws.cell(r, c).fill = HDR_BG
ws.row_dimensions[r].height = 28

r = 37
# Header: merge cols for each scenario
labels = ["Metric"] + SCENARIOS
col_starts = [2, 5, 8, 11, 14, 17]  # B, E, H, K, N, Q... but only have to P(16)
# Simplified: use cols B,E,G,I,K,M,O... or just use single columns
# Better: use fewer columns. B=label, then D,F,H,J,L for 5 scenarios
# Actually with current layout B=label(wide), then need 5 value columns
# Let me use B for label, then merge pairs for each scenario

ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = hdr_font; ws.cell(r, 2).fill = hdr_fill
ws.cell(r, 2).border = thin
for i, s in enumerate(SCENARIOS):
    cs_col = 5 + i*2
    if cs_col + 1 <= 16:
        ws.merge_cells(start_row=r, start_column=cs_col, end_row=r, end_column=min(cs_col+1, 16))
    ws.cell(r, cs_col).value = s.replace(",000 Meters", "K")
    ws.cell(r, cs_col).font = hdr_font; ws.cell(r, cs_col).fill = hdr_fill
    ws.cell(r, cs_col).alignment = Alignment(horizontal="center")
    ws.cell(r, cs_col).border = thin

# Data rows
sc_metrics = [
    ("Total Revenue (5Y)", [f"=C63", f"=D63", f"=E63", f"=F63", f"=G63"], CURR_M),
    ("Total Costs (5Y)", [f"=C61", f"=D61", f"=E61", f"=F61", f"=G61"], CURR_M),
    ("Net Profit (5Y)", [f"=C62", f"=D62", f"=E62", f"=F62", f"=G62"], CURR_M),
    ("Gross Margin", [f"=C64", f"=D64", f"=E64", f"=F64", f"=G64"], PCT),
    ("Payback Period", [f"=C65", f"=D65", f"=E65", f"=F65", f"=G65"], '0.0" mo"'),
    ("DCF Valuation", [f"=C66", f"=D66", f"=E66", f"=F66", f"=G66"], CURR_M),
]
for idx_m, (metric, formulas, fmt) in enumerate(sc_metrics):
    r += 1
    alt_fill = PatternFill("solid", fgColor="F5F5F5") if idx_m % 2 == 0 else None
    ws.cell(r, 2).value = metric; ws.cell(r, 2).font = bold_font
    if alt_fill: ws.cell(r, 2).fill = alt_fill
    ws.cell(r, 2).border = thin
    for i, formula in enumerate(formulas):
        cs_col = 5 + i*2
        if cs_col + 1 <= 16:
            ws.merge_cells(start_row=r, start_column=cs_col, end_row=r, end_column=min(cs_col+1, 16))
        c = ws.cell(r, cs_col)
        c.value = formula; c.font = label_font; c.number_format = fmt
        c.alignment = Alignment(horizontal="center")
        if alt_fill: c.fill = alt_fill
        c.border = thin
        # Highlight 10K column
        if i == 2:
            c.fill = PatternFill("solid", fgColor="E8F5E9")
            c.font = bold_font

# ════════════ DATA VALIDATION ════════════
dv = DataValidation(type="list", formula1="=$C$50:$G$50", allow_blank=False)
dv.prompt = "Select a rollout scenario"; dv.promptTitle = "Meter Rollout"
ws.add_data_validation(dv)
dv.add(ws['B6'])

# ════════════ FINALIZE ════════════
wb.move_sheet("Dashboard", offset=-3)
ws1.freeze_panes = "B4"
ws2.freeze_panes = "B3"

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.print_area = 'A1:Q44'
ws.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v7.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print("v7: Complete model with 7 cost categories, KPI cards, charts, scenario comparison")
