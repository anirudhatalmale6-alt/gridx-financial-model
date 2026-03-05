"""
GRIDx Financial Dashboard v8 — DARK NAVY THEME Dashboard
Matches client's ChatGPT-generated reference image:
  - Dark navy background (#1B2A4A) across entire dashboard
  - White text on dark background
  - KPI cards (Revenue, Profit, Margin, Payback) top-right
  - Revenue vs Costs bar chart center-top
  - Financial table with Years 1-5 columns
  - Scenario Comparison table at bottom
  - Input Parameters panel left side
Data sheets kept from v7 (Revenue, Cost Detail, 5 Year Model)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

SCENARIOS = ["3,000 Meters","5,000 Meters","10,000 Meters","20,000 Meters","50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]
NS = 5

HW_RATE=5500; INST_RATE=700; MAINT_RATE=300; WIFI_RATE=90; SMS_RATE=30; APP_RATE=624; RE_RATE=76.80
ANNUAL_RECURRING = HW_RATE+MAINT_RATE+WIFI_RATE+SMS_RATE+APP_RATE+RE_RATE
YEAR1_PER_METER = ANNUAL_RECURRING+INST_RATE
MFG_UNIT=[2000,2000,1700,1700,1462]; SPARE_RATE=4495*0.02

CLIENT_REV_5Y  = [101412000,169020000,338040000,676080000,1690200000]
CLIENT_COST_5Y = [26700000,33000000,41994047,62500000,109200000]
CLIENT_PROFIT  = [74700000,136000000,296045953,613600000,1580000000]
CLIENT_MARGIN  = [0.737,0.805,0.876,0.908,0.935]
CLIENT_PAYBACK = [7.6,7.2,4.2,3.9,3.5]
CLIENT_DCF     = [65500000,119300000,259647182,538200000,1390000000]

# ── Data sheet shared styles ──
NAVY="1B2A4A"; TEAL_C="2E75B6"
hdr_fill=PatternFill("solid",fgColor=NAVY); light_fill=PatternFill("solid",fgColor="D6E4F0")
green_fill=PatternFill("solid",fgColor="E2EFDA"); orange_fill=PatternFill("solid",fgColor="FCE4D6")
yellow_fill=PatternFill("solid",fgColor="FFF2CC")
hdr_font=Font(bold=True,color="FFFFFF",size=11,name="Calibri")
bold_font=Font(bold=True,size=10,name="Calibri"); bold11=Font(bold=True,size=11,name="Calibri")
label_font=Font(size=10,name="Calibri"); input_font=Font(color="0000FF",size=10,name="Calibri")
green_font=Font(color="008000",size=10,name="Calibri"); red_font=Font(color="C00000",size=10,name="Calibri")
title_font=Font(bold=True,color=NAVY,size=14,name="Calibri"); navy_bold=Font(bold=True,color=NAVY,size=12,name="Calibri")
thin=Border(left=Side("thin",color="D9D9D9"),right=Side("thin",color="D9D9D9"),top=Side("thin",color="D9D9D9"),bottom=Side("thin",color="D9D9D9"))
total_bdr=Border(top=Side("medium",color=NAVY),bottom=Side("double",color=NAVY))
NUM='#,##0'; PCT='0.0%'; CURR='N$#,##0'

def wh(ws,r,labels,cs=1):
    for i,l in enumerate(labels):
        c=ws.cell(r,cs+i,l);c.font=hdr_font;c.fill=hdr_fill;c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=thin
def wl(ws,r,c,t,f=label_font,fi=None):
    cell=ws.cell(r,c,t);cell.font=f
    if fi:cell.fill=fi
    cell.border=thin;return cell
def wi(ws,r,c,v,fmt=NUM):
    cell=ws.cell(r,c,v);cell.font=input_font;cell.fill=yellow_fill;cell.number_format=fmt;cell.border=thin;return cell
def wf(ws,r,c,formula,fmt=NUM,f=label_font,fi=None):
    cell=ws.cell(r,c,formula);cell.font=f
    if fi:cell.fill=fi
    cell.number_format=fmt;cell.border=thin;return cell
def fill_row(ws,r,cs,ce,fi):
    for c in range(cs,ce+1):ws.cell(r,c).fill=fi;ws.cell(r,c).border=thin

# ═══════════════════════════════════════
# SHEET 1: REVENUE
# ═══════════════════════════════════════
ws1=wb.active;ws1.title="Revenue";ws1.sheet_properties.tabColor="ED7D31"
ws1.column_dimensions['A'].width=40;ws1.column_dimensions['B'].width=22;ws1.column_dimensions['C'].width=16;ws1.column_dimensions['D'].width=22;ws1.column_dimensions['E'].width=22
wl(ws1,1,1,"Revenue Model",f=title_font)
wh(ws1,3,["Revenue Stream","Rate","Adoption","Basis","Annual per Meter (N$)"])
for r,name,rate,adopt,basis,annual in [
    (4,"Meter Hardware Sale","N$5,500/yr","100%","Annual per meter",HW_RATE),
    (5,"Installation Fee","N$700 one-time","100%","One-time per meter",INST_RATE),
    (6,"Maintenance","N$300/yr","100%","Annual",MAINT_RATE),
    (7,"Wi-Fi","N$25/mo","30%","30%×25×12",WIFI_RATE),
    (8,"SMS","N$2.50/SMS","50%","50%×2×2.50×12",SMS_RATE),
    (9,"Mobile App","N$65/mo","80%","80%×65×12",APP_RATE),
    (10,"Real Estate","N$3,200/block/mo","1/500m","(m÷500)×3200×12÷m",RE_RATE)]:
    wl(ws1,r,1,name);wl(ws1,r,2,rate);wl(ws1,r,3,adopt);wl(ws1,r,4,basis);wi(ws1,r,5,annual,'N$#,##0.00')
wl(ws1,11,1,"Annual Recurring per Meter",f=bold_font,fi=green_fill);fill_row(ws1,11,2,5,green_fill)
wf(ws1,11,5,"=E4+E6+E7+E8+E9+E10",'N$#,##0.00',bold_font,green_fill).border=total_bdr
wl(ws1,12,1,"Year 1 Total per Meter",f=bold_font,fi=green_fill);fill_row(ws1,12,2,5,green_fill)
wf(ws1,12,5,"=E11+E5",'N$#,##0.00',bold_font,green_fill).border=total_bdr
r=15;wl(ws1,r,1,"5-Year Revenue by Scenario",f=title_font);r=17;wh(ws1,r,["Revenue Stream"]+SCENARIOS)
streams=[("Hardware Sales",HW_RATE,0),("Installation Fees",0,INST_RATE),("Maintenance",MAINT_RATE,0),
    ("Wi-Fi",WIFI_RATE,0),("SMS",SMS_RATE,0),("Mobile App",APP_RATE,0),("Real Estate",RE_RATE,0)]
r=18
for name,annual,y1only in streams:
    wl(ws1,r,1,name)
    for i in range(NS):
        val=MC[i]*y1only if y1only else MC[i]*annual*5
        wi(ws1,r,2+i,val,CURR)
    r+=1
wl(ws1,r,1,"Total 5-Year Revenue",f=bold11,fi=green_fill);fill_row(ws1,r,2,6,green_fill)
for i in range(NS):
    col=get_column_letter(2+i);wf(ws1,r,2+i,f"=SUM({col}18:{col}24)",CURR,bold11,green_fill).border=total_bdr
REV_TOT=r

# ═══════════════════════════════════════
# SHEET 2: COST DETAIL
# ═══════════════════════════════════════
ws2=wb.create_sheet("Cost Detail");ws2.sheet_properties.tabColor="C00000"
ws2.column_dimensions['A'].width=42
for i in range(NS):ws2.column_dimensions[get_column_letter(2+i)].width=18
wl(ws2,1,1,"Complete Cost Structure — 5-Year",f=title_font)

# Manufacturing
r=3;wl(ws2,r,1,"MANUFACTURING COSTS",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2
wl(ws2,r,1,"Meter Hardware")
for i in range(NS):wi(ws2,r,2+i,MC[i]*MFG_UNIT[i],CURR)
r+=1;wl(ws2,r,1,"Spare Parts (2%)")
for i in range(NS):wi(ws2,r,2+i,round(MC[i]*SPARE_RATE),CURR)
r+=1;wl(ws2,r,1,"Total Manufacturing",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{r-2}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
MFG_TOT=r

# Operating
r+=2;wl(ws2,r,1,"OPERATING COSTS (5-Year)",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;os=r
for name,vals in [
    ("Personnel (N$748,800/yr)",[748800*5]*5),("Data Storage",[m*48*5 for m in MC]),
    ("Vehicles Maintenance",[v*30000*5 for v in[2,2,3,5,10]]),("Field Equipment",[k*10000*5 for k in[2,3,4,6,12]]),
    ("Office & Utilities",[240000*5]*5),("Software & Tools",[120000*5]*5),
    ("Marketing",[200000*5]*5),("Regulatory",[100000*5]*5),("Insurance",[150000*5]*5)]:
    wl(ws2,r,1,name)
    for i,v in enumerate(vals):wi(ws2,r,2+i,v,CURR)
    r+=1
wl(ws2,r,1,"Total Operating (5Y)",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{os}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
OPER_TOT=r

# Implementation
r+=2;wl(ws2,r,1,"IMPLEMENTATION (One-Time)",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;ims=r
for name,vals in [("Project Management",[500000]*5),("Site Surveys",[300000]*5),("Onboarding & Training",[250000]*5),("System Integration",[450000]*5)]:
    wl(ws2,r,1,name);[wi(ws2,r,2+i,v,CURR) for i,v in enumerate(vals)];r+=1
wl(ws2,r,1,"Total Implementation",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{ims}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
IMPL_TOT=r

# Equipment
r+=2;wl(ws2,r,1,"EQUIPMENT (One-Time)",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;eqs=r
for name,vals in [("Service Vehicles",[250000*v for v in[2,2,3,5,10]]),("Test Equipment",[85000*k for k in[2,3,4,6,12]]),
    ("Workstations",[15000*w for w in[3,5,7,12,25]]),("Network Monitoring",[120000*n for n in[1,1,2,3,5]]),
    ("Office Furniture",[50000+round(m/10000*230000) for m in MC]),("Server/Cloud",[100000+round(m/10000*150000) for m in MC])]:
    wl(ws2,r,1,name);[wi(ws2,r,2+i,v,CURR) for i,v in enumerate(vals)];r+=1
wl(ws2,r,1,"Total Equipment",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{eqs}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
EQUIP_TOT=r

# Development
r+=2;wl(ws2,r,1,"DEVELOPMENT COSTS",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;dvs=r
for name,vals in [("Electronics Mfg Dev",[312000]*5),("Software R&D",[777747]*5),("Mobile App Dev",[300000]*5),
    ("Firmware Dev",[2160000]*5),("Enclosure Mfg",[1403850]*5),("Marketing/Branding",[325000]*5),
    ("Regulatory/Legal",[1060000]*5),("Office/Logistics",[897400]*5),("Dev Contingency 5%",[615650]*5)]:
    wl(ws2,r,1,name);[wi(ws2,r,2+i,v,CURR) for i,v in enumerate(vals)];r+=1
wl(ws2,r,1,"Total Development",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{dvs}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
DEV_TOT=r

# Financing
r+=2;wl(ws2,r,1,"FINANCING",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;fns=r
for name,vals in [("Interest on Working Capital",[500000]*5),("Bank Fees",[200000]*5)]:
    wl(ws2,r,1,name);[wi(ws2,r,2+i,v,CURR) for i,v in enumerate(vals)];r+=1
wl(ws2,r,1,"Total Financing",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{fns}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
FIN_TOT=r

# Contingency
r+=2;wl(ws2,r,1,"CONTINGENCY",f=bold11,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
wh(ws2,r+1,["Cost Item"]+SCENARIOS);r+=2;cts=r
wl(ws2,r,1,"Ops Contingency 10%")
for i in range(NS):wf(ws2,r,2+i,f"={get_column_letter(2+i)}{OPER_TOT}*0.1",CURR)
r+=1;wl(ws2,r,1,"Impl Contingency 10%")
for i in range(NS):wf(ws2,r,2+i,f"={get_column_letter(2+i)}{IMPL_TOT}*0.1",CURR)
r+=1;wl(ws2,r,1,"Total Contingency",f=bold_font,fi=orange_fill);fill_row(ws2,r,2,6,orange_fill)
for i in range(NS):col=get_column_letter(2+i);wf(ws2,r,2+i,f"=SUM({col}{cts}:{col}{r-1})",CURR,bold_font,orange_fill).border=total_bdr
CONT_TOT=r

# Grand Total
r+=2;wl(ws2,r,1,"TOTAL 5-YEAR COSTS",f=navy_bold,fi=green_fill);fill_row(ws2,r,2,6,green_fill)
cost_tots=[MFG_TOT,OPER_TOT,IMPL_TOT,EQUIP_TOT,DEV_TOT,FIN_TOT,CONT_TOT]
for i in range(NS):
    col=get_column_letter(2+i);refs="+".join([f"{col}{cr}" for cr in cost_tots])
    wf(ws2,r,2+i,f"={refs}",CURR,navy_bold,green_fill).border=total_bdr
GRAND_TOT=r
print(f"Cost: Mfg={MFG_TOT} Oper={OPER_TOT} Impl={IMPL_TOT} Equip={EQUIP_TOT} Dev={DEV_TOT} Fin={FIN_TOT} Cont={CONT_TOT} Grand={GRAND_TOT}")

# ═══════════════════════════════════════
# SHEET 3: DASHBOARD (DARK NAVY THEME)
# ═══════════════════════════════════════
ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "002060"

# ── Dark theme colors ──
BG1 = "0F1B2D"   # Darkest background
BG2 = "162236"   # Dark panel background
BG3 = "1D2B42"   # Medium dark (card bg)
BG4 = "243B55"   # Lighter dark (table header)
ACC_BLUE = "3B82F6"   # Accent blue
ACC_GREEN = "22C55E"  # Accent green
ACC_RED = "EF4444"    # Accent red/cost
GOLD = "F59E0B"       # Gold accent
WHITE = "FFFFFF"
LGRAY = "94A3B8"      # Light gray for secondary text
BORDER_C = "2D3F59"   # Border color

bg1_fill = PatternFill("solid", fgColor=BG1)
bg2_fill = PatternFill("solid", fgColor=BG2)
bg3_fill = PatternFill("solid", fgColor=BG3)
bg4_fill = PatternFill("solid", fgColor=BG4)
acc_blue_fill = PatternFill("solid", fgColor=ACC_BLUE)
acc_green_fill = PatternFill("solid", fgColor="1B4332")
kpi_rev_fill = PatternFill("solid", fgColor="1E3A5F")
kpi_prof_fill = PatternFill("solid", fgColor="1E3A5F")
kpi_marg_fill = PatternFill("solid", fgColor="1E3A5F")
kpi_pay_fill = PatternFill("solid", fgColor="1E3A5F")
tbl_hdr_fill = PatternFill("solid", fgColor=BG4)
tbl_row1_fill = PatternFill("solid", fgColor=BG2)
tbl_row2_fill = PatternFill("solid", fgColor=BG3)
green_num_fill = PatternFill("solid", fgColor="0F3D2E")
red_num_fill = PatternFill("solid", fgColor="3D1F1F")

# Fonts
f_title = Font(bold=True, size=16, name="Calibri", color=WHITE)
f_sub = Font(bold=True, size=11, name="Calibri", color=LGRAY)
f_section = Font(bold=True, size=11, name="Calibri", color=ACC_BLUE)
f_white = Font(size=10, name="Calibri", color=WHITE)
f_white_b = Font(bold=True, size=10, name="Calibri", color=WHITE)
f_white_lg = Font(bold=True, size=14, name="Calibri", color=WHITE)
f_kpi_val = Font(bold=True, size=20, name="Calibri", color=WHITE)
f_kpi_lbl = Font(size=9, name="Calibri", color=LGRAY)
f_green = Font(bold=True, size=10, name="Calibri", color=ACC_GREEN)
f_red = Font(size=10, name="Calibri", color=ACC_RED)
f_gold = Font(bold=True, size=10, name="Calibri", color=GOLD)
f_drop = Font(bold=True, size=12, name="Calibri", color=WHITE)
f_tiny = Font(size=7, name="Calibri", color=BG2)  # hidden data

d_bdr = Border(left=Side("thin",color=BORDER_C),right=Side("thin",color=BORDER_C),
    top=Side("thin",color=BORDER_C),bottom=Side("thin",color=BORDER_C))

# Column widths
ws.column_dimensions['A'].width = 1.5
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 1
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 1
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 12
ws.column_dimensions['N'].width = 12
ws.column_dimensions['O'].width = 12
ws.column_dimensions['P'].width = 1.5

# Fill ENTIRE visible area with dark background
for r in range(1, 50):
    ws.row_dimensions[r].height = 18
    for c in range(1, 17):
        ws.cell(r, c).fill = bg1_fill

# ════════════ ROW 2: TITLE BAR ════════════
ws.row_dimensions[2].height = 32
ws.merge_cells('B2:O2')
ws['B2'] = "GRIDx Smart Meter - 5-Year Financial Model"
ws['B2'].font = f_title; ws['B2'].fill = bg2_fill
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 16): ws.cell(2, c).fill = bg2_fill

# ════════════ ROW 3: SCENARIO SELECTOR ════════════
ws.row_dimensions[3].height = 28
ws.merge_cells('F3:I3')
ws['F3'] = "10,000 Meters"
ws['F3'].font = f_drop; ws['F3'].fill = acc_blue_fill
ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(6, 10): ws.cell(3, c).fill = acc_blue_fill; ws.cell(3, c).border = d_bdr

ws.merge_cells('L3:O3')
ws['L3'] = "Recommended"
ws['L3'].font = Font(italic=True, size=10, name="Calibri", color=GOLD)
ws['L3'].alignment = Alignment(horizontal="center", vertical="center")

# ════════════ LEFT: INPUT PARAMETERS (rows 5-18) ════════════
ws.row_dimensions[5].height = 22
ws.merge_cells('B5:D5')
ws['B5'] = "1. Input Parameters"
ws['B5'].font = f_section

# Dropdown label
ws['B6'] = "Rollout Scenario:"; ws['B6'].font = f_white
ws.merge_cells('C6:D6')
ws['C6'] = "10,000 Meters"; ws['C6'].font = f_drop; ws['C6'].fill = PatternFill("solid", fgColor="2563EB")
ws['C6'].alignment = Alignment(horizontal="center"); ws['C6'].border = d_bdr
ws['D6'].fill = PatternFill("solid", fgColor="2563EB"); ws['D6'].border = d_bdr

# Parameter values
params = [
    (8, "Hardware Price:", f"N$5,500", "/unit/yr"),
    (9, "Installation:", f"N$700", "one-time"),
    (10, "Mfg Cost/Meter:", None, "per unit"),
    (11, "Personnel:", "N$748,800", "/year"),
    (12, "Annual Revenue:", None, "/year"),
    (13, "Total 5Y Revenue:", None, "5-year"),
    (14, "Total 5Y Costs:", None, "5-year"),
    (15, "Net Profit:", None, "5-year"),
]
for pr, label, val, unit in params:
    ws.cell(pr, 2).value = label; ws.cell(pr, 2).font = f_white
    if val:
        ws.cell(pr, 3).value = val; ws.cell(pr, 3).font = f_green
    ws.cell(pr, 4).value = unit; ws.cell(pr, 4).font = Font(size=8, color=LGRAY, name="Calibri")

# Dynamic formulas for input panel
IDX = "$C$51"
def idx(rr): return f"INDEX($C${rr}:$G${rr},1,{IDX})"

ws['C10'] = f'="N$"&TEXT({idx(53)},"#,##0")'; ws['C10'].font = f_green
ws['C12'] = f'="N$"&TEXT({idx(52)}*{ANNUAL_RECURRING},"#,##0")'; ws['C12'].font = f_green
ws['C13'] = f'="N$"&TEXT({idx(63)},"#,##0")'; ws['C13'].font = f_green
ws['C14'] = f'="N$"&TEXT({idx(61)},"#,##0")'; ws['C14'].font = f_red
ws['C15'] = f'="N$"&TEXT({idx(62)},"#,##0")'; ws['C15'].font = f_gold

# ════════════ KPI CARDS (rows 5-8, cols L-O) ════════════
kpis = [
    (5, 12, 13, "N$338M", "Total Revenue", f'="N$"&TEXT({idx(63)}/1000000,"#,##0")&"M"', kpi_rev_fill, ACC_GREEN),
    (5, 14, 15, "N$296M", "Net Profit", f'="N$"&TEXT({idx(62)}/1000000,"#,##0")&"M"', kpi_prof_fill, ACC_GREEN),
    (7, 12, 13, "87.6%", "Gross Margin", f'=TEXT({idx(64)}*100,"0.0")&"%"', kpi_marg_fill, GOLD),
    (7, 14, 15, "4.2 mo", "Payback Period", f'=TEXT({idx(65)},"0.0")&" months"', kpi_pay_fill, ACC_BLUE),
]
for kr, kc1, kc2, default, label, formula, bg, accent_c in kpis:
    for r in range(kr, kr+2):
        for c in range(kc1, kc2+1):
            ws.cell(r, c).fill = bg; ws.cell(r, c).border = d_bdr
    ws.merge_cells(start_row=kr, start_column=kc1, end_row=kr, end_column=kc2)
    ws.cell(kr, kc1).value = formula
    ws.cell(kr, kc1).font = Font(bold=True, size=18, name="Calibri", color=accent_c)
    ws.cell(kr, kc1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=kr+1, start_column=kc1, end_row=kr+1, end_column=kc2)
    ws.cell(kr+1, kc1).value = label
    ws.cell(kr+1, kc1).font = f_kpi_lbl
    ws.cell(kr+1, kc1).alignment = Alignment(horizontal="center")

# ════════════ HIDDEN DATA (rows 50+) ════════════
for i, s in enumerate(SCENARIOS): ws.cell(50, 3+i, s).font = f_tiny
ws.cell(51, 3).value = '=MATCH($C$6,$C$50:$G$50,0)'; ws.cell(51, 3).font = f_tiny
for i, m in enumerate(MC): ws.cell(52, 3+i, m).font = f_tiny
for i in range(NS): ws.cell(53, 3+i, MFG_UNIT[i]).font = f_tiny

# Cost totals from Cost Detail
for i in range(NS):
    col = get_column_letter(2+i)
    ws.cell(54, 3+i).value = f"='Cost Detail'!{col}{MFG_TOT}"; ws.cell(54, 3+i).font = f_tiny
    ws.cell(55, 3+i).value = f"='Cost Detail'!{col}{OPER_TOT}"; ws.cell(55, 3+i).font = f_tiny
    ws.cell(56, 3+i).value = f"='Cost Detail'!{col}{IMPL_TOT}"; ws.cell(56, 3+i).font = f_tiny
    ws.cell(57, 3+i).value = f"='Cost Detail'!{col}{EQUIP_TOT}"; ws.cell(57, 3+i).font = f_tiny
    ws.cell(58, 3+i).value = f"='Cost Detail'!{col}{DEV_TOT}"; ws.cell(58, 3+i).font = f_tiny
    ws.cell(59, 3+i).value = f"='Cost Detail'!{col}{FIN_TOT}"; ws.cell(59, 3+i).font = f_tiny
    ws.cell(60, 3+i).value = f"='Cost Detail'!{col}{CONT_TOT}"; ws.cell(60, 3+i).font = f_tiny
    ws.cell(61, 3+i).value = f"='Cost Detail'!{col}{GRAND_TOT}"; ws.cell(61, 3+i).font = f_tiny
    ws.cell(62, 3+i).value = f"='Revenue'!{col}{REV_TOT}-{get_column_letter(3+i)}61"; ws.cell(62, 3+i).font = f_tiny
    ws.cell(63, 3+i).value = f"='Revenue'!{col}{REV_TOT}"; ws.cell(63, 3+i).font = f_tiny
    ws.cell(64, 3+i).value = f"=IF({get_column_letter(3+i)}63>0,{get_column_letter(3+i)}62/{get_column_letter(3+i)}63,0)"
    ws.cell(64, 3+i).font = f_tiny

for i in range(NS):
    ws.cell(65, 3+i, CLIENT_PAYBACK[i]).font = f_tiny
    ws.cell(66, 3+i, CLIENT_DCF[i]).font = f_tiny

# Year-by-year chart data (millions)
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(70, 4+i, yr).font = f_tiny

# Revenue by year (M)
ws.cell(71, 3, "RevM").font = f_tiny
ws.cell(71, 4).value = f"={idx(52)}*{YEAR1_PER_METER}/1000000"
for yc in range(5, 9): ws.cell(71, yc).value = f"={idx(52)}*{ANNUAL_RECURRING}/1000000"

# Costs by year (M)
ws.cell(72, 3, "CostM").font = f_tiny
ws.cell(72, 4).value = f"=({idx(54)}+{idx(55)}/5+{idx(56)}+{idx(57)}+{idx(58)}+{idx(59)}+{idx(60)}/5)/1000000"
for yc in range(5, 9): ws.cell(72, yc).value = f"=({idx(55)}/5+{idx(60)}/5)/1000000"

# Profit by year (M)
ws.cell(73, 3, "ProfM").font = f_tiny
for yc in range(4, 9): ws.cell(73, yc).value = f"={get_column_letter(yc)}71-{get_column_letter(yc)}72"

# ════════════ CHART: Revenue vs Costs (F5:J18) ════════════
cats = Reference(ws, min_col=4, min_row=70, max_col=8, max_row=70)

bar1 = BarChart()
bar1.type = "col"; bar1.grouping = "clustered"
bar1.title = "Revenue vs. Costs (N$ Millions)"
bar1.style = 10
bar1.y_axis.numFmt = '0'; bar1.y_axis.title = "N$ Millions"

d_rev = Reference(ws, min_col=4, min_row=71, max_col=8, max_row=71)
bar1.add_data(d_rev, from_rows=True)
bar1.series[0].title = SeriesLabel(v="Revenue")
bar1.series[0].graphicalProperties.solidFill = "3B82F6"

d_cost = Reference(ws, min_col=4, min_row=72, max_col=8, max_row=72)
bar1.add_data(d_cost, from_rows=True)
bar1.series[1].title = SeriesLabel(v="Costs")
bar1.series[1].graphicalProperties.solidFill = "22C55E"

bar1.set_categories(cats)
bar1.legend.position = 'b'
bar1.width = 18; bar1.height = 12
ws.add_chart(bar1, "F5")

# ════════════ FINANCIAL TABLE (rows 19-28) ════════════
ws.row_dimensions[19].height = 22
# Table header
tbl_cols = ['B','C','D','F','G','H','I','J','L','M','N','O']
hdr_labels = ["Metric","","","Year 1","Year 2","Year 3","Year 4","Year 5","","10,000 M","",""]

# Actually, let me use a cleaner table structure
# Cols B-D: Metric name (merged)
# Cols F-J: Year 1-5 values
# Cols L-O: 5-Year totals or scenario-specific

r = 19
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = f_white_b; ws.cell(r, 2).fill = tbl_hdr_fill
for c in [2,3,4]: ws.cell(r, c).fill = tbl_hdr_fill; ws.cell(r, c).border = d_bdr
yr_cols = [6,7,8,9,10]  # F,G,H,I,J
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(r, yr_cols[i]).value = yr; ws.cell(r, yr_cols[i]).font = f_white_b
    ws.cell(r, yr_cols[i]).fill = tbl_hdr_fill; ws.cell(r, yr_cols[i]).border = d_bdr
    ws.cell(r, yr_cols[i]).alignment = Alignment(horizontal="center")
# 5-Year Total column
ws.merge_cells(f'L{r}:O{r}')
ws.cell(r, 12).value = "5-Year Total"; ws.cell(r, 12).font = f_white_b
ws.cell(r, 12).fill = tbl_hdr_fill; ws.cell(r, 12).border = d_bdr
ws.cell(r, 12).alignment = Alignment(horizontal="center")
for c in [13,14,15]: ws.cell(r, c).fill = tbl_hdr_fill; ws.cell(r, c).border = d_bdr

# Table data rows
table_data = [
    ("Hardware Sales", [f"={idx(52)}*{HW_RATE}"]*5, f"={idx(52)}*{HW_RATE}*5", True, False),
    ("Installation Fees", [f"={idx(52)}*{INST_RATE}","0","0","0","0"], f"={idx(52)}*{INST_RATE}", True, False),
    ("Service Revenue", [f"={idx(52)}*{MAINT_RATE+WIFI_RATE+SMS_RATE+APP_RATE+RE_RATE}"]*5,
        f"={idx(52)}*{MAINT_RATE+WIFI_RATE+SMS_RATE+APP_RATE+RE_RATE}*5", True, False),
    ("Manufacturing Costs", [f"=-{idx(54)}","0","0","0","0"], f"=-{idx(54)}", False, True),
    ("Operating Costs", [f"=-{idx(55)}/5"]*5, f"=-{idx(55)}", False, True),
    ("Net Profit", None, None, True, False),  # calculated
]

for ri, (label, yr_formulas, total_formula, is_positive, is_cost) in enumerate(table_data):
    r = 20 + ri
    row_fill = tbl_row1_fill if ri % 2 == 0 else tbl_row2_fill
    val_font = f_green if is_positive else f_red

    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = label; ws.cell(r, 2).font = f_white_b if label == "Net Profit" else f_white
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = d_bdr

    if label == "Net Profit":
        # Sum of revenue rows minus cost rows
        for ci, yc in enumerate(yr_cols):
            ws.cell(r, yc).value = f"={get_column_letter(yc)}20+{get_column_letter(yc)}21+{get_column_letter(yc)}22+{get_column_letter(yc)}23+{get_column_letter(yc)}24"
            ws.cell(r, yc).font = f_gold; ws.cell(r, yc).number_format = '"N$"#,##0.0,,"M"'
            ws.cell(r, yc).fill = row_fill; ws.cell(r, yc).border = d_bdr
            ws.cell(r, yc).alignment = Alignment(horizontal="center")
        # 5-year total
        ws.merge_cells(f'L{r}:O{r}')
        ws.cell(r, 12).value = f"={idx(62)}"
        ws.cell(r, 12).font = Font(bold=True, size=12, color=GOLD, name="Calibri")
        ws.cell(r, 12).number_format = '"N$"#,##0.0,,"M"'
        ws.cell(r, 12).fill = row_fill; ws.cell(r, 12).border = d_bdr
        ws.cell(r, 12).alignment = Alignment(horizontal="center")
        for c in [13,14,15]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = d_bdr
    else:
        for ci, yc in enumerate(yr_cols):
            ws.cell(r, yc).value = yr_formulas[ci]
            ws.cell(r, yc).font = val_font; ws.cell(r, yc).number_format = '"N$"#,##0.0,,"M"'
            ws.cell(r, yc).fill = row_fill; ws.cell(r, yc).border = d_bdr
            ws.cell(r, yc).alignment = Alignment(horizontal="center")
        ws.merge_cells(f'L{r}:O{r}')
        ws.cell(r, 12).value = total_formula
        ws.cell(r, 12).font = val_font; ws.cell(r, 12).number_format = '"N$"#,##0.0,,"M"'
        ws.cell(r, 12).fill = row_fill; ws.cell(r, 12).border = d_bdr
        ws.cell(r, 12).alignment = Alignment(horizontal="center")
        for c in [13,14,15]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = d_bdr

# ════════════ SCENARIO COMPARISON (rows 27-34) ════════════
r = 27
ws.merge_cells(f'B{r}:O{r}')
ws.cell(r, 2).value = "Scenario Comparison"
ws.cell(r, 2).font = f_section
ws.cell(r, 2).fill = bg1_fill
for c in range(2, 16): ws.cell(r, c).fill = bg1_fill

r = 28
# Header
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = f_white_b; ws.cell(r, 2).fill = tbl_hdr_fill
for c in [2,3,4]: ws.cell(r, c).fill = tbl_hdr_fill; ws.cell(r, c).border = d_bdr
sc_cols = [6,8,10,12,14]  # spread across
sc_short = ["3K","5K","10K","20K","50K"]
for i, (sc, label) in enumerate(zip(sc_cols, sc_short)):
    ws.cell(r, sc).value = label; ws.cell(r, sc).font = f_white_b
    ws.cell(r, sc).fill = tbl_hdr_fill; ws.cell(r, sc).border = d_bdr
    ws.cell(r, sc).alignment = Alignment(horizontal="center")

sc_rows = [
    ("5-Year Revenue", [f"=C63",f"=D63",f"=E63",f"=F63",f"=G63"], '"N$"#,##0.0,,"M"', True),
    ("5-Year Costs", [f"=C61",f"=D61",f"=E61",f"=F61",f"=G61"], '"N$"#,##0.0,,"M"', False),
    ("Net Profit", [f"=C62",f"=D62",f"=E62",f"=F62",f"=G62"], '"N$"#,##0.0,,"M"', True),
    ("Gross Margin", [f"=C64",f"=D64",f"=E64",f"=F64",f"=G64"], '0.0%', True),
    ("Payback", [f"=C65",f"=D65",f"=E65",f"=F65",f"=G65"], '0.0" mo"', True),
]
for ri, (metric, formulas, fmt, positive) in enumerate(sc_rows):
    r = 29 + ri
    row_fill = tbl_row1_fill if ri % 2 == 0 else tbl_row2_fill
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = metric; ws.cell(r, 2).font = f_white
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = d_bdr
    for i, sc in enumerate(sc_cols):
        ws.cell(r, sc).value = formulas[i]
        ws.cell(r, sc).font = f_green if positive else f_red
        ws.cell(r, sc).number_format = fmt
        ws.cell(r, sc).fill = row_fill; ws.cell(r, sc).border = d_bdr
        ws.cell(r, sc).alignment = Alignment(horizontal="center")
        # Highlight 10K column
        if i == 2:
            ws.cell(r, sc).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.cell(r, sc).font = Font(bold=True, size=10, name="Calibri", color=ACC_GREEN if positive else ACC_RED)

# ════════════ DATA VALIDATION ════════════
dv = DataValidation(type="list", formula1="=$C$50:$G$50", allow_blank=False)
dv.prompt = "Select scenario"; dv.promptTitle = "Rollout"
ws.add_data_validation(dv); dv.add(ws['C6'])

# ════════════ FINALIZE ════════════
wb.move_sheet("Dashboard", offset=-2)
ws1.freeze_panes = "B4"
ws2.freeze_panes = "B3"

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v8.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print("v8: Dark navy theme dashboard matching client reference")
