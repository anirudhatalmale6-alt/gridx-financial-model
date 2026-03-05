"""
GRIDx Financial Dashboard v4 — Interactive Dashboard
Features:
  - Dropdown selector for rollout scenario (3K/5K/10K/20K/50K meters)
  - LEFT: Parameters + 5-year year-by-year financial table (dynamic)
  - RIGHT: 3 charts that auto-update when dropdown changes
  - All formulas use INDEX/MATCH linked to dropdown
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, AreaChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
SCENARIOS = ["3,000 Meters", "5,000 Meters", "10,000 Meters", "20,000 Meters", "50,000 Meters"]
MC = [3000, 5000, 10000, 20000, 50000]
NS = 5

# ── Shared styles ──
NAVY = "1B2A4A"; DARK_BLUE = "2C3E6B"; TEAL = "2E75B6"
hdr_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor="D6E4F0")
green_fill = PatternFill("solid", fgColor="E2EFDA")
orange_fill = PatternFill("solid", fgColor="FCE4D6")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bold_font = Font(bold=True, size=10, name="Calibri")
bold11 = Font(bold=True, size=11, name="Calibri")
label_font = Font(size=10, name="Calibri")
input_font = Font(color="0000FF", size=10, name="Calibri")
green_font = Font(color="008000", size=10, name="Calibri")
title_font = Font(bold=True, color=NAVY, size=14, name="Calibri")
navy_bold = Font(bold=True, color=NAVY, size=12, name="Calibri")

thin = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
total_bdr = Border(top=Side("medium", color=NAVY), bottom=Side("double", color=NAVY))

NUM = '#,##0'; PCT = '0.0%'; CURR = 'N$#,##0'; CURR_N = 'N$#,##0;(N$#,##0);"-"'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=r, column=cs+i, value=l)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin

def wl(ws, r, c, t, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=t); cell.font = f
    if fi: cell.fill = fi
    cell.border = thin; return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(row=r, column=c, value=v); cell.font = input_font; cell.fill = yellow_fill; cell.number_format = fmt; cell.border = thin; return cell

def wf(ws, r, c, formula, fmt=NUM, f=label_font, fi=None):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = f
    if fi: cell.fill = fi
    cell.number_format = fmt; cell.border = thin; return cell

def fill_row(ws, r, cs, ce, fi):
    for c in range(cs, ce+1):
        ws.cell(r, c).fill = fi; ws.cell(r, c).border = thin

# ═══════════════════════════════════════════════════════════════
# SHEET 1: PERSONNEL COST
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Personnel Cost"; ws1.sheet_properties.tabColor = NAVY
ws1.column_dimensions['A'].width = 36
for i in range(NS): ws1.column_dimensions[get_column_letter(2+i)].width = 16
ws1.column_dimensions['H'].width = 30; ws1.column_dimensions['I'].width = 18; ws1.column_dimensions['J'].width = 18

wh(ws1, 1, ["Resource Type"] + SCENARIOS)
ws1.cell(1,1).alignment = Alignment(horizontal="left")
wl(ws1, 1, 8, "Position", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 9, "Monthly Salary (N$)", f=hdr_font, fi=hdr_fill)
wl(ws1, 1, 10, "Annual Salary (N$)", f=hdr_font, fi=hdr_fill)

departments = [
    ("LEADERSHIP & MANAGEMENT", [
        ("Chief Executive Officer (CEO)", [1,1,1,1,1], 35000),
        ("Chief Operations Officer (COO)", [0,0,1,1,1], 28000),
        ("Chief Financial Officer (CFO)", [0,1,1,1,1], 28000),
        ("Chief Technology Officer (CTO)", [0,1,1,1,1], 30000),
        ("Head of Business Development", [0,0,1,1,1], 25000),
        ("Head of Human Resources", [0,0,0,1,1], 22000),
        ("Legal & Compliance Officer", [0,0,0,1,1], 22000),
    ]),
    ("OPERATIONS & FIELD SERVICES", [
        ("Operations Manager", [0,1,1,1,1], 20000),
        ("Field Support Technicians", [2,3,4,5,8], 9000),
        ("Installation Team Lead", [0,1,1,2,3], 11000),
        ("Logistics Coordinator", [0,0,1,1,2], 8000),
    ]),
    ("ENGINEERING & QUALITY", [
        ("Engineering Manager", [0,0,1,1,1], 22000),
        ("Quality Assurance Engineers", [1,1,1,3,5], 12000),
        ("Firmware/Software Engineers", [0,1,2,3,5], 15000),
        ("Hardware Engineers", [0,0,1,2,3], 15000),
    ]),
    ("CUSTOMER SUCCESS & SUPPORT", [
        ("Customer Support Manager", [0,0,1,1,1], 18000),
        ("Customer Support Agents", [1,2,2,4,8], 7200),
        ("Technical Support Specialists", [0,0,1,2,3], 9000),
    ]),
    ("SALES & MARKETING", [
        ("Sales Manager", [0,0,1,1,1], 20000),
        ("Sales Representatives", [0,1,2,3,5], 10000),
        ("Marketing Coordinator", [0,0,1,1,2], 9000),
    ]),
    ("FINANCE & ADMINISTRATION", [
        ("Finance Manager", [0,0,1,1,1], 20000),
        ("Accountant", [0,1,1,2,3], 11000),
        ("Administrative Assistant", [0,0,1,2,3], 7000),
    ]),
]

r = 2; sal_row = 2
subtotal_rows = []
position_rows = {}

for dept_name, positions in departments:
    wl(ws1, r, 1, dept_name, f=bold11, fi=light_fill)
    fill_row(ws1, r, 2, 6, light_fill)
    r += 1
    start = r
    for name, counts, salary in positions:
        wl(ws1, r, 1, name)
        for i, cnt in enumerate(counts):
            wi(ws1, r, 2+i, cnt, NUM)
        position_rows[name] = r
        wl(ws1, sal_row, 8, name); wi(ws1, sal_row, 9, salary, CURR)
        wf(ws1, sal_row, 10, f"=I{sal_row}*12", CURR)
        sal_row += 1
        r += 1
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, f"Subtotal – {short}", f=bold_font)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws1, r, 2+i, f"=SUM({col}{start}:{col}{r-1})", NUM, bold_font)
    subtotal_rows.append(r)
    r += 2

wl(ws1, r, 1, "TOTAL PERSONNEL", f=navy_bold, fi=green_fill)
fill_row(ws1, r, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{col}{sr}" for sr in subtotal_rows])
    cell = wf(ws1, r, 2+i, f"={refs}", NUM, navy_bold, green_fill)
    cell.border = total_bdr
TOTAL_PERS_ROW = r

r += 3
wl(ws1, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=navy_bold)
r += 1
wh(ws1, r, ["Position", "Unit Cost (N$/year)"] + SCENARIOS)
r += 1

cost_subtotal_rows = []
for dept_name, positions in departments:
    short = dept_name.split(" &")[0].split(" ")[0]
    wl(ws1, r, 1, dept_name, f=bold_font, fi=light_fill)
    fill_row(ws1, r, 2, 2+NS, light_fill)
    r += 1
    dept_start = r
    for name, counts, salary in positions:
        hc_row = position_rows[name]
        annual = salary * 12
        wl(ws1, r, 1, name)
        wi(ws1, r, 2, annual, CURR)
        for i in range(NS):
            hc_col = get_column_letter(2+i)
            wf(ws1, r, 3+i, f"={hc_col}{hc_row}*$B{r}", CURR, green_font)
        r += 1
    wl(ws1, r, 1, f"{short} Subtotal", f=bold_font)
    for i in range(NS):
        col = get_column_letter(3+i)
        wf(ws1, r, 3+i, f"=SUM({col}{dept_start}:{col}{r-1})", CURR, bold_font)
    cost_subtotal_rows.append(r)
    r += 2

wl(ws1, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=navy_bold, fi=green_fill)
fill_row(ws1, r, 2, 2+NS, green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    refs = "+".join([f"{col}{sr}" for sr in cost_subtotal_rows])
    cell = wf(ws1, r, 3+i, f"={refs}", CURR, navy_bold, green_fill)
    cell.border = total_bdr
TOTAL_COST_ROW = r
print(f"Personnel: total headcount row={TOTAL_PERS_ROW}, cost row={TOTAL_COST_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: INFRASTRUCTURE & EQUIPMENT COST
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Infrastructure & Equipment"); ws2.sheet_properties.tabColor = TEAL
ws2.column_dimensions['A'].width = 36; ws2.column_dimensions['B'].width = 18
for i in range(NS): ws2.column_dimensions[get_column_letter(3+i)].width = 18

wl(ws2, 1, 1, "Infrastructure Costs (Annual)", f=title_font)
wh(ws2, 3, ["Cost Category", "Unit Cost"] + SCENARIOS)

infra = [
    ("Data Storage & Transmission", "N$48/meter", [144000,240000,480000,960000,2400000]),
    ("Office Space Rental", "N$200/sq m", [60000,120000,240000,360000,600000]),
    ("Utilities & Connectivity", "Lump sum", [36000,60000,120000,240000,480000]),
]
r = 4
for name, unit, vals in infra:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Annual Infrastructure Total", f=bold_font, fi=green_fill)
wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}4:{col}6)", CURR, bold_font, green_fill)
    cell.border = total_bdr
INFRA_TOT = r

r += 2
wl(ws2, r, 1, "Equipment Costs (One-Time Investment)", f=title_font)
r += 2
wh(ws2, r, ["Equipment Type", "Unit Cost"] + SCENARIOS)
r += 1

equip = [
    ("Service Vehicles", "N$250,000", [250000,500000,750000,750000,1250000]),
    ("Field Test Equipment Kits", "N$85,000", [85000,170000,340000,680000,1530000]),
    ("Spare Parts Inventory", "N$4,495 × 2%", [269700,449500,899000,1798000,4495000]),
    ("Workstations", "N$15,000", [75000,210000,420000,630000,930000]),
    ("Network Monitoring Tools", "N$120,000", [120000,120000,240000,360000,600000]),
    ("Office Equipment & Furniture", "Per employee", [50000,140000,280000,420000,620000]),
    ("Server/Cloud Infrastructure", "Lump sum", [100000,150000,250000,400000,800000]),
]
eq_start = r
for name, unit, vals in equip:
    wl(ws2, r, 1, name); wl(ws2, r, 2, unit)
    for i, v in enumerate(vals): wi(ws2, r, 3+i, v, CURR)
    r += 1
wl(ws2, r, 1, "Total Equipment Investment", f=bold_font, fi=green_fill)
wl(ws2, r, 2, "", fi=green_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    cell = wf(ws2, r, 3+i, f"=SUM({col}{eq_start}:{col}{r-1})", CURR, bold_font, green_fill)
    cell.border = total_bdr
EQUIP_TOT = r
print(f"Infra: {INFRA_TOT}, Equip: {EQUIP_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: MANUFACTURING COST
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Manufacturing Cost"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 28; ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 20; ws3.column_dimensions['D'].width = 36

wl(ws3, 1, 1, "Volume Discount Structure", f=title_font)
wh(ws3, 3, ["Volume Tier", "Discount", "Cost per Meter (N$)", "Calculation Basis"])
for i, (name, disc, cost, basis) in enumerate([
    ("Base (0-9,999 meters)", 0, 2000, "N$2,000 per meter"),
    ("10,000-19,999 meters", 0.15, 1700, "15% reduction from base"),
    ("20,000-49,999 meters", 0.30, 1400, "Additional 15% reduction"),
    ("50,000+ meters", 0.45, 1100, "Additional 15% reduction"),
]):
    wl(ws3, 4+i, 1, name); wi(ws3, 4+i, 2, disc, PCT); wi(ws3, 4+i, 3, cost, CURR); wl(ws3, 4+i, 4, basis)

wl(ws3, 10, 1, "Manufacturing Cost by Scenario", f=title_font)
wh(ws3, 12, ["Scenario", "Meters", "Cost per Meter (N$)", "Total Manufacturing Cost (N$)"])
mfg_data = [(3000, 2000), (5000, 2000), (10000, 1700), (20000, 1700), (50000, 1462)]
MFG_S = 13
for i, (m, c) in enumerate(mfg_data):
    r = MFG_S + i
    wl(ws3, r, 1, SCENARIOS[i]); wi(ws3, r, 2, m, NUM); wi(ws3, r, 3, c, CURR)
    wf(ws3, r, 4, f"=B{r}*C{r}", CURR, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 4: REVENUE
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Revenue"); ws4.sheet_properties.tabColor = "ED7D31"
ws4.column_dimensions['A'].width = 36; ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 22; ws4.column_dimensions['D'].width = 36; ws4.column_dimensions['E'].width = 24

wl(ws4, 1, 1, "Revenue Model Assumptions", f=title_font)
wh(ws4, 3, ["Revenue Stream", "Rate", "Adoption Rate", "Calculation Basis", "Annual Amount per Meter (N$)"])

rev_a = [
    (4, "Meter Hardware Sale", "N$5,500 (one-time)", 1.0, "One-time revenue per installed meter", 5500),
    (5, "Meter Installation Fee", "N$700 (one-time)", 1.0, "One-time fee per new meter installation", 700),
    (6, "Meter Maintenance", "N$500 / year", 0.05, "Annual maintenance fee on installed meters", 25),
    (7, "Wi-Fi Subscription", "N$25 / month", "30% of installed base", "Recurring", 90),
    (8, "SMS Notifications", "N$2.50 / SMS", "50% adoption, 2 SMS/month", "Recurring", 30),
    (9, "Mobile App Subscription", "N$65 / month", "80% of installed base", "Recurring", 624),
    (10, "Real Estate Management Platform", "N$3,200 / block / month", "1 block per 500 meters", "Recurring", 76.80),
]
for r, name, rate, adoption, basis, annual in rev_a:
    wl(ws4, r, 1, name); wl(ws4, r, 2, rate)
    if isinstance(adoption, float): wi(ws4, r, 3, adoption, PCT)
    else: wl(ws4, r, 3, adoption)
    wl(ws4, r, 4, basis); wi(ws4, r, 5, annual, 'N$#,##0.00')

wl(ws4, 11, 1, "Total Year 1 Revenue per Meter", f=bold_font, fi=green_fill)
cell = wf(ws4, 11, 5, "=SUM(E4:E10)", 'N$#,##0.00', bold_font, green_fill)
cell.border = total_bdr

wl(ws4, 14, 1, "Revenue by Year", f=title_font)
wh(ws4, 15, ["Year"] + SCENARIOS)
wl(ws4, 16, 1, "Year 1")
for i in range(NS): wf(ws4, 16, 2+i, f"={MC[i]}*E11", CURR, green_font)
for yr in range(2, 6):
    wl(ws4, 15+yr, 1, f"Year {yr}")
    for i in range(NS): wf(ws4, 15+yr, 2+i, f"={MC[i]}*(E6+E7+E8+E9+E10)", CURR, green_font)
wl(ws4, 21, 1, "5-Year Total", f=bold_font, fi=green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws4, 21, 2+i, f"=SUM({col}16:{col}20)", CURR, bold_font, green_fill)
    cell.border = total_bdr

wl(ws4, 24, 1, "5-Year Revenue Breakdown by Stream", f=title_font)
wh(ws4, 25, ["Revenue Stream"] + SCENARIOS)
bd = [
    (26, "Hardware Sales (One-Time)", "E4"),
    (27, "Installation Fees (One-Time)", "E5"),
    (28, "Maintenance (Recurring)", "E6", 5),
    (29, "Wi-Fi Subscriptions", "E7", 5),
    (30, "SMS Notifications", "E8", 5),
    (31, "Mobile App Subscriptions", "E9", 5),
    (32, "Real Estate Management", "E10", 5),
]
for item in bd:
    r = item[0]; name = item[1]; ref = item[2]; mult = item[3] if len(item) > 3 else 1
    wl(ws4, r, 1, name)
    for i in range(NS):
        wf(ws4, r, 2+i, f"={MC[i]}*{ref}*{mult}", CURR, green_font)
wl(ws4, 33, 1, "5-Year Total", f=bold_font, fi=green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws4, 33, 2+i, f"=SUM({col}26:{col}32)", CURR, bold_font, green_fill)
    cell.border = total_bdr

# ═══════════════════════════════════════════════════════════════
# SHEET 5: PROFITABILITY ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Profitability Analysis"); ws5.sheet_properties.tabColor = "C00000"
ws5.column_dimensions['A'].width = 42
for i in range(NS): ws5.column_dimensions[get_column_letter(2+i)].width = 20

wl(ws5, 1, 1, "Updated Revenue & Profitability Analysis (Year 1)", f=title_font)
wh(ws5, 3, ["Metric"] + SCENARIOS)

wl(ws5, 4, 1, "REVENUE (N$)", f=bold11, fi=green_fill); fill_row(ws5, 4, 2, 6, green_fill)
wl(ws5, 5, 1, "Hardware Sales")
for i in range(NS): wf(ws5, 5, 2+i, f"={MC[i]}*Revenue!E4", CURR, green_font)
wl(ws5, 6, 1, "Installation Fees")
for i in range(NS): wf(ws5, 6, 2+i, f"={MC[i]}*Revenue!E5", CURR, green_font)
wl(ws5, 7, 1, "Service Revenues")
for i in range(NS): wf(ws5, 7, 2+i, f"={MC[i]}*(Revenue!E6+Revenue!E7+Revenue!E8+Revenue!E9+Revenue!E10)", CURR, green_font)
wl(ws5, 8, 1, "Total Revenue", f=bold11, fi=green_fill); fill_row(ws5, 8, 2, 6, green_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 8, 2+i, f"=SUM({col}5:{col}7)", CURR, bold11, green_fill)
    cell.border = total_bdr

wl(ws5, 9, 1, "MANUFACTURING COST (N$)", f=bold11, fi=orange_fill); fill_row(ws5, 9, 2, 6, orange_fill)
wl(ws5, 10, 1, "Cost per Meter")
for i in range(NS): wf(ws5, 10, 2+i, f"='Manufacturing Cost'!C{MFG_S+i}", CURR, green_font)
wl(ws5, 11, 1, "Total Manufacturing Cost")
for i in range(NS): wf(ws5, 11, 2+i, f"='Manufacturing Cost'!D{MFG_S+i}", CURR, green_font)

wl(ws5, 12, 1, "OPERATING COSTS (N$)", f=bold11, fi=orange_fill)
for i in range(NS):
    pc = get_column_letter(3+i); ic = get_column_letter(3+i)
    cell = wf(ws5, 12, 2+i, f"='Personnel Cost'!{pc}{TOTAL_COST_ROW}+'Infrastructure & Equipment'!{ic}{INFRA_TOT}", CURR, green_font, orange_fill)

wl(ws5, 13, 1, "EQUIPMENT INVESTMENT (N$)")
for i in range(NS):
    ec = get_column_letter(3+i)
    wf(ws5, 13, 2+i, f"='Infrastructure & Equipment'!{ec}{EQUIP_TOT}", CURR, green_font)

wl(ws5, 14, 1, "PROFITABILITY (N$)", f=bold11, fi=light_fill); fill_row(ws5, 14, 2, 6, light_fill)
wl(ws5, 15, 1, "Gross Profit", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 15, 2+i, f"={col}8-{col}11", CURR, bold_font)
wl(ws5, 16, 1, "Operating Profit (EBITDA)", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 16, 2+i, f"={col}15-{col}12-{col}13", CURR, bold_font)
wl(ws5, 17, 1, "Net Profit", f=navy_bold)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 17, 2+i, f"={col}16", CURR, navy_bold)
    cell.border = total_bdr

wl(ws5, 18, 1, "MARGINS", f=bold11, fi=light_fill); fill_row(ws5, 18, 2, 6, light_fill)
for rr, label, num_row in [(19, "Gross Margin %", 15), (20, "Operating Margin %", 16), (21, "Net Margin %", 17)]:
    wl(ws5, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws5, rr, 2+i, f"=IF({col}8<>0,{col}{num_row}/{col}8,0)", PCT, bold_font)

wl(ws5, 23, 1, "5-YEAR PROJECTIONS", f=title_font)
wh(ws5, 24, ["Metric"] + SCENARIOS)
wl(ws5, 25, 1, "5-Year Total Revenue", f=bold_font)
for i in range(NS):
    rc = get_column_letter(2+i)
    wf(ws5, 25, 2+i, f"=Revenue!{rc}21", CURR, green_font)
wl(ws5, 26, 1, "5-Year Total Costs (Mfg + OpCost×5 + Equip)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 26, 2+i, f"={col}11+{col}12*5+{col}13", CURR)
wl(ws5, 27, 1, "5-Year Net Profit", f=navy_bold)
for i in range(NS):
    col = get_column_letter(2+i)
    cell = wf(ws5, 27, 2+i, f"={col}25-{col}26", CURR, navy_bold)
    cell.border = total_bdr
wl(ws5, 28, 1, "5-Year ROI %", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 28, 2+i, f"=IF({col}26<>0,{col}27/{col}26,0)", PCT, bold_font)
wl(ws5, 29, 1, "5-Year Net Margin %", f=bold_font)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws5, 29, 2+i, f"=IF({col}25<>0,{col}27/{col}25,0)", PCT, bold_font)

# ═══════════════════════════════════════════════════════════════
# SHEET 6: INTERACTIVE DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Dashboard"); ws6.sheet_properties.tabColor = "002060"

# ── Column widths ──
ws6.column_dimensions['A'].width = 1.5    # Margin
ws6.column_dimensions['B'].width = 2.5    # Section strip
ws6.column_dimensions['C'].width = 36     # Labels
ws6.column_dimensions['D'].width = 15     # Year 1
ws6.column_dimensions['E'].width = 15     # Year 2
ws6.column_dimensions['F'].width = 15     # Year 3
ws6.column_dimensions['G'].width = 15     # Year 4
ws6.column_dimensions['H'].width = 15     # Year 5
ws6.column_dimensions['I'].width = 16     # 5-Year Total
ws6.column_dimensions['J'].width = 2      # Gap

# ── Dashboard styles ──
TEAL_H  = PatternFill("solid", fgColor="00897B")
TEAL_BG = PatternFill("solid", fgColor="E0F2F1")
TEAL_TOT= PatternFill("solid", fgColor="B2DFDB")
ORNG_H  = PatternFill("solid", fgColor="E65100")
ORNG_BG = PatternFill("solid", fgColor="FFF3E0")
ORNG_TOT= PatternFill("solid", fgColor="FFE0B2")
GRN_H   = PatternFill("solid", fgColor="2E7D32")
GRN_BG  = PatternFill("solid", fgColor="E8F5E9")
GRN_TOT = PatternFill("solid", fgColor="C8E6C9")
BLUE_H  = PatternFill("solid", fgColor="1565C0")
BLUE_BG = PatternFill("solid", fgColor="E3F2FD")
DARK_H  = PatternFill("solid", fgColor="263238")
DARK_BG = PatternFill("solid", fgColor="ECEFF1")
DARK_TOT= PatternFill("solid", fgColor="CFD8DC")
GOLD_H  = PatternFill("solid", fgColor="F57F17")
GOLD_BG = PatternFill("solid", fgColor="FFFDE7")
DROP_BG = PatternFill("solid", fgColor="FFF9C4")  # Yellow highlight for dropdown

d_title  = Font(bold=True, size=16, name="Calibri", color="1B2A4A")
d_sub    = Font(italic=True, size=11, name="Calibri", color="546E7A")
d_sec    = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
d_label  = Font(size=10, name="Calibri", color="37474F")
d_val    = Font(size=10, name="Calibri", color="1B5E20")
d_bold   = Font(bold=True, size=10, name="Calibri", color="37474F")
d_total  = Font(bold=True, size=11, name="Calibri", color="1B2A4A")
d_hdr    = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
d_pct    = Font(bold=True, size=10, name="Calibri", color="0D47A1")
d_input  = Font(bold=True, size=11, name="Calibri", color="0000FF")
d_drop   = Font(bold=True, size=12, name="Calibri", color="BF360C")
d_param  = Font(size=10, name="Calibri", color="37474F")

d_thin  = Border(
    left=Side("thin", color="CFD8DC"), right=Side("thin", color="CFD8DC"),
    top=Side("thin", color="CFD8DC"), bottom=Side("thin", color="CFD8DC"))
d_total_b = Border(top=Side("medium", color="263238"), bottom=Side("double", color="263238"))
d_drop_b = Border(
    left=Side("medium", color="BF360C"), right=Side("medium", color="BF360C"),
    top=Side("medium", color="BF360C"), bottom=Side("medium", color="BF360C"))

# ── Helper functions ──
def dsec(ws, r, label, hdr_fill, cs=2, ce=9):
    for c in range(cs, ce+1):
        ws.cell(r, c).fill = hdr_fill; ws.cell(r, c).border = d_thin
    ws.cell(r, 3, label).font = d_sec

def drow(ws, r, label, strip_fill, bg_fill, font=d_label, cs=2, ce=9):
    ws.cell(r, cs).fill = strip_fill
    ws.cell(r, 3, label).font = font; ws.cell(r, 3).fill = bg_fill; ws.cell(r, 3).border = d_thin
    for c in range(4, ce+1):
        ws.cell(r, c).fill = bg_fill; ws.cell(r, c).border = d_thin

def dtotal(ws, r, label, strip_fill, bg_fill, cs=2, ce=9):
    ws.cell(r, cs).fill = strip_fill
    ws.cell(r, 3, label).font = d_total; ws.cell(r, 3).fill = bg_fill; ws.cell(r, 3).border = d_total_b
    for c in range(4, ce+1):
        ws.cell(r, c).fill = bg_fill; ws.cell(r, c).border = d_total_b

def dval(ws, r, c, formula, fmt=CURR, font=d_val):
    cell = ws.cell(r, c); cell.value = formula; cell.font = font; cell.number_format = fmt

# INDEX helper — returns formula to look up a value from reference row
REF_IDX = "$C$54"  # Scenario index cell
def idx(ref_row):
    return f"INDEX($C${ref_row}:$G${ref_row},1,{REF_IDX})"

# ═══════════ TITLE (rows 1-2) ═══════════
ws6.merge_cells('C1:I1')
ws6['C1'] = "GRIDx Financial Model — Interactive Dashboard"
ws6['C1'].font = d_title; ws6['C1'].alignment = Alignment(vertical="center")
ws6.row_dimensions[1].height = 30

ws6.merge_cells('C2:I2')
ws6['C2'] = "Pulsar Electronic Solutions — Smart Meter Deployment (Namibia)"
ws6['C2'].font = d_sub
ws6.row_dimensions[2].height = 20

# ═══════════ ROLLOUT SELECTION (rows 4-6) ═══════════
dsec(ws6, 4, "SELECT ROLLOUT SCENARIO", BLUE_H)

ws6.cell(5, 3, "Deployment Scale:").font = Font(bold=True, size=11, name="Calibri", color="37474F")
ws6.cell(5, 3).border = d_thin

# Dropdown cell D5 (merged D5:E5 for visibility)
ws6.merge_cells('D5:E5')
ws6['D5'] = "10,000 Meters"  # Default selection
ws6['D5'].font = d_drop
ws6['D5'].fill = DROP_BG
ws6['D5'].border = d_drop_b
ws6['D5'].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[5].height = 28

# Display selected meter count
ws6.cell(5, 6, "Selected:").font = d_param; ws6.cell(5, 6).border = d_thin
c = ws6.cell(5, 7)
c.value = f"={idx(55)}"
c.font = d_input; c.number_format = '#,##0 "meters"'; c.border = d_thin

# ═══════════ KEY PARAMETERS (rows 7-14) ═══════════
dsec(ws6, 7, "KEY PARAMETERS FOR SELECTED SCENARIO", GOLD_H)

params = [
    (8,  "Sale Price per Meter",       5500,                    CURR,  True),
    (9,  "Revenue per Meter (Year 1)", "=Revenue!$E$11",        'N$#,##0.00', False),
    (10, "Recurring Revenue/Meter/Year","=SUM(Revenue!$E$6:$E$10)",'N$#,##0.00', False),
    (11, "Manufacturing Cost per Meter",f"={idx(56)}",          CURR,  False),
    (12, "Annual Personnel Cost",      f"={idx(58)}",           CURR,  False),
    (13, "Annual Infrastructure Cost",  f"={idx(59)}",          CURR,  False),
    (14, "Equipment Investment (1-time)",f"={idx(60)}",         CURR,  False),
]
for pr, label, val, fmt, static in params:
    ws6.cell(pr, 2).fill = GOLD_H
    ws6.cell(pr, 3, label).font = d_param; ws6.cell(pr, 3).fill = GOLD_BG; ws6.cell(pr, 3).border = d_thin
    ws6.merge_cells(f'D{pr}:E{pr}')
    c = ws6.cell(pr, 4)
    c.value = val; c.font = d_input; c.fill = GOLD_BG; c.number_format = fmt; c.border = d_thin
    if static:
        c.font = Font(bold=True, size=10, name="Calibri", color="37474F")
    # Label "dynamic" for INDEX-based cells
    if not static and isinstance(val, str) and "INDEX" in val:
        ws6.cell(pr, 6, "(updates with selection)").font = Font(size=8, italic=True, color="90A4AE", name="Calibri")

# ═══════════ 5-YEAR FINANCIAL PROJECTION (rows 16-48) ═══════════
dsec(ws6, 16, "5-YEAR FINANCIAL PROJECTION", DARK_H)

# Column headers row 17
hdr_bg = PatternFill("solid", fgColor="37474F")
ws6.cell(17, 2).fill = hdr_bg
ws6.cell(17, 3, "").fill = hdr_bg; ws6.cell(17, 3).border = d_thin
year_labels = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "5-Year Total"]
for i, yl in enumerate(year_labels):
    c = ws6.cell(17, 4+i)
    c.value = yl; c.font = d_hdr; c.fill = hdr_bg
    c.alignment = Alignment(horizontal="center"); c.border = d_thin
ws6.row_dimensions[17].height = 22

# ── Revenue subsection (rows 18-26) ──
dsec(ws6, 18, "REVENUE (N$)", TEAL_H)

# Revenue streams: rows 19-25 (7 streams)
rev_streams = [
    (19, "Hardware Sales (one-time)",     "Revenue!$E$4",  True),   # One-time
    (20, "Installation Fees (one-time)",  "Revenue!$E$5",  True),   # One-time
    (21, "Meter Maintenance",             "Revenue!$E$6",  False),  # Recurring
    (22, "Wi-Fi Subscriptions",           "Revenue!$E$7",  False),
    (23, "SMS Notifications",             "Revenue!$E$8",  False),
    (24, "Mobile App Subscriptions",      "Revenue!$E$9",  False),
    (25, "Real Estate Management",        "Revenue!$E$10", False),
]

for rr, label, rate_ref, one_time in rev_streams:
    drow(ws6, rr, label, TEAL_H, TEAL_BG)
    # Year 1 (col D): meters × per-meter rate
    dval(ws6, rr, 4, f"={idx(55)}*{rate_ref}")
    # Years 2-5 (cols E-H)
    for yc in range(5, 9):  # E through H
        if one_time:
            dval(ws6, rr, yc, 0)
        else:
            dval(ws6, rr, yc, f"=$D${rr}")
    # Total (col I)
    dval(ws6, rr, 9, f"=SUM(D{rr}:H{rr})")

# Total Revenue row 26
dtotal(ws6, 26, "Total Revenue", TEAL_H, TEAL_TOT)
for c in range(4, 10):  # D through I
    col_l = get_column_letter(c)
    dval(ws6, 26, c, f"=SUM({col_l}19:{col_l}25)", CURR, d_total)

# ── Costs subsection (rows 28-33) ──
dsec(ws6, 28, "COSTS (N$)", ORNG_H)

# Manufacturing (row 29) — one-time Year 1
drow(ws6, 29, "Manufacturing Cost", ORNG_H, ORNG_BG)
dval(ws6, 29, 4, f"={idx(57)}")  # Year 1 = total mfg from lookup
for yc in range(5, 9):
    dval(ws6, 29, yc, 0)  # Years 2-5: no manufacturing
dval(ws6, 29, 9, f"=SUM(D29:H29)")

# Personnel (row 30) — recurring annual
drow(ws6, 30, "Personnel Cost (Annual)", ORNG_H, ORNG_BG)
dval(ws6, 30, 4, f"={idx(58)}")  # Year 1
for yc in range(5, 9):
    dval(ws6, 30, yc, f"=$D$30")  # Years 2-5: same as Year 1
dval(ws6, 30, 9, f"=SUM(D30:H30)")

# Infrastructure (row 31) — recurring annual
drow(ws6, 31, "Infrastructure Cost (Annual)", ORNG_H, ORNG_BG)
dval(ws6, 31, 4, f"={idx(59)}")
for yc in range(5, 9):
    dval(ws6, 31, yc, f"=$D$31")
dval(ws6, 31, 9, f"=SUM(D31:H31)")

# Equipment (row 32) — one-time Year 1
drow(ws6, 32, "Equipment Investment (one-time)", ORNG_H, ORNG_BG)
dval(ws6, 32, 4, f"={idx(60)}")
for yc in range(5, 9):
    dval(ws6, 32, yc, 0)
dval(ws6, 32, 9, f"=SUM(D32:H32)")

# Total Costs row 33
dtotal(ws6, 33, "Total Costs", ORNG_H, ORNG_TOT)
for c in range(4, 10):
    col_l = get_column_letter(c)
    dval(ws6, 33, c, f"=SUM({col_l}29:{col_l}32)", CURR, d_total)

# ── Profitability subsection (rows 35-39) ──
dsec(ws6, 35, "PROFITABILITY (N$)", GRN_H)

# Gross Profit (row 36)
drow(ws6, 36, "Gross Profit (Revenue − Manufacturing)", GRN_H, GRN_BG, d_bold)
for c in range(4, 10):
    col_l = get_column_letter(c)
    dval(ws6, 36, c, f"={col_l}26-{col_l}29")

# Net Profit (row 37)
dtotal(ws6, 37, "Net Profit", GRN_H, GRN_TOT)
for c in range(4, 10):
    col_l = get_column_letter(c)
    dval(ws6, 37, c, f"={col_l}26-{col_l}33", CURR, d_total)

# Net Margin % (row 38)
drow(ws6, 38, "Net Margin %", GRN_H, GRN_BG, d_pct)
for c in range(4, 10):
    col_l = get_column_letter(c)
    dval(ws6, 38, c, f"=IF({col_l}26>0,{col_l}37/{col_l}26,0)", PCT, d_pct)

# Gross Margin % (row 39)
drow(ws6, 39, "Gross Margin %", GRN_H, GRN_BG, d_pct)
for c in range(4, 10):
    col_l = get_column_letter(c)
    dval(ws6, 39, c, f"=IF({col_l}26>0,{col_l}36/{col_l}26,0)", PCT, d_pct)

# ── Cumulative (rows 41-43) ──
dsec(ws6, 41, "CUMULATIVE PERFORMANCE", DARK_H)

# Cumulative Revenue (row 42)
drow(ws6, 42, "Cumulative Revenue", DARK_H, DARK_BG, d_bold)
dval(ws6, 42, 4, "=D26")  # Year 1
for yc in range(5, 9):
    prev = get_column_letter(yc-1)
    cur = get_column_letter(yc)
    dval(ws6, 42, yc, f"={prev}42+{cur}26")
dval(ws6, 42, 9, "=H42")  # Total = last cumulative

# Cumulative Net Profit (row 43)
drow(ws6, 43, "Cumulative Net Profit", DARK_H, DARK_BG, d_bold)
dval(ws6, 43, 4, "=D37")
for yc in range(5, 9):
    prev = get_column_letter(yc-1)
    cur = get_column_letter(yc)
    dval(ws6, 43, yc, f"={prev}43+{cur}37")
dval(ws6, 43, 9, "=H43")

# ── 5-Year Summary (rows 45-49) ──
dsec(ws6, 45, "5-YEAR KEY METRICS", BLUE_H)

summary_items = [
    (46, "Total 5-Year Revenue",   "=I26",             CURR,  d_bold),
    (47, "Total 5-Year Costs",     "=I33",             CURR,  d_bold),
    (48, "Total 5-Year Net Profit","=I37",             CURR,  d_total),
    (49, "5-Year ROI %",           "=IF(I33>0,I37/I33,0)", PCT, d_pct),
    (50, "5-Year Net Margin %",    "=IF(I26>0,I37/I26,0)", PCT, d_pct),
]
for sr, label, formula, fmt, fnt in summary_items:
    drow(ws6, sr, label, BLUE_H, BLUE_BG, fnt)
    ws6.merge_cells(f'D{sr}:E{sr}')
    c = ws6.cell(sr, 4)
    c.value = formula; c.font = fnt; c.number_format = fmt; c.fill = BLUE_BG; c.border = d_thin
    if sr == 48:
        c.border = d_total_b

# ═══════════ REFERENCE DATA (rows 53-60) ═══════════
# This data supports the INDEX/MATCH lookups from the dropdown
ref_label_font = Font(size=8, italic=True, color="90A4AE", name="Calibri")
ws6.cell(52, 3, "REFERENCE DATA — DO NOT MODIFY (supports dropdown calculations)").font = ref_label_font

# Row 53: Scenario labels (must match dropdown values exactly)
for i, s in enumerate(SCENARIOS):
    ws6.cell(53, 3+i, s).font = ref_label_font

# Row 54: Scenario index (MATCH formula)
ws6.cell(54, 2, "Index:").font = ref_label_font
ws6.cell(54, 3).value = "=MATCH($D$5,$C$53:$G$53,0)"
ws6.cell(54, 3).font = ref_label_font

# Row 55: Meter counts
ws6.cell(55, 2, "Meters:").font = ref_label_font
for i, m in enumerate(MC):
    ws6.cell(55, 3+i, m).font = ref_label_font

# Row 56: Manufacturing cost per meter
ws6.cell(56, 2, "Mfg/meter:").font = ref_label_font
for i in range(NS):
    ws6.cell(56, 3+i).value = f"='Manufacturing Cost'!C{MFG_S+i}"
    ws6.cell(56, 3+i).font = ref_label_font

# Row 57: Total manufacturing cost
ws6.cell(57, 2, "Mfg total:").font = ref_label_font
for i in range(NS):
    ws6.cell(57, 3+i).value = f"='Manufacturing Cost'!D{MFG_S+i}"
    ws6.cell(57, 3+i).font = ref_label_font

# Row 58: Annual personnel cost
ws6.cell(58, 2, "Personnel:").font = ref_label_font
for i in range(NS):
    pc = get_column_letter(3+i)
    ws6.cell(58, 3+i).value = f"='Personnel Cost'!{pc}{TOTAL_COST_ROW}"
    ws6.cell(58, 3+i).font = ref_label_font

# Row 59: Annual infrastructure cost
ws6.cell(59, 2, "Infra:").font = ref_label_font
for i in range(NS):
    ic = get_column_letter(3+i)
    ws6.cell(59, 3+i).value = f"='Infrastructure & Equipment'!{ic}{INFRA_TOT}"
    ws6.cell(59, 3+i).font = ref_label_font

# Row 60: Equipment investment (one-time)
ws6.cell(60, 2, "Equipment:").font = ref_label_font
for i in range(NS):
    ec = get_column_letter(3+i)
    ws6.cell(60, 3+i).value = f"='Infrastructure & Equipment'!{ec}{EQUIP_TOT}"
    ws6.cell(60, 3+i).font = ref_label_font

# ═══════════ DATA VALIDATION (Dropdown) ═══════════
dv = DataValidation(
    type="list",
    formula1="=$C$53:$G$53",
    allow_blank=False
)
dv.prompt = "Select a rollout scenario"
dv.promptTitle = "Meter Rollout"
dv.error = "Please select a valid scenario from the dropdown"
dv.errorTitle = "Invalid Selection"
ws6.add_data_validation(dv)
dv.add(ws6['D5'])

# ═══════════ CHARTS (positioned RIGHT of table, col K+) ═══════════
cats = Reference(ws6, min_col=4, min_row=17, max_col=8, max_row=17)  # "Year 1" through "Year 5"

# ── CHART 1: 5-Year Revenue & Net Profit (Line Chart) ──
ch1 = LineChart()
ch1.title = "5-Year Revenue & Net Profit Projection"
ch1.style = 10
ch1.y_axis.title = "Amount (N$)"
ch1.y_axis.numFmt = '#,##0'
ch1.x_axis.title = "Year"
ch1.width = 26; ch1.height = 15

# Total Revenue line
d1 = Reference(ws6, min_col=4, min_row=26, max_col=8, max_row=26)
ch1.add_data(d1, from_rows=True)
ch1.series[0].title = SeriesLabel(v="Total Revenue")
ch1.series[0].graphicalProperties.line.solidFill = "00897B"
ch1.series[0].graphicalProperties.line.width = 32000

# Net Profit line
d2 = Reference(ws6, min_col=4, min_row=37, max_col=8, max_row=37)
ch1.add_data(d2, from_rows=True)
ch1.series[1].title = SeriesLabel(v="Net Profit")
ch1.series[1].graphicalProperties.line.solidFill = "2E7D32"
ch1.series[1].graphicalProperties.line.width = 28000

# Total Costs line
d3 = Reference(ws6, min_col=4, min_row=33, max_col=8, max_row=33)
ch1.add_data(d3, from_rows=True)
ch1.series[2].title = SeriesLabel(v="Total Costs")
ch1.series[2].graphicalProperties.line.solidFill = "E65100"
ch1.series[2].graphicalProperties.line.width = 28000
ch1.series[2].graphicalProperties.line.dashStyle = "dash"

ch1.set_categories(cats)
ch1.legend.position = 'b'
ws6.add_chart(ch1, "K3")

# ── CHART 2: Revenue Breakdown by Year (Stacked Bar) ──
ch2 = BarChart()
ch2.type = "col"; ch2.grouping = "stacked"
ch2.title = "Revenue Breakdown by Year"
ch2.style = 10
ch2.y_axis.title = "Amount (N$)"
ch2.y_axis.numFmt = '#,##0'
ch2.width = 26; ch2.height = 15

rev_chart_series = [
    (19, "Hardware Sales",  "004D40"),
    (20, "Installation",    "00695C"),
    (21, "Maintenance",     "00897B"),
    (22, "Wi-Fi",           "26A69A"),
    (23, "SMS",             "4DB6AC"),
    (24, "Mobile App",      "80CBC4"),
    (25, "Real Estate",     "B2DFDB"),
]
for rr, nm, color in rev_chart_series:
    d = Reference(ws6, min_col=4, min_row=rr, max_col=8, max_row=rr)
    ch2.add_data(d, from_rows=True)
    ch2.series[-1].title = SeriesLabel(v=nm)
    ch2.series[-1].graphicalProperties.solidFill = color

ch2.set_categories(cats)
ch2.legend.position = 'b'
ws6.add_chart(ch2, "K20")

# ── CHART 3: Cumulative Revenue & Profit (Area Chart) ──
ch3 = AreaChart()
ch3.title = "Cumulative Revenue & Profit Growth"
ch3.style = 10
ch3.y_axis.title = "Cumulative Amount (N$)"
ch3.y_axis.numFmt = '#,##0'
ch3.width = 26; ch3.height = 15

d_cr = Reference(ws6, min_col=4, min_row=42, max_col=8, max_row=42)
ch3.add_data(d_cr, from_rows=True)
ch3.series[0].title = SeriesLabel(v="Cumulative Revenue")
ch3.series[0].graphicalProperties.solidFill = "B2DFDB"
ch3.series[0].graphicalProperties.line.solidFill = "00897B"

d_cp = Reference(ws6, min_col=4, min_row=43, max_col=8, max_row=43)
ch3.add_data(d_cp, from_rows=True)
ch3.series[1].title = SeriesLabel(v="Cumulative Net Profit")
ch3.series[1].graphicalProperties.solidFill = "C8E6C9"
ch3.series[1].graphicalProperties.line.solidFill = "2E7D32"

ch3.set_categories(cats)
ch3.legend.position = 'b'
ws6.add_chart(ch3, "K37")

# ═══════════ FINALIZE ═══════════
wb.move_sheet("Dashboard", offset=-5)
ws1.freeze_panes = "B2"
ws5.freeze_panes = "B4"
ws6.freeze_panes = "C5"

ws6.page_setup.orientation = "landscape"
ws6.page_setup.fitToWidth = 1
ws6.page_setup.fitToHeight = 1

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v4.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print("Interactive Dashboard: dropdown selector + dynamic 5-year projections + 3 charts")
print(f"Total reference data rows: 53-60, scenario index at C54")
