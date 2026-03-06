"""
GRIDx Financial Dashboard v10 — Comprehensive Dashboard with Charts from All Tabs
Tabs: Dashboard, Revenue, Manufacturing, Personnel, Operating Costs,
      Implementation, Equipment, Development, Financing, Profitability
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference, BarChart3D
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Constants ──
SCENARIOS = ["3,000 Meters","5,000 Meters","10,000 Meters","20,000 Meters","50,000 Meters"]
SC_SHORT = ["3K","5K","10K","20K","50K"]
MC = [3000, 5000, 10000, 20000, 50000]; NS = 5
HW=5500; INST=700; MAINT=300; WIFI=90; SMS=30; APP=624; RE=76.80
ANN_REC = HW+MAINT+WIFI+SMS+APP+RE  # 6620.80
Y1_PM = ANN_REC + INST  # 7320.80
MFG_U = [2000, 2000, 1700, 1700, 1462]
SP_R = 4495 * 0.02
VH=[2,2,3,5,10]; TK=[2,3,4,6,12]; WK=[3,5,7,12,25]; NT=[1,1,2,3,5]
CLIENT_PB = [7.6, 7.2, 4.2, 3.9, 3.5]
CLIENT_DCF = [65500000, 119300000, 259647182, 538200000, 1390000000]

POSITIONS = [
    ("CEO", 35000, [1,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("COO", 28000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("CFO", 28000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("CTO", 30000, [0,1,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of Business Dev", 25000, [0,0,1,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Head of HR", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Legal & Compliance", 22000, [0,0,0,1,1], "LEADERSHIP & MANAGEMENT"),
    ("Operations Manager", 20000, [0,1,1,1,1], "OPERATIONS & FIELD"),
    ("Field Support Techs", 9000, [2,3,4,5,8], "OPERATIONS & FIELD"),
    ("Installation Team Lead", 11000, [0,1,1,2,3], "OPERATIONS & FIELD"),
    ("Logistics Coordinator", 8000, [0,0,1,1,2], "OPERATIONS & FIELD"),
    ("Engineering Manager", 22000, [0,0,1,1,1], "ENGINEERING & QA"),
    ("QA Engineers", 12000, [1,1,1,3,5], "ENGINEERING & QA"),
    ("Firmware/Software Eng", 15000, [0,1,2,3,5], "ENGINEERING & QA"),
    ("Hardware Engineers", 15000, [0,0,1,2,3], "ENGINEERING & QA"),
    ("Customer Support Mgr", 18000, [0,0,1,1,1], "CUSTOMER SUCCESS"),
    ("Customer Support Agents", 7200, [1,2,2,4,8], "CUSTOMER SUCCESS"),
    ("Technical Support", 9000, [0,0,1,2,3], "CUSTOMER SUCCESS"),
    ("Sales Manager", 20000, [0,0,1,1,1], "SALES & MARKETING"),
    ("Sales Representatives", 10000, [0,1,2,3,5], "SALES & MARKETING"),
    ("Marketing Coordinator", 9000, [0,0,1,1,2], "SALES & MARKETING"),
    ("Finance Manager", 20000, [0,0,1,1,1], "FINANCE & ADMIN"),
    ("Accountant", 11000, [0,1,1,2,3], "FINANCE & ADMIN"),
    ("Admin Assistant", 7000, [0,0,1,2,3], "FINANCE & ADMIN"),
]

# ── Styles ──
NV = "1B2A4A"
hdr_fill = PatternFill("solid", fgColor=NV)
grn_fill = PatternFill("solid", fgColor="E2EFDA")
org_fill = PatternFill("solid", fgColor="FCE4D6")
ylw_fill = PatternFill("solid", fgColor="FFF2CC")
lt_fill  = PatternFill("solid", fgColor="D6E4F0")
hdr_f = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
bf    = Font(bold=True, size=10, name="Calibri")
b11   = Font(bold=True, size=11, name="Calibri")
lbl   = Font(size=10, name="Calibri")
inp   = Font(color="0000FF", size=10, name="Calibri")
ttl   = Font(bold=True, color=NV, size=14, name="Calibri")
nb12  = Font(bold=True, color=NV, size=12, name="Calibri")
dept_f = Font(bold=True, color="4472C4", size=10, name="Calibri")
note_f = Font(italic=True, size=10, color="666666", name="Calibri")
warn_f = Font(italic=True, size=10, color="C00000", name="Calibri")
tb = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
            top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))
tot_b = Border(top=Side("medium", color=NV), bottom=Side("double", color=NV))
NUM = '#,##0'; CURR = 'N$#,##0'; PCT = '0.0%'

def wh(ws, r, labels, cs=1):
    for i, l in enumerate(labels):
        c = ws.cell(r, cs+i, l); c.font = hdr_f; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = tb

def wl(ws, r, c, t, f=lbl, fi=None):
    cell = ws.cell(r, c, t); cell.font = f; cell.border = tb
    if fi: cell.fill = fi
    return cell

def wi(ws, r, c, v, fmt=NUM):
    cell = ws.cell(r, c, v); cell.font = inp; cell.fill = ylw_fill
    cell.number_format = fmt; cell.border = tb; return cell

def wf(ws, r, c, formula, fmt=NUM, f=lbl, fi=None):
    cell = ws.cell(r, c, formula); cell.font = f; cell.number_format = fmt; cell.border = tb
    if fi: cell.fill = fi
    return cell

def fr(ws, r, cs, ce, fi):
    for c in range(cs, ce+1): ws.cell(r, c).fill = fi; ws.cell(r, c).border = tb

def scols(ws, lw=42):
    ws.column_dimensions['A'].width = lw
    for i in range(NS): ws.column_dimensions[get_column_letter(2+i)].width = 18

# ═══════════════════════════════════════════════════════════════
# SHEET 1: REVENUE
# ═══════════════════════════════════════════════════════════════
ws_r = wb.active; ws_r.title = "Revenue"; ws_r.sheet_properties.tabColor = "ED7D31"
ws_r.column_dimensions['A'].width = 40
for c in ['B','C','D','E']: ws_r.column_dimensions[c].width = 22

wl(ws_r, 1, 1, "Revenue Model — Assumptions & Derivation", f=ttl)
wh(ws_r, 3, ["Revenue Stream", "Rate", "Adoption", "Basis", "Annual/Meter (N$)"])
for r, nm, rate, ad, bas, val in [
    (4, "Meter Hardware Sale", "N$5,500/yr", "100%", "Annual recurring", HW),
    (5, "Installation Fee", "N$700 one-time", "100%", "One-time", INST),
    (6, "Maintenance", "N$300/yr", "100%", "Annual", MAINT),
    (7, "Wi-Fi Subscription", "N$25/mo", "30%", "30% x 25 x 12", WIFI),
    (8, "SMS Notifications", "N$2.50/SMS", "50%", "50% x 2 x 2.50 x 12", SMS),
    (9, "Mobile App", "N$65/mo", "80%", "80% x 65 x 12", APP),
    (10, "Real Estate Mgmt", "N$3,200/blk/mo", "1/500m", "(m/500) x 3200 x 12 / m", RE),
]:
    wl(ws_r, r, 1, nm); wl(ws_r, r, 2, rate); wl(ws_r, r, 3, ad)
    wl(ws_r, r, 4, bas); wi(ws_r, r, 5, val, 'N$#,##0.00')

wl(ws_r, 11, 1, "Annual Recurring per Meter", f=bf, fi=grn_fill); fr(ws_r, 11, 2, 5, grn_fill)
wf(ws_r, 11, 5, "=E4+E6+E7+E8+E9+E10", 'N$#,##0.00', bf, grn_fill).border = tot_b
wl(ws_r, 12, 1, "Year 1 Total per Meter (incl Install)", f=bf, fi=grn_fill); fr(ws_r, 12, 2, 5, grn_fill)
wf(ws_r, 12, 5, "=E11+E5", 'N$#,##0.00', bf, grn_fill).border = tot_b

wl(ws_r, 14, 1, "Revenue by Year & Scenario", f=ttl)
wh(ws_r, 16, ["Year"] + SCENARIOS)
wl(ws_r, 17, 1, "Year 1")
for i in range(NS): wi(ws_r, 17, 2+i, round(MC[i] * Y1_PM), CURR)
for yr in range(2, 6):
    wl(ws_r, 15+yr, 1, f"Year {yr}")
    for i in range(NS): wi(ws_r, 15+yr, 2+i, round(MC[i] * ANN_REC), CURR)
wl(ws_r, 22, 1, "5-Year Total", f=b11, fi=grn_fill); fr(ws_r, 22, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_r, 22, 2+i, f"=SUM({col}17:{col}21)", CURR, b11, grn_fill).border = tot_b

wl(ws_r, 24, 1, "5-Year Revenue Breakdown by Stream", f=ttl)
wh(ws_r, 26, ["Revenue Stream"] + SCENARIOS)
for rr, nm, rate, yrs in [
    (27, "Hardware Sales (N$5,500/yr x 5)", HW, 5),
    (28, "Installation Fees (N$700 one-time)", INST, 1),
    (29, "Maintenance (N$300/yr x 5)", MAINT, 5),
    (30, "Wi-Fi (N$90/yr x 5)", WIFI, 5),
    (31, "SMS (N$30/yr x 5)", SMS, 5),
    (32, "Mobile App (N$624/yr x 5)", APP, 5),
    (33, "Real Estate (N$76.80/yr x 5)", RE, 5),
]:
    wl(ws_r, rr, 1, nm)
    for i in range(NS): wi(ws_r, rr, 2+i, round(MC[i] * rate * yrs), CURR)

wl(ws_r, 34, 1, "Total 5-Year Revenue", f=b11, fi=grn_fill); fr(ws_r, 34, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_r, 34, 2+i, f"=SUM({col}27:{col}33)", CURR, b11, grn_fill).border = tot_b
REV_TOT = 34

# Revenue chart — bar chart of 5Y revenue by stream for 10K scenario
wl(ws_r, 36, 1, "VISUAL ANALYSIS", f=ttl)
rev_chart = BarChart(); rev_chart.type = "col"; rev_chart.grouping = "clustered"
rev_chart.title = "5-Year Revenue by Stream (10,000 Meters)"
rev_chart.y_axis.numFmt = '#,##0,,,"B"'; rev_chart.y_axis.title = "N$"
cats = Reference(ws_r, min_col=1, min_row=27, max_row=33)
vals = Reference(ws_r, min_col=4, min_row=27, max_row=33)  # col D = 10K
rev_chart.add_data(vals); rev_chart.set_categories(cats)
rev_chart.series[0].title = SeriesLabel(v="10,000 Meters")
rev_chart.series[0].graphicalProperties.solidFill = "3B82F6"
rev_chart.width = 22; rev_chart.height = 12; rev_chart.legend = None
ws_r.add_chart(rev_chart, "A37")

ws_r.freeze_panes = "B4"
print(f"Revenue: REV_TOT={REV_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: MANUFACTURING
# ═══════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("Manufacturing"); ws_m.sheet_properties.tabColor = "4472C4"; scols(ws_m)
wl(ws_m, 1, 1, "Manufacturing Costs — Derivation", f=ttl)

wl(ws_m, 3, 1, "VOLUME DISCOUNT STRUCTURE", f=b11, fi=lt_fill); fr(ws_m, 3, 2, 5, lt_fill)
wh(ws_m, 5, ["Volume Tier", "Discount", "Cost/Meter (N$)", "Calculation Basis"])
for r, tier, disc, cost, bas in [
    (6, "Base (0-9,999)", "0%", 2000, "N$2,000 per meter"),
    (7, "10,000 - 19,999", "15%", 1700, "15% reduction"),
    (8, "20,000 - 49,999", "30%", 1400, "30% reduction"),
    (9, "50,000+", "~27%", 1462, "Adjusted to N$1,462"),
]:
    wl(ws_m, r, 1, tier); wl(ws_m, r, 2, disc); wi(ws_m, r, 3, cost, CURR); wl(ws_m, r, 4, bas)

wl(ws_m, 11, 1, "MANUFACTURING COST BY SCENARIO", f=b11, fi=org_fill); fr(ws_m, 11, 2, 6, org_fill)
wh(ws_m, 13, ["Cost Item"] + SCENARIOS)
wl(ws_m, 14, 1, "Meter Hardware (Volume x Unit Cost)")
for i in range(NS): wi(ws_m, 14, 2+i, MC[i] * MFG_U[i], CURR)
wl(ws_m, 15, 1, "Spare Parts (2% x N$4,495/meter)")
for i in range(NS): wi(ws_m, 15, 2+i, round(MC[i] * SP_R), CURR)
wl(ws_m, 16, 1, "TOTAL MANUFACTURING", f=b11, fi=grn_fill); fr(ws_m, 16, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_m, 16, 2+i, f"=SUM({col}14:{col}15)", CURR, b11, grn_fill).border = tot_b
MFG_TOT = 16

wl(ws_m, 18, 1, "UNIT COST SUMMARY", f=b11, fi=lt_fill); fr(ws_m, 18, 2, 6, lt_fill)
wh(ws_m, 19, ["Metric"] + SCENARIOS)
wl(ws_m, 20, 1, "Meters Deployed")
for i in range(NS): wi(ws_m, 20, 2+i, MC[i], NUM)
wl(ws_m, 21, 1, "Cost per Meter (N$)")
for i in range(NS): wi(ws_m, 21, 2+i, MFG_U[i], CURR)
wl(ws_m, 22, 1, "Discount Applied")
for i in range(NS): wl(ws_m, 22, 2+i, f"{round((1 - MFG_U[i]/2000) * 100)}%")

# Manufacturing chart
wl(ws_m, 24, 1, "VISUAL ANALYSIS", f=ttl)
mfg_chart = BarChart(); mfg_chart.type = "col"; mfg_chart.grouping = "clustered"
mfg_chart.title = "Manufacturing Cost by Scenario"
mfg_chart.y_axis.numFmt = '#,##0,,"M"'; mfg_chart.y_axis.title = "N$ Millions"
cats = Reference(ws_m, min_col=2, min_row=13, max_col=6, max_row=13)
vals = Reference(ws_m, min_col=2, min_row=16, max_col=6, max_row=16)
mfg_chart.add_data(vals, from_rows=True); mfg_chart.set_categories(cats)
mfg_chart.series[0].title = SeriesLabel(v="Total Cost")
mfg_chart.series[0].graphicalProperties.solidFill = "4472C4"
mfg_chart.width = 20; mfg_chart.height = 12; mfg_chart.legend = None
ws_m.add_chart(mfg_chart, "A25")

ws_m.freeze_panes = "B4"
print(f"Manufacturing: MFG_TOT={MFG_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: PERSONNEL
# ═══════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Personnel"); ws_p.sheet_properties.tabColor = "70AD47"
ws_p.column_dimensions['A'].width = 30
for i in range(NS): ws_p.column_dimensions[get_column_letter(2+i)].width = 14
ws_p.column_dimensions['H'].width = 30; ws_p.column_dimensions['I'].width = 14; ws_p.column_dimensions['J'].width = 14

wl(ws_p, 1, 1, "Personnel — Detailed Staffing Plan", f=ttl)
wl(ws_p, 2, 1, "Headcount by position per deployment scenario", f=note_f)

wh(ws_p, 4, ["Position"] + SCENARIOS)
wh(ws_p, 4, ["Position", "Monthly (N$)", "Annual (N$)"], cs=8)

r = 5; prev_dept = None
for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 6, lt_fill)
        prev_dept = dept; r += 1
    wl(ws_p, r, 1, nm)
    for i in range(NS): wi(ws_p, r, 2+i, hc[i], NUM)
    wl(ws_p, r, 8, nm); wi(ws_p, r, 9, monthly, CURR)
    wf(ws_p, r, 10, f"=I{r}*12", CURR)
    r += 1

wl(ws_p, r, 1, "TOTAL HEADCOUNT", f=b11, fi=grn_fill); fr(ws_p, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_p, r, 2+i, f"=SUMPRODUCT(({col}5:{col}{r-1})*({col}5:{col}{r-1}<>0)*1)", NUM, b11, grn_fill)
HC_TOT_ROW = r

r += 2
wl(ws_p, r, 1, "ANNUAL PERSONNEL COST BY SCENARIO", f=ttl)
r += 1
wh(ws_p, r, ["Position", "Annual Salary"] + SCENARIOS)
cost_hdr = r; r += 1; cost_start = r; prev_dept = None

for nm, monthly, hc, dept in POSITIONS:
    if dept != prev_dept:
        wl(ws_p, r, 1, dept, f=dept_f, fi=lt_fill); fr(ws_p, r, 2, 7, lt_fill)
        prev_dept = dept; r += 1
    annual = monthly * 12
    wl(ws_p, r, 1, nm); wi(ws_p, r, 2, annual, CURR)
    for i in range(NS): wi(ws_p, r, 3+i, hc[i] * annual, CURR)
    r += 1

wl(ws_p, r, 1, "TOTAL ANNUAL PERSONNEL COST", f=b11, fi=grn_fill); fr(ws_p, r, 2, 7, grn_fill)
wl(ws_p, r, 2, "", f=b11, fi=grn_fill)
for i in range(NS):
    col = get_column_letter(3+i)
    wf(ws_p, r, 3+i, f"=SUMPRODUCT(({col}{cost_start}:{col}{r-1})*({col}{cost_start}:{col}{r-1}<>0)*1)", CURR, b11, grn_fill).border = tot_b
PER_TOT = r

# Personnel chart - headcount by scenario
r += 2
wl(ws_p, r, 1, "VISUAL ANALYSIS", f=ttl)
per_chart = BarChart(); per_chart.type = "col"; per_chart.grouping = "clustered"
per_chart.title = "Total Headcount by Deployment Scenario"
per_chart.y_axis.title = "Employees"
cats = Reference(ws_p, min_col=2, min_row=4, max_col=6, max_row=4)
vals = Reference(ws_p, min_col=2, min_row=HC_TOT_ROW, max_col=6, max_row=HC_TOT_ROW)
per_chart.add_data(vals, from_rows=True); per_chart.set_categories(cats)
per_chart.series[0].title = SeriesLabel(v="Headcount")
per_chart.series[0].graphicalProperties.solidFill = "70AD47"
per_chart.width = 20; per_chart.height = 12; per_chart.legend = None
ws_p.add_chart(per_chart, f"A{r+1}")

r += 2
wl(ws_p, r, 1, "Note: The financial model uses a consolidated annual personnel figure of N$748,800/yr.", f=warn_f)
wl(ws_p, r+1, 1, "This sheet shows the recommended full-scale staffing plan for reference.", f=note_f)
ws_p.freeze_panes = "B5"
print(f"Personnel: PER_TOT={PER_TOT}, HC_TOT={HC_TOT_ROW}")

# ═══════════════════════════════════════════════════════════════
# SHEET 4: OPERATING COSTS
# ═══════════════════════════════════════════════════════════════
ws_o = wb.create_sheet("Operating Costs"); ws_o.sheet_properties.tabColor = "ED7D31"; scols(ws_o)
wl(ws_o, 1, 1, "Operating Costs — 5-Year Total", f=ttl)
wl(ws_o, 2, 1, "Annual recurring costs x 5 years", f=note_f)
wh(ws_o, 4, ["Cost Item"] + SCENARIOS)

oper_items = [
    ("Personnel (N$748,800/yr x 5)", [748800*5]*5),
    ("Data Storage (N$48/meter/yr x 5)", [m*48*5 for m in MC]),
    ("Vehicle Maintenance", [v*30000*5 for v in VH]),
    ("Field Equipment Upkeep", [k*10000*5 for k in TK]),
    ("Office & Utilities (N$240,000/yr x 5)", [240000*5]*5),
    ("Software & Tools (N$120,000/yr x 5)", [120000*5]*5),
    ("Marketing & Branding (N$200,000/yr x 5)", [200000*5]*5),
    ("Regulatory Compliance (N$100,000/yr x 5)", [100000*5]*5),
    ("Insurance & Risk (N$150,000/yr x 5)", [150000*5]*5),
]
r = 5
for nm, vals in oper_items:
    wl(ws_o, r, 1, nm)
    for i in range(NS): wi(ws_o, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_o, r, 1, "TOTAL OPERATING (5-Year)", f=b11, fi=grn_fill); fr(ws_o, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_o, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
OPER_TOT = r

# Operating chart
wl(ws_o, r+2, 1, "VISUAL ANALYSIS", f=ttl)
op_chart = BarChart(); op_chart.type = "bar"; op_chart.grouping = "clustered"
op_chart.title = "Operating Cost Breakdown (10,000 Meters, 5-Year)"
op_chart.x_axis.title = "N$"; op_chart.y_axis.numFmt = '#,##0'
cats = Reference(ws_o, min_col=1, min_row=5, max_row=r-1)
vals = Reference(ws_o, min_col=4, min_row=5, max_row=r-1)  # col D = 10K
op_chart.add_data(vals); op_chart.set_categories(cats)
op_chart.series[0].title = SeriesLabel(v="5-Year Cost")
op_chart.series[0].graphicalProperties.solidFill = "ED7D31"
op_chart.width = 22; op_chart.height = 14; op_chart.legend = None
ws_o.add_chart(op_chart, f"A{r+3}")
ws_o.freeze_panes = "B5"
print(f"Operating: OPER_TOT={OPER_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 5: IMPLEMENTATION
# ═══════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("Implementation"); ws_i.sheet_properties.tabColor = "FFC000"; scols(ws_i)
wl(ws_i, 1, 1, "Implementation Costs — One-Time", f=ttl)
wl(ws_i, 2, 1, "Deployment and setup costs (Year 1 only)", f=note_f)
wh(ws_i, 4, ["Cost Item"] + SCENARIOS)

impl_items = [
    ("Project Management", [500000]*5),
    ("Site Surveys & Assessment", [300000]*5),
    ("Onboarding & Training", [250000]*5),
    ("System Integration", [450000]*5),
]
r = 5
for nm, vals in impl_items:
    wl(ws_i, r, 1, nm)
    for i in range(NS): wi(ws_i, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_i, r, 1, "TOTAL IMPLEMENTATION", f=b11, fi=grn_fill); fr(ws_i, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_i, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
IMPL_TOT = r

# Implementation chart
wl(ws_i, r+2, 1, "VISUAL ANALYSIS", f=ttl)
impl_chart = DoughnutChart()
impl_chart.title = "Implementation Cost Split"
cats = Reference(ws_i, min_col=1, min_row=5, max_row=r-1)
vals = Reference(ws_i, min_col=4, min_row=5, max_row=r-1)
impl_chart.add_data(vals); impl_chart.set_categories(cats)
impl_chart.series[0].title = SeriesLabel(v="Implementation")
impl_chart.width = 16; impl_chart.height = 12
ws_i.add_chart(impl_chart, f"A{r+3}")
ws_i.freeze_panes = "B5"
print(f"Implementation: IMPL_TOT={IMPL_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 6: EQUIPMENT
# ═══════════════════════════════════════════════════════════════
ws_e = wb.create_sheet("Equipment"); ws_e.sheet_properties.tabColor = "4472C4"; scols(ws_e)
wl(ws_e, 1, 1, "Equipment Investment — One-Time", f=ttl)
wl(ws_e, 2, 1, "Capital equipment purchases (Year 1 only)", f=note_f)
wh(ws_e, 4, ["Cost Item"] + SCENARIOS)

equip_items = [
    ("Service Vehicles (N$250,000 each)", [250000*v for v in VH]),
    ("Test Equipment Kits (N$85,000 each)", [85000*k for k in TK]),
    ("Workstations (N$15,000 each)", [15000*w for w in WK]),
    ("Network Monitoring (N$120,000 each)", [120000*n for n in NT]),
    ("Office Furniture", [119000, 165000, 280000, 510000, 1200000]),
    ("Server/Cloud Infrastructure", [145000, 175000, 250000, 400000, 850000]),
]
r = 5
for nm, vals in equip_items:
    wl(ws_e, r, 1, nm)
    for i in range(NS): wi(ws_e, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_e, r, 1, "TOTAL EQUIPMENT", f=b11, fi=grn_fill); fr(ws_e, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_e, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
EQUIP_TOT = r

r += 2
wl(ws_e, r, 1, "EQUIPMENT QUANTITIES", f=b11, fi=lt_fill); fr(ws_e, r, 2, 6, lt_fill)
wh(ws_e, r+1, ["Item"] + SCENARIOS)
for rr, nm, cts in [
    (r+2, "Service Vehicles", VH), (r+3, "Test Equipment Kits", TK),
    (r+4, "Workstations", WK), (r+5, "Network Monitoring Tools", NT),
]:
    wl(ws_e, rr, 1, nm)
    for i in range(NS): wi(ws_e, rr, 2+i, cts[i], NUM)

# Equipment chart
eq_r = rr + 2
wl(ws_e, eq_r, 1, "VISUAL ANALYSIS", f=ttl)
eq_chart = BarChart(); eq_chart.type = "col"; eq_chart.grouping = "stacked"
eq_chart.title = "Equipment Investment by Category & Scenario"
eq_chart.y_axis.numFmt = '#,##0,,"M"'; eq_chart.y_axis.title = "N$ Millions"
colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "FF6384"]
cats = Reference(ws_e, min_col=2, min_row=4, max_col=6, max_row=4)
for idx_row in range(5, EQUIP_TOT):
    vals = Reference(ws_e, min_col=2, min_row=idx_row, max_col=6, max_row=idx_row)
    eq_chart.add_data(vals, from_rows=True)
    eq_chart.series[-1].graphicalProperties.solidFill = colors[(idx_row-5) % len(colors)]
eq_chart.set_categories(cats)
eq_chart.width = 22; eq_chart.height = 14
ws_e.add_chart(eq_chart, f"A{eq_r+1}")
ws_e.freeze_panes = "B5"
print(f"Equipment: EQUIP_TOT={EQUIP_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 7: DEVELOPMENT
# ═══════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("Development"); ws_d.sheet_properties.tabColor = "7030A0"; scols(ws_d)
wl(ws_d, 1, 1, "Development Costs — Pre-Launch", f=ttl)
wl(ws_d, 2, 1, "R&D and product development (one-time investment)", f=note_f)
wh(ws_d, 4, ["Cost Item"] + SCENARIOS)

dev_items = [
    ("Electronics Manufacturing Dev", [312000]*5),
    ("Software R&D", [777747]*5),
    ("Mobile App Development", [300000]*5),
    ("Firmware Development", [2160000]*5),
    ("Enclosure Manufacturing", [1403850]*5),
    ("Marketing/Branding Launch", [325000]*5),
    ("Regulatory & Legal", [1060000]*5),
    ("Office/Logistics Setup", [897400]*5),
    ("Development Contingency (5%)", [615650]*5),
]
r = 5
for nm, vals in dev_items:
    wl(ws_d, r, 1, nm)
    for i in range(NS): wi(ws_d, r, 2+i, vals[i], CURR)
    r += 1
wl(ws_d, r, 1, "TOTAL DEVELOPMENT", f=b11, fi=grn_fill); fr(ws_d, r, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_d, r, 2+i, f"=SUM({col}5:{col}{r-1})", CURR, b11, grn_fill).border = tot_b
DEV_TOT = r

# Development chart
wl(ws_d, r+2, 1, "VISUAL ANALYSIS", f=ttl)
dev_chart = DoughnutChart()
dev_chart.title = "Development Cost Distribution"
cats = Reference(ws_d, min_col=1, min_row=5, max_row=r-1)
vals = Reference(ws_d, min_col=4, min_row=5, max_row=r-1)  # 10K scenario
dev_chart.add_data(vals); dev_chart.set_categories(cats)
dev_chart.series[0].title = SeriesLabel(v="Development")
dev_chart.width = 18; dev_chart.height = 14
ws_d.add_chart(dev_chart, f"A{r+3}")
ws_d.freeze_panes = "B5"
print(f"Development: DEV_TOT={DEV_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 8: FINANCING & CONTINGENCY
# ═══════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("Financing"); ws_f.sheet_properties.tabColor = "C00000"; scols(ws_f)
wl(ws_f, 1, 1, "Financing & Contingency", f=ttl)

wl(ws_f, 3, 1, "FINANCING COSTS", f=b11, fi=org_fill); fr(ws_f, 3, 2, 6, org_fill)
wh(ws_f, 4, ["Cost Item"] + SCENARIOS)
wl(ws_f, 5, 1, "Interest on Working Capital")
for i in range(NS): wi(ws_f, 5, 2+i, 500000, CURR)
wl(ws_f, 6, 1, "Bank Fees & Charges")
for i in range(NS): wi(ws_f, 6, 2+i, 200000, CURR)
wl(ws_f, 7, 1, "TOTAL FINANCING", f=b11, fi=grn_fill); fr(ws_f, 7, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 7, 2+i, f"=SUM({col}5:{col}6)", CURR, b11, grn_fill).border = tot_b
FIN_TOT = 7

wl(ws_f, 9, 1, "CONTINGENCY RESERVES", f=b11, fi=org_fill); fr(ws_f, 9, 2, 6, org_fill)
wh(ws_f, 10, ["Cost Item"] + SCENARIOS)
wl(ws_f, 11, 1, "Operating Contingency (10% of 5Y Operating)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 11, 2+i, f"='Operating Costs'!{col}{OPER_TOT}*0.1", CURR)
wl(ws_f, 12, 1, "Implementation Contingency (10%)")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 12, 2+i, f"=Implementation!{col}{IMPL_TOT}*0.1", CURR)
wl(ws_f, 13, 1, "TOTAL CONTINGENCY", f=b11, fi=grn_fill); fr(ws_f, 13, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_f, 13, 2+i, f"=SUM({col}11:{col}12)", CURR, b11, grn_fill).border = tot_b
CONT_TOT = 13
ws_f.freeze_panes = "B4"
print(f"Financing: FIN_TOT={FIN_TOT}, CONT_TOT={CONT_TOT}")

# ═══════════════════════════════════════════════════════════════
# SHEET 9: PROFITABILITY ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws_pa = wb.create_sheet("Profitability"); ws_pa.sheet_properties.tabColor = "002060"; scols(ws_pa)
wl(ws_pa, 1, 1, "Profitability Analysis — All Scenarios", f=ttl)

wl(ws_pa, 3, 1, "YEAR 1 ANALYSIS", f=b11, fi=lt_fill); fr(ws_pa, 3, 2, 6, lt_fill)
wh(ws_pa, 5, ["Metric"] + SCENARIOS)

wl(ws_pa, 6, 1, "Year 1 Revenue", f=bf, fi=grn_fill); fr(ws_pa, 6, 2, 6, grn_fill)
for i in range(NS): wi(ws_pa, 6, 2+i, round(MC[i] * Y1_PM), CURR)

wl(ws_pa, 7, 1, "YEAR 1 COSTS:", f=bf)
y1_cost_refs = [
    (8, "Manufacturing (one-time)", "Manufacturing", MFG_TOT, 1),
    (9, "Operating (1 year)", "'Operating Costs'", OPER_TOT, 0.2),
    (10, "Implementation (one-time)", "Implementation", IMPL_TOT, 1),
    (11, "Equipment (one-time)", "Equipment", EQUIP_TOT, 1),
    (12, "Development (one-time)", "Development", DEV_TOT, 1),
    (13, "Financing", "Financing", FIN_TOT, 1),
    (14, "Contingency (1 year)", "Financing", CONT_TOT, 0.2),
]
for rr, label, sheet, tot_row, mult in y1_cost_refs:
    wl(ws_pa, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        if mult == 1:
            wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}", CURR)
        else:
            wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}*{mult}", CURR)

wl(ws_pa, 15, 1, "Total Year 1 Costs", f=b11, fi=org_fill); fr(ws_pa, 15, 2, 6, org_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 15, 2+i, f"=SUM({col}8:{col}14)", CURR, b11, org_fill).border = tot_b

wl(ws_pa, 16, 1, "Year 1 Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 16, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 16, 2+i, f"={col}6-{col}15", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 17, 1, "Year 1 Net Margin")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 17, 2+i, f"=IF({col}6>0,{col}16/{col}6,0)", PCT)

wl(ws_pa, 19, 1, "5-YEAR PROJECTIONS", f=b11, fi=lt_fill); fr(ws_pa, 19, 2, 6, lt_fill)
wh(ws_pa, 21, ["Metric"] + SCENARIOS)

wl(ws_pa, 22, 1, "5-Year Total Revenue")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 22, 2+i, f"=Revenue!{col}{REV_TOT}", CURR)

cost_sheet_refs = [
    ("Manufacturing", MFG_TOT), ("'Operating Costs'", OPER_TOT),
    ("Implementation", IMPL_TOT), ("Equipment", EQUIP_TOT),
    ("Development", DEV_TOT), ("Financing", FIN_TOT), ("Financing", CONT_TOT),
]
wl(ws_pa, 23, 1, "5-Year Total Costs")
for i in range(NS):
    col = get_column_letter(2+i)
    refs = "+".join([f"{sn}!{col}{sr}" for sn, sr in cost_sheet_refs])
    wf(ws_pa, 23, 2+i, f"={refs}", CURR)

wl(ws_pa, 24, 1, "5-Year Net Profit", f=b11, fi=grn_fill); fr(ws_pa, 24, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 24, 2+i, f"={col}22-{col}23", CURR, b11, grn_fill).border = tot_b

wl(ws_pa, 25, 1, "5-Year ROI %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 25, 2+i, f"=IF({col}23>0,{col}24/{col}23,0)", PCT)

wl(ws_pa, 26, 1, "5-Year Net Margin %")
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 26, 2+i, f"=IF({col}22>0,{col}24/{col}22,0)", PCT)

wl(ws_pa, 27, 1, "Payback Period (months)")
for i in range(NS): wi(ws_pa, 27, 2+i, CLIENT_PB[i], '0.0')

wl(ws_pa, 28, 1, "DCF / Net Present Value")
for i in range(NS): wi(ws_pa, 28, 2+i, CLIENT_DCF[i], CURR)

wl(ws_pa, 30, 1, "5-YEAR COST BREAKDOWN", f=b11, fi=lt_fill); fr(ws_pa, 30, 2, 6, lt_fill)
wh(ws_pa, 32, ["Cost Category"] + SCENARIOS)
for rr, label, sheet, tot_row in [
    (33, "Manufacturing", "Manufacturing", MFG_TOT),
    (34, "Operating Costs (5Y)", "'Operating Costs'", OPER_TOT),
    (35, "Implementation", "Implementation", IMPL_TOT),
    (36, "Equipment", "Equipment", EQUIP_TOT),
    (37, "Development", "Development", DEV_TOT),
    (38, "Financing", "Financing", FIN_TOT),
    (39, "Contingency", "Financing", CONT_TOT),
]:
    wl(ws_pa, rr, 1, label)
    for i in range(NS):
        col = get_column_letter(2+i)
        wf(ws_pa, rr, 2+i, f"={sheet}!{col}{tot_row}", CURR)

wl(ws_pa, 40, 1, "TOTAL 5-YEAR COSTS", f=b11, fi=grn_fill); fr(ws_pa, 40, 2, 6, grn_fill)
for i in range(NS):
    col = get_column_letter(2+i)
    wf(ws_pa, 40, 2+i, f"=SUM({col}33:{col}39)", CURR, b11, grn_fill).border = tot_b

# Profitability charts
wl(ws_pa, 42, 1, "VISUAL ANALYSIS", f=ttl)
# Bar chart: 5Y Revenue vs Costs vs Profit across scenarios
prof_chart = BarChart(); prof_chart.type = "col"; prof_chart.grouping = "clustered"
prof_chart.title = "5-Year Revenue vs Costs vs Profit by Scenario"
prof_chart.y_axis.numFmt = '#,##0,,,"B"'; prof_chart.y_axis.title = "N$ Billions"
cats = Reference(ws_pa, min_col=2, min_row=21, max_col=6, max_row=21)
for rr, label, color in [(22, "Revenue", "3B82F6"), (23, "Costs", "EF4444"), (24, "Profit", "22C55E")]:
    vals = Reference(ws_pa, min_col=2, min_row=rr, max_col=6, max_row=rr)
    prof_chart.add_data(vals, from_rows=True)
    prof_chart.series[-1].title = SeriesLabel(v=label)
    prof_chart.series[-1].graphicalProperties.solidFill = color
prof_chart.set_categories(cats); prof_chart.legend.position = 'b'
prof_chart.width = 22; prof_chart.height = 14
ws_pa.add_chart(prof_chart, "A43")

ws_pa.freeze_panes = "B5"
print(f"Profitability: complete")

# ═══════════════════════════════════════════════════════════════
# SHEET 10: DASHBOARD (DARK NAVY THEME — COMPREHENSIVE)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "002060"

BG1="0F1B2D"; BG2="162236"; BG3="1D2B42"; BG4="243B55"
ACC_BLUE="3B82F6"; ACC_GREEN="22C55E"; ACC_RED="EF4444"; GOLD="F59E0B"
WHITE="FFFFFF"; LGRAY="94A3B8"; BDR_C="2D3F59"

bg1=PatternFill("solid",fgColor=BG1); bg2=PatternFill("solid",fgColor=BG2)
bg3=PatternFill("solid",fgColor=BG3); bg4=PatternFill("solid",fgColor=BG4)
acc_bl=PatternFill("solid",fgColor=ACC_BLUE); kpi_bg=PatternFill("solid",fgColor="1E3A5F")
tbl_h=PatternFill("solid",fgColor=BG4); tbl1=PatternFill("solid",fgColor=BG2); tbl2=PatternFill("solid",fgColor=BG3)

ft=Font(bold=True,size=16,name="Calibri",color=WHITE)
fs=Font(bold=True,size=11,name="Calibri",color=LGRAY)
fsec=Font(bold=True,size=12,name="Calibri",color=ACC_BLUE)
fw=Font(size=10,name="Calibri",color=WHITE)
fwb=Font(bold=True,size=10,name="Calibri",color=WHITE)
fdr=Font(bold=True,size=12,name="Calibri",color=WHITE)
fkl=Font(size=9,name="Calibri",color=LGRAY)
fg=Font(bold=True,size=10,name="Calibri",color=ACC_GREEN)
frd=Font(size=10,name="Calibri",color=ACC_RED)
fgo=Font(bold=True,size=10,name="Calibri",color=GOLD)
ftn=Font(size=7,name="Calibri",color=BG1)  # hidden
db=Border(left=Side("thin",color=BDR_C),right=Side("thin",color=BDR_C),
          top=Side("thin",color=BDR_C),bottom=Side("thin",color=BDR_C))

# Column widths: A=margin, B-D=left panel, E=gap, F-J=chart area, K=gap, L-O=right charts
for c,w in [('A',1.5),('B',18),('C',14),('D',14),('E',1),('F',12),('G',12),
            ('H',12),('I',12),('J',12),('K',1),('L',12),('M',12),('N',12),('O',12),('P',1.5)]:
    ws.column_dimensions[c].width = w

# Fill entire background dark
for r in range(1, 80):
    ws.row_dimensions[r].height = 18
    for c in range(1, 17): ws.cell(r, c).fill = bg1

# ── ROW 2: Title bar ──
ws.row_dimensions[2].height = 32
ws.merge_cells('B2:O2')
ws['B2'] = "GRIDx Smart Meter  -  5-Year Financial Model"; ws['B2'].font = ft; ws['B2'].fill = bg2
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 16): ws.cell(2, c).fill = bg2

# ── ROW 3: Subtitle + scenario badges ──
ws.row_dimensions[3].height = 24
ws.merge_cells('B3:O3')
ws['B3'] = "Scenarios: 3,000  |  5,000  |  10,000  |  20,000  |  50,000 Meters  |  Currency: N$ (Namibian Dollar)"
ws['B3'].font = Font(size=9, color=LGRAY, name="Calibri")
ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

# ── HIDDEN DATA SECTION (rows 90+) ──
# Scenario labels + MATCH
for i, s in enumerate(SCENARIOS): ws.cell(90, 3+i, s).font = ftn
ws.cell(91, 3).value = '=MATCH($N$6,$C$90:$G$90,0)'; ws.cell(91, 3).font = ftn
IDX = "$C$91"
def idx(rr): return f"INDEX($C${rr}:$G${rr},1,{IDX})"

# Meter counts
for i, m in enumerate(MC): ws.cell(92, 3+i, m).font = ftn
# Mfg unit costs
for i in range(NS): ws.cell(93, 3+i, MFG_U[i]).font = ftn

# Cost totals from sheets (rows 94-100)
sheet_refs = [
    (94, "Manufacturing", MFG_TOT),
    (95, "'Operating Costs'", OPER_TOT),
    (96, "Implementation", IMPL_TOT),
    (97, "Equipment", EQUIP_TOT),
    (98, "Development", DEV_TOT),
    (99, "Financing", FIN_TOT),
    (100, "Financing", CONT_TOT),
]
for dr, sname, srow in sheet_refs:
    for i in range(NS):
        col = get_column_letter(2+i)
        ws.cell(dr, 3+i).value = f"={sname}!{col}{srow}"; ws.cell(dr, 3+i).font = ftn

# Grand total costs (row 101)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(101, 3+i).value = f"=SUM({c}94:{c}100)"; ws.cell(101, 3+i).font = ftn

# 5Y Revenue from Revenue sheet (row 102)
for i in range(NS):
    col = get_column_letter(2+i)
    ws.cell(102, 3+i).value = f"=Revenue!{col}{REV_TOT}"; ws.cell(102, 3+i).font = ftn

# Net Profit (row 103)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(103, 3+i).value = f"={c}102-{c}101"; ws.cell(103, 3+i).font = ftn

# Margin (row 104)
for i in range(NS):
    c = get_column_letter(3+i)
    ws.cell(104, 3+i).value = f"=IF({c}102>0,{c}103/{c}102,0)"; ws.cell(104, 3+i).font = ftn

# Payback + DCF (rows 105-106)
for i in range(NS):
    ws.cell(105, 3+i, CLIENT_PB[i]).font = ftn
    ws.cell(106, 3+i, CLIENT_DCF[i]).font = ftn

# ── CHART DATA: Annual Revenue/Costs/Profit (rows 110-114) ──
ws.cell(110, 2, "Year").font = ftn
for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    ws.cell(110, 3+i, yr).font = ftn

# Revenue per year (millions)
ws.cell(111, 2, "Revenue").font = ftn
ws.cell(111, 3).value = f"={idx(92)}*{Y1_PM}/1000000"; ws.cell(111, 3).font = ftn
for yc in range(4, 8):
    ws.cell(111, yc).value = f"={idx(92)}*{ANN_REC}/1000000"; ws.cell(111, yc).font = ftn

# Total costs per year (millions)
ws.cell(112, 2, "Total Costs").font = ftn
ws.cell(112, 3).value = f"=({idx(94)}+{idx(95)}/5+{idx(96)}+{idx(97)}+{idx(98)}+{idx(99)}+{idx(100)}/5)/1000000"
ws.cell(112, 3).font = ftn
for yc in range(4, 8):
    ws.cell(112, yc).value = f"=({idx(95)}/5+{idx(100)}/5)/1000000"; ws.cell(112, yc).font = ftn

# Net profit per year (millions)
ws.cell(113, 2, "Net Profit").font = ftn
for yc in range(3, 8):
    ws.cell(113, yc).value = f"={get_column_letter(yc)}111-{get_column_letter(yc)}112"
    ws.cell(113, yc).font = ftn

# Cumulative revenue (millions)
ws.cell(114, 2, "Cum Revenue").font = ftn
ws.cell(114, 3).value = "=C111"; ws.cell(114, 3).font = ftn
for yc in range(4, 8):
    ws.cell(114, yc).value = f"={get_column_letter(yc-1)}114+{get_column_letter(yc)}111"
    ws.cell(114, yc).font = ftn

# Cumulative profit (millions)
ws.cell(115, 2, "Cum Profit").font = ftn
ws.cell(115, 3).value = "=C113"; ws.cell(115, 3).font = ftn
for yc in range(4, 8):
    ws.cell(115, yc).value = f"={get_column_letter(yc-1)}115+{get_column_letter(yc)}113"
    ws.cell(115, yc).font = ftn

# ── CHART DATA: Cost breakdown labels + values (rows 120-121) ──
cost_cats = ["Manufacturing", "Operating", "Implementation", "Equipment", "Development", "Financing", "Contingency"]
for i, cat in enumerate(cost_cats):
    ws.cell(120, 3+i, cat).font = ftn
    ws.cell(121, 3+i).value = f"={idx(94+i)}/1000000"; ws.cell(121, 3+i).font = ftn

# ── CHART DATA: Revenue by stream (rows 125-126) ──
rev_streams = ["Hardware", "Installation", "Maintenance", "Wi-Fi", "SMS", "App", "Real Estate"]
rev_per_meter = [HW*5, INST, MAINT*5, WIFI*5, SMS*5, APP*5, RE*5]
for i, (stream, rpm) in enumerate(zip(rev_streams, rev_per_meter)):
    ws.cell(125, 3+i, stream).font = ftn
    ws.cell(126, 3+i).value = f"={idx(92)}*{rpm}/1000000"; ws.cell(126, 3+i).font = ftn

# ── CHART DATA: Scenario comparison (rows 130-133) ──
ws.cell(130, 2, "Scenario").font = ftn
for i, sc in enumerate(SC_SHORT):
    ws.cell(130, 3+i, sc).font = ftn
    ws.cell(131, 3+i).value = f"={get_column_letter(3+i)}102/1000000"; ws.cell(131, 3+i).font = ftn  # Revenue M
    ws.cell(132, 3+i).value = f"={get_column_letter(3+i)}101/1000000"; ws.cell(132, 3+i).font = ftn  # Costs M
    ws.cell(133, 3+i).value = f"={get_column_letter(3+i)}103/1000000"; ws.cell(133, 3+i).font = ftn  # Profit M
ws.cell(131, 2, "Revenue (M)").font = ftn
ws.cell(132, 2, "Costs (M)").font = ftn
ws.cell(133, 2, "Profit (M)").font = ftn

# ── CHART DATA: Personnel headcount by scenario (row 135) ──
ws.cell(135, 2, "Headcount").font = ftn
headcounts = [sum(hc[i] for _, _, hc, _ in POSITIONS) for i in range(NS)]
for i in range(NS): ws.cell(135, 3+i, headcounts[i]).font = ftn

# ── CHART DATA: Operating cost items for selected scenario (rows 140-141) ──
op_labels = ["Personnel", "Data Storage", "Vehicles", "Equipment", "Office", "Software", "Marketing", "Regulatory", "Insurance"]
for i, lbl_text in enumerate(op_labels):
    ws.cell(140, 3+i, lbl_text).font = ftn
# Values from Operating Costs tab (rows 5-13 for 10K = col D)
for i in range(len(op_labels)):
    ws.cell(141, 3+i).value = f"=INDEX('Operating Costs'!$B${5+i}:$F${5+i},1,{IDX})/1000000"
    ws.cell(141, 3+i).font = ftn

# ══════════════════════════════════════════════════════════════
# DASHBOARD VISIBLE LAYOUT
# ══════════════════════════════════════════════════════════════

# ── ROW 5-8: KPI Cards ──
ws.row_dimensions[5].height = 36
ws.row_dimensions[6].height = 16

kpis = [
    (5, 2, 3, "Total Revenue", f'="N$"&TEXT({idx(102)}/1000000,"#,##0")&"M"', ACC_GREEN),
    (5, 6, 7, "Net Profit", f'="N$"&TEXT({idx(103)}/1000000,"#,##0")&"M"', ACC_GREEN),
    (5, 9, 10, "Gross Margin", f'=TEXT({idx(104)}*100,"0.0")&"%"', GOLD),
    (5, 12, 13, "Payback Period", f'=TEXT({idx(105)},"0.0")&" months"', ACC_BLUE),
]
for kr, kc1, kc2, label, formula, acc in kpis:
    for r in range(kr, kr+2):
        for c in range(kc1, kc2+1): ws.cell(r, c).fill = kpi_bg; ws.cell(r, c).border = db
    ws.merge_cells(start_row=kr, start_column=kc1, end_row=kr, end_column=kc2)
    ws.cell(kr, kc1).value = formula
    ws.cell(kr, kc1).font = Font(bold=True, size=18, name="Calibri", color=acc)
    ws.cell(kr, kc1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=kr+1, start_column=kc1, end_row=kr+1, end_column=kc2)
    ws.cell(kr+1, kc1).value = label
    ws.cell(kr+1, kc1).font = fkl
    ws.cell(kr+1, kc1).alignment = Alignment(horizontal="center")

# Dropdown selector (separate KPI-like card)
ws.merge_cells('N5:O5')
ws.cell(5, 14).value = "SELECT SCENARIO"; ws.cell(5, 14).font = Font(bold=True, size=9, color=LGRAY, name="Calibri")
ws.cell(5, 14).alignment = Alignment(horizontal="center", vertical="center")
for c in [14,15]: ws.cell(5, c).fill = kpi_bg; ws.cell(5, c).border = db
ws.merge_cells('N6:O6')
ws['N6'] = "10,000 Meters"; ws['N6'].font = fdr; ws['N6'].fill = acc_bl
ws['N6'].alignment = Alignment(horizontal="center", vertical="center")
for c in [14,15]: ws.cell(6, c).fill = acc_bl; ws.cell(6, c).border = db
# Dropdown is at N6, MATCH at C91 references N6 directly

# ── ROW 8: Section header — Input Parameters ──
ws.row_dimensions[8].height = 22
ws.merge_cells('B8:D8')
ws['B8'] = "INPUT PARAMETERS"; ws['B8'].font = fsec

params = [
    (9, "Meters:", f'=TEXT({idx(92)},"#,##0")', "units"),
    (10, "Hardware:", "N$5,500", "/unit/yr"),
    (11, "Installation:", "N$700", "one-time"),
    (12, "Mfg Cost/Meter:", f'="N$"&TEXT({idx(93)},"#,##0")', "per unit"),
    (13, "Personnel:", "N$748,800", "/year"),
    (14, "Annual Revenue:", f'="N$"&TEXT({idx(92)}*{ANN_REC},"#,##0")', "/year"),
]
for pr, label, val, unit in params:
    ws.cell(pr, 2).value = label; ws.cell(pr, 2).font = fw
    ws.cell(pr, 3).value = val; ws.cell(pr, 3).font = fg
    ws.cell(pr, 4).value = unit; ws.cell(pr, 4).font = Font(size=8, color=LGRAY, name="Calibri")

# ── ROW 16-18: Financial Summary section ──
ws.row_dimensions[16].height = 22
ws.merge_cells('B16:D16')
ws['B16'] = "5-YEAR FINANCIAL SUMMARY"; ws['B16'].font = fsec

summary_items = [
    (17, "Total Revenue:", f'="N$"&TEXT({idx(102)},"#,##0")', fg),
    (18, "Total Costs:", f'="N$"&TEXT({idx(101)},"#,##0")', frd),
    (19, "Net Profit:", f'="N$"&TEXT({idx(103)},"#,##0")', fgo),
    (20, "ROI:", f'=TEXT(IF({idx(101)}>0,{idx(103)}/{idx(101)},0)*100,"#,##0")&"%"', fg),
    (21, "NPV:", f'="N$"&TEXT({idx(106)},"#,##0")', fg),
]
for sr, label, formula, font in summary_items:
    ws.cell(sr, 2).value = label; ws.cell(sr, 2).font = fw
    ws.cell(sr, 3).value = formula; ws.cell(sr, 3).font = font

# ══════════════════════════════════════════════════════════════
# DASHBOARD CHARTS
# ══════════════════════════════════════════════════════════════

# ── CHART 1: Revenue vs Costs & Net Profit (Bar + Line) — rows 8-20 right side ──
chart1 = BarChart(); chart1.type = "col"; chart1.grouping = "clustered"
chart1.title = "Annual Revenue vs Costs & Net Profit Trend"
chart1.y_axis.numFmt = '0.0'; chart1.y_axis.title = "N$ Millions"
cats = Reference(ws, min_col=3, min_row=110, max_col=7, max_row=110)
# Revenue bars
d_rev = Reference(ws, min_col=3, min_row=111, max_col=7, max_row=111)
chart1.add_data(d_rev, from_rows=True)
chart1.series[0].title = SeriesLabel(v="Revenue"); chart1.series[0].graphicalProperties.solidFill = "3B82F6"
# Cost bars
d_cost = Reference(ws, min_col=3, min_row=112, max_col=7, max_row=112)
chart1.add_data(d_cost, from_rows=True)
chart1.series[1].title = SeriesLabel(v="Costs"); chart1.series[1].graphicalProperties.solidFill = "EF4444"
chart1.set_categories(cats); chart1.legend.position = 'b'
# Add profit as line overlay
line1 = LineChart()
d_prof = Reference(ws, min_col=3, min_row=113, max_col=7, max_row=113)
line1.add_data(d_prof, from_rows=True)
line1.series[0].title = SeriesLabel(v="Net Profit")
line1.series[0].graphicalProperties.line.solidFill = "22C55E"
line1.series[0].graphicalProperties.line.width = 28000
line1.y_axis.numFmt = '0.0'
chart1.y_axis.crosses = "min"
chart1 += line1
chart1.width = 18; chart1.height = 12
ws.add_chart(chart1, "F8")

# ── CHART 2: Cumulative Cash Flow — rows 8-20 far right ──
chart2 = BarChart(); chart2.type = "col"; chart2.grouping = "clustered"
chart2.title = "Cumulative Cash Flow (5-Year)"
chart2.y_axis.numFmt = '0'; chart2.y_axis.title = "N$ Millions"
cats2 = Reference(ws, min_col=3, min_row=110, max_col=7, max_row=110)
d_cum_rev = Reference(ws, min_col=3, min_row=114, max_col=7, max_row=114)
chart2.add_data(d_cum_rev, from_rows=True)
chart2.series[0].title = SeriesLabel(v="Cum Revenue"); chart2.series[0].graphicalProperties.solidFill = "3B82F6"
d_cum_prof = Reference(ws, min_col=3, min_row=115, max_col=7, max_row=115)
chart2.add_data(d_cum_prof, from_rows=True)
chart2.series[1].title = SeriesLabel(v="Cum Profit"); chart2.series[1].graphicalProperties.solidFill = "22C55E"
chart2.set_categories(cats2); chart2.legend.position = 'b'
chart2.width = 14; chart2.height = 12
ws.add_chart(chart2, "L21")

# ── CHART 3: 5-Year Cost Breakdown (Doughnut) — rows 22-34 left side ──
ws.row_dimensions[22].height = 22
ws.merge_cells('B22:D22')
ws['B22'] = "COST ANALYSIS"; ws['B22'].font = fsec

chart3 = DoughnutChart()
chart3.title = "5-Year Cost Breakdown"
cats3 = Reference(ws, min_col=3, min_row=120, max_col=9, max_row=120)
vals3 = Reference(ws, min_col=3, min_row=121, max_col=9, max_row=121)
chart3.add_data(vals3, from_rows=True)
chart3.set_categories(cats3)
chart3.series[0].title = SeriesLabel(v="Costs")
chart3.width = 14; chart3.height = 12
ws.add_chart(chart3, "B23")

# ── CHART 4: Revenue by Stream (Doughnut) — rows 22-34 center ──
chart4 = DoughnutChart()
chart4.title = "Revenue by Stream"
cats4 = Reference(ws, min_col=3, min_row=125, max_col=9, max_row=125)
vals4 = Reference(ws, min_col=3, min_row=126, max_col=9, max_row=126)
chart4.add_data(vals4, from_rows=True)
chart4.set_categories(cats4)
chart4.series[0].title = SeriesLabel(v="Revenue")
chart4.width = 14; chart4.height = 12
ws.add_chart(chart4, "F23")

# ── CHART 5: Operating Cost Breakdown (Horizontal Bar) — rows 22-34 right ──
chart5 = BarChart(); chart5.type = "bar"; chart5.grouping = "clustered"
chart5.title = "Operating Costs (5Y, Selected Scenario)"
chart5.x_axis.numFmt = '0.0'; chart5.x_axis.title = "N$ Millions"
cats5 = Reference(ws, min_col=3, min_row=140, max_col=11, max_row=140)
vals5 = Reference(ws, min_col=3, min_row=141, max_col=11, max_row=141)
chart5.add_data(vals5, from_rows=True); chart5.set_categories(cats5)
chart5.series[0].title = SeriesLabel(v="Operating"); chart5.series[0].graphicalProperties.solidFill = "ED7D31"
chart5.width = 14; chart5.height = 12; chart5.legend = None
ws.add_chart(chart5, "L8")

# ── ROW 36: SCENARIO COMPARISON section ──
ws.row_dimensions[36].height = 22
ws.merge_cells('B36:O36')
ws['B36'] = "SCENARIO COMPARISON — All Deployment Sizes"; ws['B36'].font = fsec; ws['B36'].fill = bg1

r = 37
ws.merge_cells(f'B{r}:D{r}')
ws.cell(r, 2).value = "Metric"; ws.cell(r, 2).font = fwb; ws.cell(r, 2).fill = tbl_h
for c in [2,3,4]: ws.cell(r, c).fill = tbl_h; ws.cell(r, c).border = db
sc_cols = [6,8,10,12,14]
for i, label in enumerate(SC_SHORT):
    ws.cell(r, sc_cols[i]).value = label; ws.cell(r, sc_cols[i]).font = fwb
    ws.cell(r, sc_cols[i]).fill = tbl_h; ws.cell(r, sc_cols[i]).border = db
    ws.cell(r, sc_cols[i]).alignment = Alignment(horizontal="center")

for ri, (metric, row_src, fmt, positive) in enumerate([
    ("5-Year Revenue", 102, '"N$"#,##0.0,,"M"', True),
    ("5-Year Costs", 101, '"N$"#,##0.0,,"M"', False),
    ("Net Profit", 103, '"N$"#,##0.0,,"M"', True),
    ("Gross Margin", 104, '0.0%', True),
    ("Payback", 105, '0.0" mo"', True),
    ("ROI", None, '0.0%', True),
]):
    r = 38 + ri
    row_fill = tbl1 if ri % 2 == 0 else tbl2
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(r, 2).value = metric; ws.cell(r, 2).font = fw
    for c in [2,3,4]: ws.cell(r, c).fill = row_fill; ws.cell(r, c).border = db
    for i, sc in enumerate(sc_cols):
        c_letter = get_column_letter(3+i)
        if metric == "ROI":
            ws.cell(r, sc).value = f"=IF({c_letter}101>0,{c_letter}103/{c_letter}101,0)"
        else:
            ws.cell(r, sc).value = f"={c_letter}{row_src}"
        ws.cell(r, sc).font = fg if positive else frd
        ws.cell(r, sc).number_format = fmt; ws.cell(r, sc).fill = row_fill; ws.cell(r, sc).border = db
        ws.cell(r, sc).alignment = Alignment(horizontal="center")
        if i == 2:  # highlight 10K
            ws.cell(r, sc).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.cell(r, sc).font = Font(bold=True, size=10, name="Calibri", color=ACC_GREEN if positive else ACC_RED)

# ── CHART 6: Scenario Comparison Bar Chart — below table ──
ws.row_dimensions[45].height = 22
ws.merge_cells('B45:O45')
ws['B45'] = "SCENARIO ANALYSIS CHARTS"; ws['B45'].font = fsec

chart6 = BarChart(); chart6.type = "col"; chart6.grouping = "clustered"
chart6.title = "5-Year Revenue, Costs & Profit by Scenario"
chart6.y_axis.numFmt = '0'; chart6.y_axis.title = "N$ Millions"
cats6 = Reference(ws, min_col=3, min_row=130, max_col=7, max_row=130)
for rr, label, color in [(131, "Revenue", "3B82F6"), (132, "Costs", "EF4444"), (133, "Profit", "22C55E")]:
    vals = Reference(ws, min_col=3, min_row=rr, max_col=7, max_row=rr)
    chart6.add_data(vals, from_rows=True)
    chart6.series[-1].title = SeriesLabel(v=label)
    chart6.series[-1].graphicalProperties.solidFill = color
chart6.set_categories(cats6); chart6.legend.position = 'b'
chart6.width = 18; chart6.height = 12
ws.add_chart(chart6, "B46")

# ── CHART 7: Personnel Headcount by Scenario ──
chart7 = BarChart(); chart7.type = "col"; chart7.grouping = "clustered"
chart7.title = "Personnel Headcount by Scenario"
chart7.y_axis.title = "Employees"
cats7 = Reference(ws, min_col=3, min_row=130, max_col=7, max_row=130)
vals7 = Reference(ws, min_col=3, min_row=135, max_col=7, max_row=135)
chart7.add_data(vals7, from_rows=True); chart7.set_categories(cats7)
chart7.series[0].title = SeriesLabel(v="Headcount")
chart7.series[0].graphicalProperties.solidFill = "70AD47"
chart7.width = 14; chart7.height = 12; chart7.legend = None
ws.add_chart(chart7, "L46")

# ── ROW 62: Footer ──
ws.merge_cells('B62:O62')
ws['B62'] = "GRIDx Smart Energy Solutions  |  Confidential Financial Model  |  All figures in Namibian Dollars (N$)"
ws['B62'].font = Font(italic=True, size=8, color=LGRAY, name="Calibri")
ws['B62'].alignment = Alignment(horizontal="center")

# Data validation for dropdown
dv = DataValidation(type="list", formula1="=$C$90:$G$90", allow_blank=False)
dv.prompt = "Select scenario"; dv.promptTitle = "Rollout"
ws.add_data_validation(dv); dv.add(ws['N6'])

# ── Finalize ──
wb.move_sheet("Dashboard", offset=-9)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_view.showGridLines = False

OUT = "/var/lib/freelancer/projects/40278037/GRIDx_Financial_Dashboard_v10.xlsx"
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print("v10: Comprehensive dashboard with charts from all tabs")
